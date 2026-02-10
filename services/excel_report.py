import logging
from datetime import datetime, timedelta, date, time

import config
import database.db_api as db

logger = logging.getLogger(__name__)


_UA_MONTHS = {
    1: "СІЧЕНЬ",
    2: "ЛЮТИЙ",
    3: "БЕРЕЗЕНЬ",
    4: "КВІТЕНЬ",
    5: "ТРАВЕНЬ",
    6: "ЧЕРВЕНЬ",
    7: "ЛИПЕНЬ",
    8: "СЕРПЕНЬ",
    9: "ВЕРЕСЕНЬ",
    10: "ЖОВТЕНЬ",
    11: "ЛИСТОПАД",
    12: "ГРУДЕНЬ",
}

# 4 intervals for the template
_SHIFT_ORDER = ["m", "d", "e", "x"]


def _month_range_kyiv(period: str) -> tuple[date, date, str]:
    """Return (start_date, end_date, label) for current/prev month in Kyiv tz."""
    now = datetime.now(config.KYIV)

    if period == "current":
        start = now.replace(day=1).date()
        if start.month == 12:
            next_m = date(start.year + 1, 1, 1)
        else:
            next_m = date(start.year, start.month + 1, 1)
        end = next_m - timedelta(days=1)
        label = _UA_MONTHS.get(start.month, str(start.month))
        return start, end, label

    first_day_current = now.replace(day=1).date()
    end = first_day_current - timedelta(days=1)
    start = end.replace(day=1)
    label = _UA_MONTHS.get(start.month, str(start.month))
    return start, end, label


def _parse_ts(ts: str) -> datetime | None:
    try:
        return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=config.KYIV)
    except Exception:
        return None


def _dt_to_excel_time(dt: datetime | None) -> time | None:
    """Return datetime.time to store as real Excel time (not text)."""
    return dt.timetz().replace(tzinfo=None) if dt else None


def _date_key(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def _iso(d: date) -> str:
    return d.strftime("%Y-%m-%d")


async def generate_report(period: str):
    """Generate Excel report from DB using template-like layout (4 intervals).

    Keeps columns: НОМЕР ЧЕКА, ПАЛИВО ПРЕВІЗ.
    """
    try:
        start_d, end_d, month_label = _month_range_kyiv(period)
        start_date, end_date = start_d.isoformat(), end_d.isoformat()
        logs = db.get_logs_for_period(start_date, end_date)

        try:
            fuel_rate = float(getattr(config, "FUEL_CONSUMPTION", 5.3) or 5.3)
            if fuel_rate <= 0:
                fuel_rate = 5.3
        except Exception:
            fuel_rate = 5.3

        # Collect shifts + refills
        days: dict[str, dict] = {}
        refills_by_day: dict[str, list] = {}

        for (event_type, ts, user_name, value, driver_name, receipt_number) in logs:
            dt = _parse_ts(ts)
            if not dt:
                continue

            d = dt.date()
            day_key = _date_key(d)
            day_iso = _iso(d)

            if day_key not in days:
                days[day_key] = {
                    "shifts": {c: {"start": None, "end": None} for c in _SHIFT_ORDER},
                }

            if event_type in ("m_start", "d_start", "e_start", "x_start"):
                c = event_type.split("_", 1)[0]
                days[day_key]["shifts"][c]["start"] = dt
                continue

            if event_type in ("m_end", "d_end", "e_end", "x_end"):
                c = event_type.split("_", 1)[0]
                days[day_key]["shifts"][c]["end"] = dt
                continue

            if event_type == "refill":
                try:
                    liters = float(str(value or "0").replace(",", "."))
                except Exception:
                    liters = 0.0
                refills_by_day.setdefault(day_iso, []).append(
                    {
                        "liters": liters,
                        "receipt": receipt_number or "",
                        "carrier": driver_name or "",  # stored in driver_name field in DB
                    }
                )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = month_label

        # Two-row header like template (simplified, with merges)
        ws["A1"].value = "ДАТА"
        ws.merge_cells("A1:A2")

        # Groups 1..4 with start/end under each
        groups = [("1", "B", "C"), ("2", "D", "E"), ("3", "F", "G"), ("4", "H", "I")]
        for g, c1, c2 in groups:
            ws[f"{c1}1"].value = int(g)
            ws.merge_cells(f"{c1}1:{c2}1")
            ws[f"{c1}2"].value = "ПОЧАТОК, Г"
            ws[f"{c2}2"].value = "КІНЕЦЬ, Г"

        ws["J1"].value = "РОЗХІД"
        ws.merge_cells("J1:J2")

        ws["K1"].value = fuel_rate
        ws.merge_cells("K1:K2")

        ws["L1"].value = "ВИТРАТИ ПАЛИВА"
        ws.merge_cells("L1:L2")

        # Second row (actual column headers for the rest)
        ws["J2"].value = "ВСЬОГО ГОДИН"
        ws["K2"].value = "ЗАЛИШОК ПАЛИВА НА РАНОК"
        ws["L2"].value = "ВИТРАТИ ПАЛИВА"
        ws["M2"].value = "ЗАЛИШОК"
        ws["N2"].value = "ПРИВЕЗЕНО ПАЛИВА"
        ws["O2"].value = "ЗАЛИШОК ПАЛИВА ВЕЧІР"
        ws["P2"].value = "НОМЕР ЧЕКА"
        ws["Q2"].value = "ПАЛИВО ПРЕВІЗ"

        # Style header
        fill = PatternFill("solid", fgColor="1F4E79")
        for r in (1, 2):
            for cell in ws[r]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column indices for formulas
        COL_DATE = 1
        COL_S1_START, COL_S1_END = 2, 3
        COL_S2_START, COL_S2_END = 4, 5
        COL_S3_START, COL_S3_END = 6, 7
        COL_S4_START, COL_S4_END = 8, 9
        COL_TOTAL_HOURS = 10  # J
        COL_FUEL_MORNING = 11  # K
        COL_FUEL_SPENT = 12  # L
        COL_FUEL_LEFT = 13  # M
        COL_REFILL = 14  # N
        COL_FUEL_EVENING = 15  # O
        COL_RECEIPT = 16  # P
        COL_CARRIER = 17  # Q

        first_data_row = 3

        # Generate full month grid
        cur = start_d
        row = first_data_row
        while cur <= end_d:
            key = _date_key(cur)
            shifts = (days.get(key) or {}).get("shifts") or {c: {"start": None, "end": None} for c in _SHIFT_ORDER}

            # Map shifts to 4 intervals by code order (m,d,e,x)
            starts = [_dt_to_excel_time(shifts[c]["start"]) for c in _SHIFT_ORDER]
            ends = [_dt_to_excel_time(shifts[c]["end"]) for c in _SHIFT_ORDER]

            # Refills: sum liters, join receipt/carrier
            refills = refills_by_day.get(_iso(cur), [])
            refill_liters = sum(r.get("liters", 0.0) for r in refills)
            receipt = ", ".join([r.get("receipt", "") for r in refills if r.get("receipt")])
            carrier = ", ".join([r.get("carrier", "") for r in refills if r.get("carrier")])

            ws.cell(row=row, column=COL_DATE).value = cur
            ws.cell(row=row, column=COL_DATE).number_format = "DD.MM.YYYY"

            for i, (s, e) in enumerate(zip(starts, ends), start=0):
                s_col = COL_S1_START + i * 2
                e_col = COL_S1_END + i * 2
                ws.cell(row=row, column=s_col).value = s
                ws.cell(row=row, column=e_col).value = e
                ws.cell(row=row, column=s_col).number_format = "HH:MM"
                ws.cell(row=row, column=e_col).number_format = "HH:MM"

            ws.cell(row=row, column=COL_REFILL).value = round(refill_liters, 1) if refill_liters else ""
            ws.cell(row=row, column=COL_RECEIPT).value = receipt
            ws.cell(row=row, column=COL_CARRIER).value = carrier

            # Helpers
            def c(col: int, rr: int = row) -> str:
                return f"{get_column_letter(col)}{rr}"

            # Total hours: sum of 4 intervals
            parts = []
            for (s_col, e_col) in [
                (COL_S1_START, COL_S1_END),
                (COL_S2_START, COL_S2_END),
                (COL_S3_START, COL_S3_END),
                (COL_S4_START, COL_S4_END),
            ]:
                parts.append(f"IF(OR({c(s_col)}=\"\",{c(e_col)}=\"\"),0,{c(e_col)}-{c(s_col)})")
            ws.cell(row=row, column=COL_TOTAL_HOURS).value = "=" + "+".join(parts)
            ws.cell(row=row, column=COL_TOTAL_HOURS).number_format = "[h]:mm:ss"

            # Fuel spent = total_hours*24 * fuel_rate (fuel rate in K1)
            ws.cell(row=row, column=COL_FUEL_SPENT).value = f"={c(COL_TOTAL_HOURS)}*24*$K$1"

            # Morning fuel:
            # - First day of period: leave blank for manual input (as you requested)
            # - Others: prev evening
            if row == first_data_row:
                ws.cell(row=row, column=COL_FUEL_MORNING).value = ""
            else:
                ws.cell(row=row, column=COL_FUEL_MORNING).value = f"={c(COL_FUEL_EVENING, row-1)}"

            # Fuel left = morning - spent
            ws.cell(row=row, column=COL_FUEL_LEFT).value = f"=IF({c(COL_FUEL_MORNING)}=\"\",\"\",{c(COL_FUEL_MORNING)}-{c(COL_FUEL_SPENT)})"

            # Evening = left + refill
            ws.cell(row=row, column=COL_FUEL_EVENING).value = f"=IF({c(COL_FUEL_LEFT)}=\"\",\"\",{c(COL_FUEL_LEFT)}+{c(COL_REFILL)})"

            row += 1
            cur += timedelta(days=1)

        # Autosize (limit)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 32)

        # Events sheet
        ws2 = wb.create_sheet("ПОДІЇ")
        ws2.append(["ID", "Дата/час", "Тип події", "Користувач", "Літри", "Чек", "Водій", "Значення"])
        for (event_type, ts, user_name, value, driver_name, receipt_number) in logs:
            liters_col = value if event_type == "refill" else ""
            ws2.append(["", ts, event_type, user_name, liters_col, receipt_number, driver_name, value])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        ts = datetime.now(config.KYIV).strftime("%Y%m%d_%H%M%S")
        filename = f"report_{period}_{month_label}_{ts}.xlsx"
        wb.save(filename)

        caption = (
            f"📊 <b>Звіт з БД</b>\n"
            f"🗓 Період: <b>{start_date}</b> — <b>{end_date}</b>\n"
            f"📁 Файл: <code>{filename}</code>\n"
            f"ℹ️ На перший день періоду залишок палива на ранок заповнюється вручну."
        )
        return filename, caption

    except Exception as e:
        logger.error(f"❌ Помилка генерації звіту: {e}", exc_info=True)
        return None, f"❌ Помилка генерації звіту: {str(e)}"
