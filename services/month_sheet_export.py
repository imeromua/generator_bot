import logging
import os
from datetime import datetime

import aiohttp
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

import config

logger = logging.getLogger(__name__)


def _build_creds() -> Credentials:
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    return Credentials.from_service_account_file("service_account.json", scopes=scopes)


async def _export_spreadsheet_xlsx(file_id: str, out_path: str, creds: Credentials) -> None:
    creds.refresh(GoogleRequest())

    url = f"https://www.googleapis.com/drive/v3/files/{file_id}/export"
    params = {
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    headers = {
        "Authorization": f"Bearer {creds.token}",
    }

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, headers=headers, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Drive export failed: status={resp.status}, body={text[:500]}")

            data = await resp.read()
            with open(out_path, "wb") as f:
                f.write(data)


async def generate_month_sheet_xlsx(sheet_name: str | None = None):
    """Export monthly sheet (e.g. 'ЛЮТИЙ') to XLSX preserving original styles."""

    if not config.SHEET_ID:
        return None, "❌ SHEET_ID не знайдено"

    if not os.path.exists("service_account.json"):
        return None, "❌ Файл service_account.json не знайдено"

    title = sheet_name or config.SHEET_NAME

    ts = datetime.now(config.KYIV).strftime("%Y%m%d_%H%M%S")
    exported = f"_source_export_{ts}.xlsx"
    out_file = f"month_{title}_{ts}.xlsx"

    try:
        creds = _build_creds()
        await _export_spreadsheet_xlsx(config.SHEET_ID, exported, creds)

        wb = load_workbook(exported)
        if title not in wb.sheetnames:
            # fallback to active
            title = wb.active.title

        # Keep only chosen sheet
        for name in list(wb.sheetnames):
            if name != title:
                ws = wb[name]
                wb.remove(ws)

        wb.save(out_file)

        caption = (
            f"📤 <b>Місячний лист (Excel)</b>\n"
            f"Джерело: <b>{title}</b>"
        )

        return out_file, caption

    except Exception as e:
        logger.error(f"❌ Month XLSX export error: {e}", exc_info=True)
        return None, f"❌ Помилка експорту: {str(e)}"

    finally:
        try:
            if os.path.exists(exported):
                os.remove(exported)
        except Exception:
            pass
