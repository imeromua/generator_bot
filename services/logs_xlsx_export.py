import logging
import os
from copy import copy
from datetime import datetime

import aiohttp
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.service_account import Credentials
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

import config

logger = logging.getLogger(__name__)


_REQUIRED_COLS = ["timestamp", "event_type", "user_name", "liters", "receipt", "driver", "value_raw"]


def _normalize(val) -> str:
    if val is None:
        return ""
    return str(val).strip().lower()


def _build_creds() -> Credentials:
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    return Credentials.from_service_account_file("service_account.json", scopes=scopes)


async def _export_spreadsheet_xlsx(file_id: str, out_path: str, creds: Credentials) -> None:
    creds.refresh(GoogleRequest())

    url = f"https://www.googleapis.com/drive/v3/files/{file_id}/export"
    params = {
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    headers = {
        "Authorization": f"Bearer {creds.token}",
    }

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, headers=headers, timeout=120) as resp:
            if resp.status != 200:
                text = await resp.text()
                raise RuntimeError(f"Drive export failed: status={resp.status}, body={text[:500]}")

            data = await resp.read()
            with open(out_path, "wb") as f:
                f.write(data)


def _find_header_row(ws, scan_rows: int = 50) -> int | None:
    """Try to locate a header row by looking for known column names."""

    expected = {
        "timestamp": {"timestamp", "дата", "дата/час", "дата та час", "час", "time"},
        "event_type": {"event_type", "тип", "тип події", "подія", "event"},
        "user_name": {"user_name", "користувач", "працівник", "оператор", "хто", "user"},
    }

    max_r = min(ws.max_row or 1, scan_rows)
    for r in range(1, max_r + 1):
        row_vals = [_normalize(c.value) for c in ws[r]]
        if not any(row_vals):
            continue

        hits = 0
        for variants in expected.values():
            if any(v in variants for v in row_vals):
                hits += 1

        if hits >= 2:
            return r

    return None


def _map_columns_by_header(ws, header_row: int) -> dict:
    """Return mapping output_col -> source_col_index (1-based)."""

    header_cells = ws[header_row]
    header_vals = [_normalize(c.value) for c in header_cells]

    synonyms = {
        "timestamp": {"timestamp", "дата", "дата/час", "дата та час", "час", "time"},
        "event_type": {"event_type", "тип", "тип події", "подія", "event"},
        "user_name": {"user_name", "користувач", "працівник", "оператор", "хто", "user"},
        "liters": {"liters", "літри", "літрів", "л", "l", "об'єм", "обʼєм", "обєм"},
        "receipt": {"receipt", "чек", "квитанція", "№ чека", "номер чека"},
        "driver": {"driver", "водій", "driver_name"},
        "value_raw": {"value_raw", "value", "значення"},
        "log_id": {"log_id", "id", "#"},
    }

    mapping = {}
    for out_col, variants in synonyms.items():
        for idx, hv in enumerate(header_vals, start=1):
            if hv in variants:
                mapping[out_col] = idx
                break

    # Fallback: assume events sheet format: (id), ts, type, user, liters, receipt, driver, value
    if all(k not in mapping for k in _REQUIRED_COLS):
        start_idx = 2 if (header_vals and header_vals[0] in {"log_id", "id", "#"}) else 1
        mapping = {
            "timestamp": start_idx,
            "event_type": start_idx + 1,
            "user_name": start_idx + 2,
            "liters": start_idx + 3,
            "receipt": start_idx + 4,
            "driver": start_idx + 5,
            "value_raw": start_idx + 6,
        }

    return mapping


def _cell_to_date(val):
    if val is None or val == "":
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, (int, float)):
        try:
            dt = from_excel(val, offset=0)
            if isinstance(dt, datetime):
                return dt.date()
        except Exception:
            return None

    s = str(val).strip()
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            continue

    return None


def _cell_to_datetime(val) -> datetime | None:
    if val is None or val == "":
        return None

    if isinstance(val, datetime):
        return val

    if isinstance(val, (int, float)):
        try:
            dt = from_excel(val, offset=0)
            if isinstance(dt, datetime):
                return dt
        except Exception:
            return None

    s = str(val).strip()
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]

    for f in fmts:
        try:
            return datetime.strptime(s, f)
        except Exception:
            continue

    return None


def _copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def _copy_column_dimensions(src_ws, dst_ws, src_col_idx: int, dst_col_idx: int):
    src_letter = src_ws.cell(row=1, column=src_col_idx).column_letter
    dst_letter = dst_ws.cell(row=1, column=dst_col_idx).column_letter

    src_dim = src_ws.column_dimensions.get(src_letter)
    if src_dim:
        dst_dim = dst_ws.column_dimensions[dst_letter]
        dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel


def _pretty_shift(val: str) -> str:
    v = (val or "").strip()
    if not v:
        return ""

    key = v.lower()
    if key in {"m", "1", "зміна 1", "зміна1"}:
        return "🟦 Зміна 1"
    if key in {"d", "2", "зміна 2", "зміна2"}:
        return "🟩 Зміна 2"
    if key in {"e", "3", "зміна 3", "зміна3"}:
        return "🟪 Зміна 3"
    if key in {"x", "екстра", "extra"}:
        return "⚡ Екстра"

    return v


def _format_event_line(ts, event_type, user_name, liters, receipt, driver, value_raw) -> str:
    dt = _cell_to_datetime(ts)
    dt_part = dt.strftime("%d.%m %H:%M") if dt else ""

    et = _normalize(event_type)
    user = ("" if user_name is None else str(user_name).strip())

    def suffix_user() -> str:
        return f" ({user})" if user else ""

    # Fuel/refill
    if any(k in et for k in {"палив", "refill", "fuel", "прийом"}):
        l = "" if liters in (None, "") else str(liters).strip()
        chk = "" if receipt in (None, "") else str(receipt).strip()
        drv = "" if driver in (None, "") else str(driver).strip()

        parts = []
        if l:
            parts.append(f"{l} л")
        if chk:
            parts.append(f"чек {chk}")
        if drv:
            parts.append(f"водій {drv}")

        body = ": ".join([
            "⛽ Прийом палива",
            ", ".join(parts) if parts else "",
        ]).rstrip(": ")
        return f"• {dt_part} — {body}{suffix_user()}".strip()

    # Start/stop
    if any(k in et for k in {"старт", "start", "почат"}):
        shift = _pretty_shift("" if value_raw is None else str(value_raw))
        return f"• {dt_part} — ▶️ Старт: {shift}{suffix_user()}".strip()

    if any(k in et for k in {"стоп", "stop", "кінец", "конец", "end"}):
        shift = _pretty_shift("" if value_raw is None else str(value_raw))
        return f"• {dt_part} — ⏹ Стоп: {shift}{suffix_user()}".strip()

    # Fallback
    label = "Тип події" if not event_type else str(event_type).strip()
    val = ("" if value_raw is None else str(value_raw).strip())
    tail = f": {val}" if val else ""
    return f"• {dt_part} — {label}{tail}{suffix_user()}".strip()


async def generate_logs_report_xlsx(
    start_date: str,
    end_date: str,
    sheet_name: str | None = None,
):
    """Generate XLSX events report from sheet (default: 'ПОДІЇ') for a given date range."""

    if not config.SHEET_ID:
        return None, "❌ SHEET_ID не знайдено"

    if not os.path.exists("service_account.json"):
        return None, "❌ Файл service_account.json не знайдено"

    try:
        start_d = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
    except Exception:
        return None, "❌ Некоректний формат дати (очікується YYYY-MM-DD)"

    period_title = sheet_name or "ПОДІЇ"

    ts = datetime.now(config.KYIV).strftime("%Y%m%d_%H%M%S")
    exported = f"_source_export_{ts}.xlsx"
    out_file = f"logs_{start_date}_to_{end_date}_{ts}.xlsx"

    try:
        creds = _build_creds()
        await _export_spreadsheet_xlsx(config.SHEET_ID, exported, creds)

        src_wb = load_workbook(exported)
        if period_title in src_wb.sheetnames:
            src_ws = src_wb[period_title]
        else:
            src_ws = src_wb.worksheets[0]
            period_title = src_ws.title

        header_row = _find_header_row(src_ws) or 1
        col_map = _map_columns_by_header(src_ws, header_row)

        ts_col = col_map.get("timestamp", 1)
        data_rows = []
        for r in range(header_row + 1, (src_ws.max_row or 1) + 1):
            dt = _cell_to_date(src_ws.cell(row=r, column=ts_col).value)
            if not dt:
                continue
            if start_d <= dt <= end_d:
                data_rows.append(r)

        if not data_rows:
            return None, "ℹ️ Немає подій за цей період"

        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        out_ws = out_wb.create_sheet(title="ПОДІЇ")

        for dst_i, key in enumerate(_REQUIRED_COLS, start=1):
            src_i = col_map.get(key)
            if src_i:
                _copy_column_dimensions(src_ws, out_ws, src_i, dst_i)

        for dst_i, key in enumerate(_REQUIRED_COLS, start=1):
            src_i = col_map.get(key)
            if not src_i:
                continue
            _copy_cell(src_ws.cell(row=header_row, column=src_i), out_ws.cell(row=1, column=dst_i))

        if src_ws.row_dimensions.get(header_row):
            out_ws.row_dimensions[1].height = src_ws.row_dimensions[header_row].height

        out_r = 2
        preview_lines: list[str] = []

        for src_r in data_rows:
            if src_ws.row_dimensions.get(src_r):
                out_ws.row_dimensions[out_r].height = src_ws.row_dimensions[src_r].height

            row_vals = {}
            for dst_i, key in enumerate(_REQUIRED_COLS, start=1):
                src_i = col_map.get(key)
                if not src_i:
                    continue

                src_cell = src_ws.cell(row=src_r, column=src_i)
                dst_cell = out_ws.cell(row=out_r, column=dst_i)
                _copy_cell(src_cell, dst_cell)
                row_vals[key] = src_cell.value

            preview_lines.append(
                _format_event_line(
                    row_vals.get("timestamp"),
                    row_vals.get("event_type"),
                    row_vals.get("user_name"),
                    row_vals.get("liters"),
                    row_vals.get("receipt"),
                    row_vals.get("driver"),
                    row_vals.get("value_raw"),
                )
            )

            out_r += 1

        out_wb.save(out_file)

        # Telegram caption is limited, so cap preview
        preview_lines = [ln for ln in preview_lines if ln]
        preview_tail = preview_lines[-15:]

        start_ui = datetime.strptime(start_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        end_ui = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d.%m.%Y")

        caption = (
            f"📤 <b>Експорт подій (Excel)</b>\n"
            f"Період: <b>{start_ui}</b> — <b>{end_ui}</b>\n"
            f"Джерело: <b>{period_title}</b>\n"
            f"Подій: <b>{len(data_rows)}</b>\n\n"
            f"🕘 <b>Останні події ({len(preview_tail)})</b>\n"
            + "\n".join(preview_tail)
        )

        # hard safety
        if len(caption) > 950:
            caption = caption[:940] + "…"

        return out_file, caption

    except Exception as e:
        logger.error(f"❌ XLSX export error: {e}", exc_info=True)
        return None, f"❌ Помилка експорту: {str(e)}"

    finally:
        try:
            if os.path.exists(exported):
                os.remove(exported)
        except Exception:
            pass
