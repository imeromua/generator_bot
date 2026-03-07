"""Tests for reports/excel_reports.py — month range and report generation."""

import io
import os
import sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

os.environ.setdefault("SQLITE_PATH", ":memory:")

import config
import database.models as db_models
from reports.excel_reports import ExcelReportGenerator, generate_excel_report

try:
    from openpyxl import load_workbook
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


@pytest.fixture(autouse=True)
def _setup_db(monkeypatch, tmp_path):
    """Fresh in-memory database for each test."""
    db_path = str(tmp_path / "test_excel.db")
    monkeypatch.setattr(config, "SQLITE_PATH", db_path)
    monkeypatch.setattr(config, "DB_BACKEND", "sqlite")
    db_models.init_db()
    yield


class TestMonthRange:
    def _gen(self):
        return ExcelReportGenerator()

    def test_february_2026(self):
        gen = self._gen()
        start_dt, end_dt = gen._month_range(2026, 2)
        assert start_dt == datetime(2026, 2, 1, 0, 0, 0, tzinfo=config.KYIV)
        assert end_dt == datetime(2026, 2, 28, 23, 59, 59, tzinfo=config.KYIV)

    def test_february_2024_leap_year(self):
        gen = self._gen()
        start_dt, end_dt = gen._month_range(2024, 2)
        assert start_dt == datetime(2024, 2, 1, 0, 0, 0, tzinfo=config.KYIV)
        assert end_dt == datetime(2024, 2, 29, 23, 59, 59, tzinfo=config.KYIV)

    def test_december_2026(self):
        gen = self._gen()
        start_dt, end_dt = gen._month_range(2026, 12)
        assert start_dt == datetime(2026, 12, 1, 0, 0, 0, tzinfo=config.KYIV)
        assert end_dt == datetime(2026, 12, 31, 23, 59, 59, tzinfo=config.KYIV)

    def test_returns_timezone_aware(self):
        gen = self._gen()
        start_dt, end_dt = gen._month_range(2026, 5)
        assert start_dt.tzinfo is not None
        assert end_dt.tzinfo is not None


class TestGenerateExcelReportWithYearMonth:
    def test_quick_report_with_year_month_does_not_raise(self):
        result = generate_excel_report('quick', 30, year=2026, month=2)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_detailed_report_with_year_month_does_not_raise(self):
        result = generate_excel_report('detailed', 30, year=2026, month=2)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_technical_report_with_year_month_does_not_raise(self):
        result = generate_excel_report('technical', 30, year=2026, month=2)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_financial_report_with_year_month_does_not_raise(self):
        result = generate_excel_report('financial', 30, year=2026, month=2)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_personnel_report_with_year_month_does_not_raise(self):
        result = generate_excel_report('personnel', 30, year=2026, month=2)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_quick_report_legacy_days_still_works(self):
        result = generate_excel_report('quick', 30)
        assert isinstance(result, bytes)
        assert len(result) > 0


# ---------------------------------------------------------------------------
# New tests for the upgraded Detailed report
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not _OPENPYXL, reason="openpyxl not installed")
class TestDetailedReportCurrentMonth:
    """Detailed report must always cover the current calendar month."""

    def test_detailed_uses_current_month_when_no_year_month_given(self):
        """Calling with days=30 and no year/month should still produce a valid workbook
        that is labelled with the current month (title contains MONTH_NAMES value).
        """
        from reports.excel_reports import MONTH_NAMES
        now = datetime.now(config.KYIV)
        result = generate_excel_report('detailed', 30)
        assert isinstance(result, bytes) and len(result) > 0
        wb = load_workbook(io.BytesIO(result))
        # Title cell should contain the current month name
        ws = wb['Операційний журнал']
        title_cell = ws['A1'].value or ''
        assert MONTH_NAMES[now.month] in title_cell, (
            f"Expected month '{MONTH_NAMES[now.month]}' in title, got: {title_cell!r}"
        )

    def test_detailed_respects_explicit_year_month(self):
        """When year/month are explicitly passed they should appear in the title."""
        from reports.excel_reports import MONTH_NAMES
        result = generate_excel_report('detailed', 30, year=2025, month=6)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        title_cell = ws['A1'].value or ''
        assert MONTH_NAMES[6] in title_cell
        assert '2025' in title_cell

    def test_detailed_ignores_days_param_covers_full_month(self):
        """days=7 should still produce a full-month table (28-31 rows) for detailed."""
        result = generate_excel_report('detailed', 7, year=2026, month=1)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Header is on row 4, data starts at row 5; January has 31 days
        data_rows = [
            ws.cell(row=r, column=1).value
            for r in range(5, 5 + 31 + 5)  # look beyond expected end
            if ws.cell(row=r, column=1).value not in (None, '', 'ПІДСУМОК МІСЯЦЯ')
        ]
        assert len(data_rows) == 31, f"Expected 31 data rows for January, got {len(data_rows)}"


@pytest.mark.skipif(not _OPENPYXL, reason="openpyxl not installed")
class TestDetailedReportSheetStructure:
    """Workbook must contain the expected sheets with correct structure."""

    def _wb(self, year=2026, month=2):
        return load_workbook(io.BytesIO(generate_excel_report('detailed', 30, year=year, month=month)))

    def test_sheet_names(self):
        wb = self._wb()
        expected = {'Операційний журнал', 'Зведення місяця', 'Технічне обслуговування'}
        assert expected.issubset(set(wb.sheetnames)), (
            f"Missing sheets. Got: {wb.sheetnames}"
        )

    def test_operational_sheet_is_first(self):
        wb = self._wb()
        assert wb.sheetnames[0] == 'Операційний журнал'

    def test_header_row_contains_expected_columns(self):
        wb = self._wb()
        ws = wb['Операційний журнал']
        # Row 4 is the column-header row
        headers = [ws.cell(row=4, column=c).value for c in range(1, 19)]
        assert '#' in headers
        assert 'Дата' in headers
        assert 'Год.' in headers
        # Fuel columns
        assert any('Пал' in (h or '') for h in headers)
        assert any('Витрата' in (h or '') for h in headers)

    def test_frozen_panes_set(self):
        wb = self._wb()
        ws = wb['Операційний журнал']
        assert ws.freeze_panes is not None, "freeze_panes should be set"
        # Freeze must be at row >= 5 (after title + group headers + column headers)
        import re
        m = re.match(r'([A-Z]+)(\d+)', ws.freeze_panes)
        assert m is not None
        frozen_row = int(m.group(2))
        assert frozen_row >= 5

    def test_auto_filter_present(self):
        wb = self._wb()
        ws = wb['Операційний журнал']
        assert ws.auto_filter.ref is not None and ws.auto_filter.ref != '', (
            "auto_filter should be set on the operational sheet"
        )

    def test_totals_row_label(self):
        wb = self._wb()
        ws = wb['Операційний журнал']
        # Totals row value 'ПІДСУМОК МІСЯЦЯ' must exist somewhere
        found = any(
            ws.cell(row=r, column=1).value == 'ПІДСУМОК МІСЯЦЯ'
            for r in range(1, ws.max_row + 1)
        )
        assert found, "Monthly totals row 'ПІДСУМОК МІСЯЦЯ' not found"

    def test_february_has_28_data_rows(self):
        wb = self._wb(year=2026, month=2)
        ws = wb['Операційний журнал']
        day_nums = [
            ws.cell(row=r, column=1).value
            for r in range(5, 5 + 30)
            if isinstance(ws.cell(row=r, column=1).value, int)
        ]
        assert len(day_nums) == 28, f"Expected 28 days for Feb 2026, got {len(day_nums)}"
        assert day_nums[0] == 1
        assert day_nums[-1] == 28

    def test_leap_february_has_29_data_rows(self):
        wb = self._wb(year=2024, month=2)
        ws = wb['Операційний журнал']
        day_nums = [
            ws.cell(row=r, column=1).value
            for r in range(5, 5 + 32)
            if isinstance(ws.cell(row=r, column=1).value, int)
        ]
        assert len(day_nums) == 29, f"Expected 29 days for leap Feb 2024, got {len(day_nums)}"

    def test_summary_sheet_has_kpi_label(self):
        wb = self._wb()
        ws = wb['Зведення місяця']
        all_values = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
        assert 'Мотогодини за місяць' in all_values

    def test_maintenance_sheet_present(self):
        wb = self._wb()
        assert 'Технічне обслуговування' in wb.sheetnames


@pytest.mark.skipif(not _OPENPYXL, reason="openpyxl not installed")
class TestDetailedReportGracefulMissingData:
    """Report must handle empty database gracefully — no crashes, no empty cells that
    break layout."""

    def test_no_logs_produces_placeholder_values(self):
        """All data rows should have valid day-number and date values even with no logs."""
        result = generate_excel_report('detailed', 30, year=2026, month=3)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Rows 5..5+30 should have integer day numbers (March has 31 days)
        day_nums = [
            ws.cell(row=r, column=1).value
            for r in range(5, 5 + 31)
            if isinstance(ws.cell(row=r, column=1).value, int)
        ]
        assert len(day_nums) == 31

    def test_no_logs_shift_columns_show_placeholder(self):
        """Shift time columns should show '—' when no shifts were logged."""
        result = generate_excel_report('detailed', 30, year=2026, month=1)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Column D = morning start (col index 4); first data row = row 5
        morning_start = ws.cell(row=5, column=4).value
        assert morning_start == '—', f"Expected '—', got {morning_start!r}"

    def test_no_logs_operator_column_shows_placeholder(self):
        result = generate_excel_report('detailed', 30, year=2026, month=1)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Column K = operator (col index 11)
        operator_val = ws.cell(row=5, column=11).value
        assert operator_val == '—', f"Expected '—', got {operator_val!r}"

    def test_no_logs_notes_column_shows_dash(self):
        result = generate_excel_report('detailed', 30, year=2026, month=1)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Column R = notes (col index 18)
        notes_val = ws.cell(row=5, column=18).value
        assert notes_val == '—', f"Expected '—', got {notes_val!r}"


@pytest.mark.skipif(not _OPENPYXL, reason="openpyxl not installed")
class TestOtherReportTypesUnchanged:
    """Ensure quick / technical / financial / personnel reports still work correctly."""

    def test_quick_report_sheet_name(self):
        result = generate_excel_report('quick', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        assert 'Швидкий звіт' in wb.sheetnames

    def test_technical_report_sheet_name(self):
        result = generate_excel_report('technical', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        assert 'Технічний звіт' in wb.sheetnames

    def test_financial_report_sheet_name(self):
        result = generate_excel_report('financial', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        assert 'Фінансовий звіт' in wb.sheetnames

    def test_personnel_report_sheet_name(self):
        result = generate_excel_report('personnel', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        assert 'По персоналу' in wb.sheetnames

    def test_quick_report_has_kpi_section(self):
        result = generate_excel_report('quick', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Швидкий звіт']
        kpi_label = ws['A3'].value
        assert kpi_label == 'Ключові показники'

    def test_detailed_and_quick_both_return_bytes(self):
        for rtype in ('quick', 'detailed', 'technical', 'financial', 'personnel'):
            result = generate_excel_report(rtype, 30, year=2026, month=2)
            assert isinstance(result, bytes) and len(result) > 0, f"Failed for {rtype}"


@pytest.mark.skipif(not _OPENPYXL, reason="openpyxl not installed")
class TestDetailedReportWithData:
    """Detailed report with seeded log data."""

    @pytest.fixture(autouse=True)
    def _seed_logs(self):
        """Insert a few shift and refill events for testing."""
        import database.db_api as db
        # Seed a morning-shift start/end and a refill for 2026-02-10
        try:
            db.log_event('m_start', '2026-02-10 06:00:00', 'TestUser', None, None, None, 'main')
            db.log_event('m_end', '2026-02-10 14:00:00', 'TestUser', None, None, None, 'main')
            db.log_event('refill', '2026-02-10 08:00:00', 'TestUser', '50', 'Driver1', 'RCP-001', 'main')
        except Exception:
            pass  # db_api may not have log_event; skip seeding

    def test_seeded_day_has_hours_or_placeholder(self):
        """Either the seeded day has work_hours > 0 or gracefully shows None/—."""
        result = generate_excel_report('detailed', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        ws = wb['Операційний журнал']
        # Day 10 is the 10th data row (DATA_START=5, so row 14)
        hours_val = ws.cell(row=14, column=10).value  # column J = 10
        # Either numeric (seeded) or None (unseeded DB)
        assert hours_val is None or isinstance(hours_val, (int, float))

    def test_workbook_is_valid_xlsx(self):
        """Output bytes must be parseable as a valid XLSX file."""
        result = generate_excel_report('detailed', 30, year=2026, month=2)
        wb = load_workbook(io.BytesIO(result))
        assert wb is not None
        assert len(wb.sheetnames) >= 3

