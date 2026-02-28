"""
Telegram Mini App — веб-сервер.

Забезпечує:
  • роздачу статичних файлів (webapp/)
  • REST API для мініаппу (/api/*)
  • валідацію Telegram WebApp initData (HMAC-SHA256)

Запуск:
    python webapp_server.py          # порт за замовчуванням 8080
    WEBAPP_PORT=3000 python webapp_server.py
"""

import io
import json
import logging
import math
import os
import re
import sys
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import AsyncGenerator
from urllib.parse import quote

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    MergedCell = None
    get_column_letter = None

# Додаємо кореневу директорію проєкту до sys.path
_PROJECT_ROOT = Path(__file__).resolve().parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import config  # noqa: E402
import database.models as db_models  # noqa: E402
import database.db_api as db  # noqa: E402

from reports.excel_reports import generate_excel_report, EXCEL_AVAILABLE as _EXCEL_RPT_AVAILABLE  # noqa: E402
from webapp.utils.validation import (
    validate_init_data as _validate_init_data,
    extract_user as _extract_user,
)  # noqa: E402
from webapp.utils.permissions import is_admin as _is_admin  # noqa: E402
from webapp.utils.db_helpers import atomic_transaction, get_admin_info as _get_admin_info  # noqa: E402
from webapp.middleware.rate_limit import RateLimitMiddleware  # noqa: E402
from get_build_version import BUILD_VERSION  # noqa: E402

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Константи
# ---------------------------------------------------------------------------
MAX_EVENTS_LIMIT = 100
MAX_NAME_LENGTH = 100

# ---------------------------------------------------------------------------
# Module-level static file pre-loading
# ---------------------------------------------------------------------------
_webapp_dir = _PROJECT_ROOT / "webapp"
_sw_path = _webapp_dir / "service-worker.js"
try:
    with open(_sw_path, 'r', encoding='utf-8') as f:
        _sw_raw: str | None = f.read()
except FileNotFoundError:
    _sw_raw = None

_sw_content = re.sub(
    r"(const CACHE_VERSION\s*=\s*')[^']*(')",
    rf"\g<1>{BUILD_VERSION}\2",
    _sw_raw or "",
)


async def index_handler(request: Request):
    return FileResponse(str(_webapp_dir / "index.html"))


async def block_handler(request: Request):
    return FileResponse(str(_webapp_dir / "block.html"))


async def sw_handler(request: Request):
    """
    Serve service-worker.js with dynamic cache version injected.
    """
    if _sw_raw is None:
        return Response(content='Service Worker not found', status_code=404, media_type='text/plain')

    return Response(
        content=_sw_content,
        media_type='application/javascript',
        headers={
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0',
            'Service-Worker-Allowed': '/',
        },
    )


# ---------------------------------------------------------------------------
# API — endpoints
# ---------------------------------------------------------------------------


async def api_status(request: Request):
    """GET /api/status — поточний стан генератора."""
    try:
        state = db.get_state()
        active_gen = db.get_active_generator()
        gen_name = db.get_generator_name(active_gen)
        completed = db.get_today_completed_shifts()
        fuel_rate = db.get_fuel_consumption_rate()

        # Оцінка палива під час роботи
        current_fuel = float(state.get("current_fuel", 0))
        status = state.get("status", "OFF")
        estimated_fuel = current_fuel

        if status == "ON":
            start_time_str = state.get("start_time", "")
            start_date_str = state.get("start_date", "")
            if start_time_str:
                try:
                    if start_date_str:
                        start_dt = datetime.strptime(f"{start_date_str} {start_time_str}", "%Y-%m-%d %H:%M")
                    else:
                        logger.warning(
                            f"start_date_str відсутній для активної зміни {state.get('active_shift')}! "
                            f"Використовую поточну дату."
                        )
                        start_dt = datetime.strptime(
                            f"{datetime.now(config.KYIV).strftime('%Y-%m-%d')} {start_time_str}",
                            "%Y-%m-%d %H:%M",
                        )
                    start_dt = start_dt.replace(tzinfo=config.KYIV)
                    now = datetime.now(config.KYIV)
                    elapsed_h = (now - start_dt).total_seconds() / 3600
                    if 0 < elapsed_h < 24:
                        estimated_fuel = max(0, current_fuel - elapsed_h * fuel_rate)
                except (ValueError, TypeError):
                    pass

        # Мотогодини — використовуємо дані активного генератора
        gen_stats = db.get_generator_stats(active_gen)
        total_hours = float(gen_stats.get("total_hours", 0))

        payload = {
            "status": status,
            "generator": active_gen,
            "generator_name": gen_name,
            "current_fuel": round(current_fuel, 1),
            "estimated_fuel": round(estimated_fuel, 1),
            "fuel_rate": fuel_rate,
            "total_hours": round(total_hours, 1),
            "active_shift": state.get("active_shift", "none"),
            "completed_shifts": list(completed),
            "start_time": state.get("start_time", ""),
            "work_start": config.WORK_START_TIME,
            "work_end": config.WORK_END_TIME,
        }
        return payload
    except Exception as e:
        logger.exception("api_status error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_schedule(request: Request):
    """GET /api/schedule?date=YYYY-MM-DD — графік відключень."""
    try:
        date_str = request.query_params.get("date")
        if not date_str:
            now = datetime.now(config.KYIV)
            date_str = now.strftime("%Y-%m-%d")

        # Валідація формату та реальності дати
        try:
            parsed = datetime.strptime(date_str, "%Y-%m-%d")
            if parsed.year < 2000 or parsed.year > 2100:
                raise ValueError("Дата поза допустимим діапазоном")
        except ValueError:
            return JSONResponse(
                content={"error": "Невірний формат або нереальна дата. Використовуйте YYYY-MM-DD"}, status_code=400
            )

        schedule = db.get_schedule(date_str)
        hours = []
        for h in range(24):
            end_h = "24:00" if h == 23 else f"{(h + 1):02d}:00"
            hours.append(
                {
                    "hour": h,
                    "label": f"{h:02d}:00 — {end_h}",
                    "off": schedule.get(h, 0) == 1,
                }
            )

        return {"date": date_str, "hours": hours}
    except Exception as e:
        logger.exception("api_schedule error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_events(request: Request):
    """GET /api/events?limit=20 — останні події."""
    try:
        limit = min(int(request.query_params.get("limit", "20")), MAX_EVENTS_LIMIT)
    except (ValueError, TypeError):
        limit = 20

    try:
        rows = db.get_last_logs(limit)
        events = []
        for row in rows:
            events.append(
                {
                    "event_type": row[0] if len(row) > 0 else "",
                    "timestamp": row[1] if len(row) > 1 else "",
                    "actor": row[2] if len(row) > 2 else "",
                    "value": row[3] if len(row) > 3 else "",
                    "driver": row[4] if len(row) > 4 else "",
                    "receipt": row[5] if len(row) > 5 else "",
                }
            )
        return {"events": events, "count": len(events)}
    except Exception as e:
        logger.exception("api_events error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_maintenance(request: Request):
    """GET /api/maintenance — стан технічного обслуговування."""
    try:
        active_gen = db.get_active_generator()
        stats = db.get_maintenance_stats(active_gen)
        history = db.get_maintenance_history(active_gen, 10)

        history_list = []
        for row in history:
            history_list.append(
                {
                    "id": row[0] if len(row) > 0 else None,
                    "date": row[1] if len(row) > 1 else "",
                    "type": row[2] if len(row) > 2 else "",
                    "hours": row[3] if len(row) > 3 else 0,
                    "admin": row[4] if len(row) > 4 else "",
                }
            )

        # Додаємо інтервали ТО з конфігурації для прогрес-барів
        stats["oil_interval"] = config.OIL_CHANGE_INTERVAL
        stats["spark_interval"] = config.SPARK_CHANGE_INTERVAL
        stats["maintenance_interval"] = config.MAINTENANCE_INTERVAL

        return {
            "generator": active_gen,
            "stats": stats,
            "history": history_list,
        }
    except Exception as e:
        logger.exception("api_maintenance error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_schedule_week(request: Request):
    """GET /api/schedule/week — графік на тиждень (сьогодні + 6 днів)."""
    try:
        now = datetime.now(config.KYIV)
        days = []
        for i in range(7):
            day = now + timedelta(days=i)
            date_str = day.strftime("%Y-%m-%d")
            schedule = db.get_schedule(date_str)
            off_count = sum(1 for v in schedule.values() if v == 1)
            days.append(
                {
                    "date": date_str,
                    "weekday": ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"][day.weekday()],
                    "off_hours": off_count,
                }
            )
        return {"days": days}
    except Exception as e:
        logger.exception("api_schedule_week error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Нові API-ендпоінти для повного функціоналу Mini App
# ---------------------------------------------------------------------------


def _within_work_window(now_t, start_t, end_t) -> bool:
    """True якщо now_t знаходиться в [start_t, end_t)."""
    if start_t <= end_t:
        return start_t <= now_t < end_t
    return now_t >= start_t or now_t < end_t


async def api_user_role(request: Request):
    """GET /api/user/role — роль поточного користувача."""
    user = _extract_user(request)
    try:
        user_id = int(user.get("id", 0)) if user else None
    except (TypeError, ValueError):
        user_id = None

    is_admin = _is_admin(user)
    personnel = db.get_personnel_for_user(user_id) if user_id else None

    return {
        "user_id": user_id,
        "is_admin": is_admin,
        "personnel": personnel,
        "has_personnel": bool(personnel),
        "first_name": user.get("first_name", "") if user else "",
    }


async def api_drivers(request: Request):
    """GET /api/drivers — список водіїв."""
    try:
        drivers = db.get_drivers()
        return {"drivers": list(drivers) if drivers else []}
    except Exception as e:
        logger.exception("api_drivers error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_generators(request: Request):
    """GET /api/generators — статистика обох генераторів."""
    try:
        active_gen = db.get_active_generator()
        main_stats = db.get_generator_stats("main")
        emerg_stats = db.get_generator_stats("emergency")

        def _fmt_stats(stats):
            return {
                "total_hours": round(float(stats.get("total_hours", 0)), 1),
                "last_oil_change": round(float(stats.get("last_oil_change", 0)), 1),
                "last_spark_change": round(float(stats.get("last_spark_change", 0)), 1),
            }

        return {
            "active": active_gen,
            "main": {"name": db.get_generator_name("main"), **_fmt_stats(main_stats)},
            "emergency": {"name": db.get_generator_name("emergency"), **_fmt_stats(emerg_stats)},
        }
    except Exception as e:
        logger.exception("api_generators error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_personnel_me(request: Request):
    """GET /api/personnel/me — персонал поточного користувача."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)

    user_id = int(user.get("id", 0))
    personnel = db.get_personnel_for_user(user_id)
    all_personnel = db.get_personnel_names()

    return {
        "personnel": personnel,
        "all_names": all_personnel,
    }


async def api_action_start(request: Request):
    """POST /api/action/start — старт зміни генератора."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    shift_code = (body.get("shift") or "").strip()
    if shift_code not in ("m", "d", "e", "x"):
        return JSONResponse(content={"error": "Невірний код зміни"}, status_code=400)

    user_id = int(user.get("id", 0))
    personnel = db.get_personnel_for_user(user_id)
    if not personnel:
        return JSONResponse(
            content={"error": "Нема прив'язки до персоналу. Зверніться до адміністратора."}, status_code=400
        )

    now = datetime.now(config.KYIV)

    # Перевірка робочого часу
    try:
        start_t = datetime.strptime(config.WORK_START_TIME, "%H:%M").time()
        end_t = datetime.strptime(config.WORK_END_TIME, "%H:%M").time()
        if not _within_work_window(now.time(), start_t, end_t):
            return JSONResponse(
                content={"error": f"Заборонено поза робочим часом ({config.WORK_START_TIME}–{config.WORK_END_TIME})"},
                status_code=400,
            )
    except Exception:
        pass

    event_type = shift_code + "_start"
    res = db.try_start_shift(event_type, personnel, now)
    if not res.get("ok"):
        reason = res.get("reason", "error")
        if reason == "already_on":
            active = res.get("active_shift", "none")
            return JSONResponse(content={"error": f"Генератор вже працює (активна зміна: {active})"}, status_code=400)
        return JSONResponse(content={"error": "Помилка старту зміни"}, status_code=400)

    return {
        "ok": True,
        "message": f"Зміна запущена о {now.strftime('%H:%M')}",
        "shift": shift_code,
        "time": now.strftime("%H:%M"),
    }


async def api_action_stop(request: Request):
    """POST /api/action/stop — зупинка зміни генератора."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    shift_code = (body.get("shift") or "").strip()
    if shift_code not in ("m", "d", "e", "x"):
        return JSONResponse(content={"error": "Невірний код зміни"}, status_code=400)

    user_id = int(user.get("id", 0))
    personnel = db.get_personnel_for_user(user_id)
    if not personnel:
        return JSONResponse(
            content={"error": "Нема прив'язки до персоналу. Зверніться до адміністратора."}, status_code=400
        )

    now = datetime.now(config.KYIV)
    event_type = shift_code + "_end"
    res = db.try_stop_shift(event_type, personnel, now)
    if not res.get("ok"):
        reason = res.get("reason", "error")
        if reason == "already_off":
            return JSONResponse(content={"error": "Генератор вже вимкнено"}, status_code=400)
        if reason == "wrong_shift":
            active = res.get("active_shift", "none")
            return JSONResponse(content={"error": f"Зараз активна інша зміна: {active}"}, status_code=400)
        return JSONResponse(content={"error": "Помилка зупинки зміни"}, status_code=400)

    duration_hours = res.get("duration_hours", 0.0)
    fuel_consumed = res.get("fuel_consumed", 0.0)
    h = int(duration_hours)
    m = int((duration_hours - h) * 60)

    return {
        "ok": True,
        "message": f"Зміна закрита о {now.strftime('%H:%M')}",
        "shift": shift_code,
        "duration": f"{h:02d}:{m:02d}",
        "fuel_consumed": round(fuel_consumed, 1),
    }


async def api_action_refill(request: Request):
    """POST /api/action/refill — прийом палива."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    driver = (body.get("driver") or "").strip()
    receipt = (body.get("receipt") or "").strip()
    try:
        liters = float(body.get("liters", 0))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірна кількість літрів"}, status_code=400)

    if not driver:
        return JSONResponse(content={"error": "Оберіть водія"}, status_code=400)
    if not receipt or len(receipt) > 50:
        return JSONResponse(content={"error": "Введіть коректний номер чека"}, status_code=400)
    if liters <= 0 or liters > 500:
        return JSONResponse(content={"error": "Кількість літрів має бути від 1 до 500"}, status_code=400)

    user_id = int(user.get("id", 0))
    personnel = db.get_personnel_for_user(user_id)
    if not personnel:
        return JSONResponse(
            content={"error": "Нема прив'язки до персоналу. Зверніться до адміністратора."}, status_code=400
        )

    # Перевірка робочого часу
    now = datetime.now(config.KYIV)
    try:
        start_t = datetime.strptime(config.WORK_START_TIME, "%H:%M").time()
        end_t = datetime.strptime(config.WORK_END_TIME, "%H:%M").time()
        if not _within_work_window(now.time(), start_t, end_t):
            return JSONResponse(
                content={
                    "error": f"Прийом палива заборонено поза робочим часом ({config.WORK_START_TIME}–{config.WORK_END_TIME})"
                },
                status_code=400,
            )
    except Exception:
        pass

    try:
        with atomic_transaction() as conn:
            db.add_log("refill", personnel, str(liters), driver, receipt=receipt, conn=conn)
            db.update_fuel(liters, conn=conn)
    except Exception as e:
        logger.exception("api_action_refill error")
        return JSONResponse(content={"error": str(e)}, status_code=500)

    return {
        "ok": True,
        "message": f"Прийнято {liters:.1f} л палива (Водій: {driver}, Чек: {receipt})",
        "liters": liters,
        "driver": driver,
        "receipt": receipt,
    }


async def api_schedule_toggle(request: Request):
    """POST /api/schedule/toggle — перемикання години графіка відключень."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    date_str = (body.get("date") or "").strip()
    try:
        hour = int(body.get("hour", -1))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірна година"}, status_code=400)

    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return JSONResponse(content={"error": "Невірний формат дати"}, status_code=400)

    if not (0 <= hour <= 23):
        return JSONResponse(content={"error": "Година повинна бути від 0 до 23"}, status_code=400)

    try:
        db.toggle_schedule(date_str, hour)
        schedule = db.get_schedule(date_str)
        new_state = bool(schedule.get(hour, 0))
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "schedule_toggle",
            f"Перемикання графіка {date_str} {hour:02d}:00 → {'відключення' if new_state else 'подача'}",
            target_entity=f"schedule:{date_str}:{hour}",
            new_value={"off": new_state},
        )
        return {
            "ok": True,
            "date": date_str,
            "hour": hour,
            "off": new_state,
            "schedule": {str(h): bool(v) for h, v in schedule.items()},
        }
    except Exception as e:
        logger.exception("api_schedule_toggle error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_generator_switch(request: Request):
    """POST /api/generator/switch — перемикання активного генератора."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    target = (body.get("target") or "").strip()
    if target not in ("main", "emergency"):
        return JSONResponse(content={"error": "Невірний генератор (main або emergency)"}, status_code=400)

    user_id = int(user.get("id", 0))
    user_info = db.get_user(user_id)
    admin_name = user_info[1] if user_info else user.get("first_name", "Адмін")

    try:
        prev_gen = db.get_active_generator()
        success, message = db.switch_generator(target, admin_name)
        db.log_admin_action(
            user_id,
            admin_name,
            "gen_switch",
            f"Перемикання генератора: {prev_gen} → {target}",
            target_entity=f"generator:{target}",
            old_value=prev_gen,
            new_value=target,
            success=success,
        )
        if success:
            return {"ok": True, "message": message, "active": target}
        return JSONResponse(content={"error": message}, status_code=400)
    except Exception as e:
        logger.exception("api_generator_switch error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_maintenance_perform(request: Request):
    """POST /api/maintenance/perform — виконання технічного обслуговування."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    action = (body.get("action") or "").strip()
    generator_id = (body.get("generator") or "main").strip()

    if action not in ("oil", "spark", "maintenance"):
        return JSONResponse(content={"error": "Невірний тип ТО (oil, spark, maintenance)"}, status_code=400)
    if generator_id not in ("main", "emergency"):
        return JSONResponse(content={"error": "Невірний генератор"}, status_code=400)

    user_id = int(user.get("id", 0))
    user_info = db.get_user(user_id)
    actor = user_info[1] if user_info else user.get("first_name", "Адмін")

    try:
        db.record_maintenance(action, actor, generator_id)
        action_names = {"oil": "Заміна мастила", "spark": "Заміна свічок", "maintenance": "Планове ТО"}
        db.log_admin_action(
            user_id,
            actor,
            "maintenance_perform",
            f"{action_names.get(action, action)} на генераторі {generator_id}",
            target_entity=f"generator:{generator_id}",
            new_value={"action": action, "generator": generator_id},
        )
        return {
            "ok": True,
            "message": f"{action_names.get(action, action)} виконано",
        }
    except Exception as e:
        logger.exception("api_maintenance_perform error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_maintenance_set_hours(request: Request):
    """POST /api/maintenance/set-hours — встановлення мотогодин генератора."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    generator_id = (body.get("generator") or "main").strip()
    try:
        hours = float(body.get("hours", -1))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірне значення мотогодин"}, status_code=400)

    if generator_id not in ("main", "emergency"):
        return JSONResponse(content={"error": "Невірний генератор"}, status_code=400)
    if hours < 0 or hours > 100000:
        return JSONResponse(
            content={"error": "Значення мотогодин поза допустимим діапазоном (0–100000)"}, status_code=400
        )

    try:
        old_stats = db.get_generator_stats(generator_id)
        old_hours = float(old_stats.get("total_hours", 0))
        db.set_total_hours(hours, generator_id)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "mnt_set_hours",
            f"Корекція мотогодин генератора {generator_id}: {old_hours:.1f} → {hours:.1f} год",
            target_entity=f"generator:{generator_id}",
            old_value=old_hours,
            new_value=hours,
        )
        return {
            "ok": True,
            "message": f"Мотогодини встановлено: {hours:.1f} год",
            "hours": hours,
        }
    except Exception as e:
        logger.exception("api_maintenance_set_hours error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_fuel_set(request: Request):
    """POST /api/fuel/set — встановлення поточного рівня палива (адмін)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    try:
        fuel = float(body.get("fuel", -1))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірне значення палива"}, status_code=400)

    if fuel < 0 or fuel > 10000:
        return JSONResponse(content={"error": "Значення палива поза допустимим діапазоном"}, status_code=400)

    try:
        old_state = db.get_state()
        old_fuel = float(old_state.get("current_fuel", 0))
        db.set_state("current_fuel", str(fuel))
        user_id = int(user.get("id", 0))
        user_info = db.get_user(user_id)
        actor = user_info[1] if user_info else user.get("first_name", "Адмін")
        db.add_log("corr_fuel_set", actor, str(fuel))
        db.log_admin_action(
            user_id,
            actor,
            "fuel_set",
            f"Корекція палива: {old_fuel:.1f} → {fuel:.1f} л",
            target_entity="fuel",
            old_value=old_fuel,
            new_value=fuel,
        )
        return {"ok": True, "message": f"Паливо встановлено: {fuel:.1f} л"}
    except Exception as e:
        logger.exception("api_fuel_set error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


def _build_daily_report_wb(generator_id: str, period_days: int, now: datetime) -> "Workbook":
    """Будує Excel-книгу з детальним щоденним звітом для одного генератора.

    Стовпці: Дата | Зміна 1 (поч/кін) | Зміна 2 | Зміна 3 | Екстра |
             Залишок ранок | Витрата | Залишок вечір | Мотогодини |
             Заправка (л) | Хто привіз | № чека
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl не встановлено")

    from collections import defaultdict

    wb = Workbook()

    gen_name = db.get_generator_name(generator_id)
    end_date = now.strftime("%Y-%m-%d")
    start_date = (now - timedelta(days=period_days)).strftime("%Y-%m-%d")

    # --- Кольори ---
    BLUE_FILL = PatternFill(start_color="2481CC", end_color="2481CC", fill_type="solid")
    LBLUE_FILL = PatternFill(start_color="D6E8FA", end_color="D6E8FA", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True, size=11)
    BORDER_SIDE = None
    try:
        from openpyxl.styles import Border, Side

        thin = Side(style="thin", color="AAAAAA")
        BORDER_SIDE = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass

    def _style_header(cell, fill=BLUE_FILL):
        cell.font = WHITE_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if BORDER_SIDE:
            cell.border = BORDER_SIDE

    def _style_data(cell, bold=False, align="center"):
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if bold:
            cell.font = Font(bold=True)
        if BORDER_SIDE:
            cell.border = BORDER_SIDE

    # ---- Аркуш «Щоденний звіт» ----
    ws = wb.active
    ws.title = "Щоденний звіт"

    # Шапка
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    header_text = f"Звіт генератора «{gen_name}» за {period_days} днів | Сформовано: {now.strftime('%d.%m.%Y %H:%M')}"
    ws["A1"] = header_text
    ws["A1"].font = Font(bold=True, size=13, color="1A1A2E")
    ws.merge_cells("A1:M1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")

    # Рядки заголовків стовпців
    col_headers_r2 = [
        "Дата",
        "Зміна 1\nпочаток",
        "Зміна 1\nкінець",
        "Зміна 2\nпочаток",
        "Зміна 2\nкінець",
        "Зміна 3\nпочаток",
        "Зміна 3\nкінець",
        "Залишок\nранок, л",
        "Витрата\nза день, л",
        "Залишок\nвечір, л",
        "Мотогодини\n(накопичено)",
        "Заправка\n(прихід), л",
        "Хто привіз / № чека",
    ]
    for ci, h in enumerate(col_headers_r2, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        if ci == 1:
            _style_header(c, BLUE_FILL)
        elif ci in (2, 3):
            _style_header(c, PatternFill(start_color="1A7A44", end_color="1A7A44", fill_type="solid"))
        elif ci in (4, 5):
            _style_header(c, PatternFill(start_color="D4AC0D", end_color="D4AC0D", fill_type="solid"))
        elif ci in (6, 7):
            _style_header(c, PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"))
        elif ci in (8, 9, 10):
            _style_header(c, PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid"))
        elif ci == 11:
            _style_header(c, PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid"))
        elif ci in (12, 13):
            _style_header(c, PatternFill(start_color="117A65", end_color="117A65", fill_type="solid"))
        ws.row_dimensions[2].height = 40

    # Ширини стовпців
    col_widths = [12, 11, 11, 11, 11, 11, 11, 13, 13, 13, 15, 13, 30]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Отримуємо всі логи за період для цього генератора
    logs = db.get_logs_for_period(start_date, end_date, generator_id)

    # Агрегуємо по датах
    days_data = defaultdict(
        lambda: {
            "shifts": {"m": {}, "d": {}, "e": {}, "x": {}},
            "refills": [],
            "morning_fuel": None,
            "evening_fuel": None,
            "hours_start": None,
            "hours_end": None,
        }
    )

    for row_data in logs:
        event_type, ts_str, user_name, value, driver_name, receipt_number, *_ = row_data
        if not ts_str:
            continue
        try:
            ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        date_str = ts.strftime("%Y-%m-%d")
        day = days_data[date_str]

        if event_type.endswith("_start"):
            shift = event_type.split("_")[0]
            if shift in day["shifts"]:
                day["shifts"][shift]["start"] = ts.strftime("%H:%M")
        elif event_type.endswith("_end"):
            shift = event_type.split("_")[0]
            if shift in day["shifts"]:
                day["shifts"][shift]["end"] = ts.strftime("%H:%M")
        elif event_type == "refill":
            try:
                liters = float(value or 0)
            except Exception:
                liters = 0.0
            day["refills"].append((liters, (driver_name or "").strip(), (receipt_number or "").strip()))
        elif event_type == "corr_fuel_set":
            # Використовуємо останню корекцію дня як залишок
            try:
                day["evening_fuel"] = float(value or 0)
            except Exception:
                pass

    # Рядок початку — отримуємо поточний стан
    state = db.get_state()
    current_fuel = float(state.get("current_fuel", 0))

    # Генеруємо рядки за відсортованими датами
    data_row = 3
    prev_fuel = None
    fuel_rate = db.get_fuel_consumption_rate()

    for date_str in sorted(days_data.keys()):
        day = days_data[date_str]
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_fmt = dt.strftime("%d.%m.%Y")
        except Exception:
            date_fmt = date_str

        # Розраховуємо витрату
        total_shift_mins = 0
        for shift_data in day["shifts"].values():
            s_str = shift_data.get("start")
            e_str = shift_data.get("end")
            if s_str and e_str:
                try:
                    s_t = datetime.strptime(s_str, "%H:%M")
                    e_t = datetime.strptime(e_str, "%H:%M")
                    diff = (e_t - s_t).total_seconds() / 60
                    if diff < 0:
                        diff += 24 * 60
                    total_shift_mins += diff
                except Exception:
                    pass

        total_hours = round(total_shift_mins / 60, 2)
        consumption = round(total_hours * fuel_rate, 1) if total_hours > 0 else 0.0
        refill_total = round(sum(r[0] for r in day["refills"]), 1) if day["refills"] else 0.0

        morning_fuel = day.get("morning_fuel") or prev_fuel
        if morning_fuel is not None:
            evening_fuel = round(float(morning_fuel) + refill_total - consumption, 1)
        else:
            morning_fuel = ""
            evening_fuel = ""

        prev_fuel = evening_fuel if isinstance(evening_fuel, float) else None

        drivers_str = ", ".join(f"{drv} (чек {rec})" if rec else drv for _, drv, rec in day["refills"] if drv) or "—"

        row_vals = [
            date_fmt,
            day["shifts"]["m"].get("start", ""),
            day["shifts"]["m"].get("end", ""),
            day["shifts"]["d"].get("start", ""),
            day["shifts"]["d"].get("end", ""),
            day["shifts"]["e"].get("start", ""),
            day["shifts"]["e"].get("end", ""),
            morning_fuel if morning_fuel != "" else "—",
            consumption if consumption > 0 else "—",
            evening_fuel if evening_fuel != "" else "—",
            total_hours if total_hours > 0 else "—",
            refill_total if refill_total > 0 else "—",
            drivers_str,
        ]

        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=data_row, column=ci, value=val)
            bold = ci == 1
            align = "left" if ci == 13 else "center"
            _style_data(c, bold=bold, align=align)
            # Підсвітлення критичних залишків
            if ci == 10 and isinstance(val, float):
                if val < 15:
                    c.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                elif val < 40:
                    c.fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # ---- Аркуш ТО ----
    ws_mnt = wb.create_sheet("Технічне обслуговування")
    stats = db.get_maintenance_stats(generator_id)
    mnt_history = db.get_maintenance_history(generator_id, 100)

    ws_mnt["A1"] = f"Технічне обслуговування — {gen_name}"
    ws_mnt["A1"].font = Font(bold=True, size=13)
    ws_mnt.merge_cells("A1:E1")
    ws_mnt["A1"].alignment = Alignment(horizontal="center")
    ws_mnt["A1"].fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")

    ws_mnt["A3"] = "Мотогодини (загалом):"
    ws_mnt["B3"] = f"{float(stats.get('total_hours', 0)):.1f} год"
    ws_mnt["A3"].font = BOLD_FONT
    ws_mnt["B3"].font = Font(size=11)

    mnt_col_hdrs = ["Дата", "Тип ТО", "Мотогодини на момент ТО", "Виконав", "Примітки"]
    for ci, h in enumerate(mnt_col_hdrs, start=1):
        c = ws_mnt.cell(row=5, column=ci, value=h)
        _style_header(c)

    ws_mnt.column_dimensions["A"].width = 14
    ws_mnt.column_dimensions["B"].width = 22
    ws_mnt.column_dimensions["C"].width = 26
    ws_mnt.column_dimensions["D"].width = 20
    ws_mnt.column_dimensions["E"].width = 20

    mnt_map = {"oil": "Заміна мастила", "spark": "Заміна свічок", "maintenance": "Планове ТО"}
    for ri, rec in enumerate(mnt_history, start=6):
        rec_id, date_s, action, hours, admin_name, *_ = rec
        ws_mnt.cell(row=ri, column=1, value=date_s)
        ws_mnt.cell(row=ri, column=2, value=mnt_map.get(action, action))
        ws_mnt.cell(row=ri, column=3, value=f"{float(hours):.1f} год")
        ws_mnt.cell(row=ri, column=4, value=admin_name or "—")
        for ci in range(1, 5):
            _style_data(ws_mnt.cell(row=ri, column=ci))

    return wb


async def api_report_excel(request: Request):
    """GET /api/report/excel?days=30&generator=main — завантаження Excel-звіту.

    Параметр ``generator`` може бути ``main``, ``emergency`` або ``all``
    (за замовчуванням — активний генератор).
    """
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    if not EXCEL_AVAILABLE:
        return JSONResponse(content={"error": "Модуль openpyxl не встановлено"}, status_code=500)

    try:
        period_days = int(request.query_params.get("days", "30"))
        if period_days < 1:
            period_days = 30
        if period_days > 365:
            period_days = 365
    except (ValueError, TypeError):
        period_days = 30

    generator_param = (request.query_params.get("generator") or "").strip().lower()
    if generator_param not in ("main", "emergency"):
        generator_param = db.get_active_generator()

    try:
        now = datetime.now(config.KYIV)
        gen_name = db.get_generator_name(generator_param)
        wb = _build_daily_report_wb(generator_param, period_days, now)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_gen = "main" if generator_param == "main" else "backup"
        filename = f"report_{safe_gen}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "export_excel",
            f"Експорт Excel-звіту: {gen_name} за {period_days} дн.",
            target_entity=f"generator:{generator_param}",
            new_value={"days": period_days, "generator": generator_param},
        )
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("api_report_excel error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Admin CRUD: drivers & personnel
# ---------------------------------------------------------------------------


async def api_admin_drivers_list(request: Request):
    """GET /api/admin/drivers — список водіїв (лише для адмінів)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        drivers = db.get_drivers()
        return {"drivers": list(drivers) if drivers else []}
    except Exception as e:
        logger.exception("api_admin_drivers_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_drivers_add(request: Request):
    """POST /api/admin/drivers — додати водія."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name or len(name) > MAX_NAME_LENGTH:
        return JSONResponse(content={"error": "Невірне ім'я водія (1–100 символів)"}, status_code=400)

    try:
        ok = db.add_driver(name)
        if not ok:
            return JSONResponse(content={"error": f"Водій «{name}» вже існує"}, status_code=409)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "driver_add",
            f"Додано водія «{name}»",
            target_entity=f"driver:{name}",
            new_value=name,
        )
        return {"ok": True, "message": f"Водія «{name}» додано"}
    except Exception as e:
        logger.exception("api_admin_drivers_add error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_drivers_delete(request: Request):
    """DELETE /api/admin/drivers — видалити водія."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name:
        return JSONResponse(content={"error": "Ім'я водія обов'язкове"}, status_code=400)

    try:
        ok = db.delete_driver(name)
        if not ok:
            return JSONResponse(content={"error": f"Водія «{name}» не знайдено"}, status_code=404)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "driver_delete",
            f"Видалено водія «{name}»",
            target_entity=f"driver:{name}",
            old_value=name,
        )
        return {"ok": True, "message": f"Водія «{name}» видалено"}
    except Exception as e:
        logger.exception("api_admin_drivers_delete error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_list(request: Request):
    """GET /api/admin/personnel — список персоналу (лише для адмінів)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        names = db.get_personnel_names()
        users_with_p = db.get_all_users_with_personnel()
        users_list = [{"user_id": row[0], "full_name": row[1] or "", "personnel": row[2] or ""} for row in users_with_p]
        return {"personnel": names, "users": users_list}
    except Exception as e:
        logger.exception("api_admin_personnel_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_add(request: Request):
    """POST /api/admin/personnel — додати ПІБ персоналу."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name or len(name) > MAX_NAME_LENGTH:
        return JSONResponse(content={"error": "Невірне ім'я (1–100 символів)"}, status_code=400)

    try:
        ok = db.add_personnel_name(name)
        if not ok:
            return JSONResponse(content={"error": f"Персонал «{name}» вже існує"}, status_code=409)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_add",
            f"Додано персонал «{name}»",
            target_entity=f"personnel:{name}",
            new_value=name,
        )
        return {"ok": True, "message": f"Персонал «{name}» додано"}
    except Exception as e:
        logger.exception("api_admin_personnel_add error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_delete(request: Request):
    """DELETE /api/admin/personnel — видалити ПІБ персоналу."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name:
        return JSONResponse(content={"error": "Ім'я обов'язкове"}, status_code=400)

    try:
        ok = db.delete_personnel_name(name)
        if not ok:
            return JSONResponse(content={"error": f"Персонал «{name}» не знайдено"}, status_code=404)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_delete",
            f"Видалено персонал «{name}»",
            target_entity=f"personnel:{name}",
            old_value=name,
        )
        return {"ok": True, "message": f"Персонал «{name}» видалено"}
    except Exception as e:
        logger.exception("api_admin_personnel_delete error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_assign(request: Request):
    """POST /api/admin/personnel/assign — прив'язати персонал до Telegram-користувача."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    try:
        target_user_id = int(body.get("user_id", 0))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірний user_id"}, status_code=400)

    personnel_name = (body.get("personnel") or "").strip() or None

    if not target_user_id:
        return JSONResponse(content={"error": "user_id обов'язковий"}, status_code=400)

    try:
        old_personnel = db.get_personnel_for_user(target_user_id)
        db.set_personnel_for_user(target_user_id, personnel_name)
        admin_id, admin_name = _get_admin_info(user)
        if personnel_name:
            msg = f"Прив'язано: user {target_user_id} → «{personnel_name}»"
        else:
            msg = f"Прив'язку для user {target_user_id} знято"
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_assign",
            msg,
            target_entity=f"user:{target_user_id}",
            old_value=old_personnel,
            new_value=personnel_name,
        )
        return {"ok": True, "message": msg}
    except Exception as e:
        logger.exception("api_admin_personnel_assign error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_sync(request: Request):
    """POST /api/admin/sync — запуск синхронізації з Google Sheets (експорт)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from services.sheets_export import full_export

        result = full_export()
        updated = result.get("updated", [])
        skipped = result.get("skipped", [])
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "export_sheets",
            f"Синхронізація з Google Sheets: {len(updated)} дн. оновлено",
            new_value={"updated": len(updated), "skipped": len(skipped)},
        )
        return {
            "ok": True,
            "message": f"Синхронізовано: {len(updated)} дн., пропущено: {len(skipped)} дн.",
            "updated": updated,
            "skipped": skipped,
        }
    except Exception as e:
        logger.exception("api_admin_sync error")
        return JSONResponse(content={"error": f"Помилка синхронізації: {e}"}, status_code=500)


# ---------------------------------------------------------------------------
# Admin Audit Log endpoints
# ---------------------------------------------------------------------------


async def api_admin_audit(request: Request):
    """GET /api/admin/audit — журнал дій адміністраторів.

    Query params:
        limit      (int, default 50, max 200)
        offset     (int, default 0)
        action_type (str, optional filter)
        admin_id   (int, optional filter by admin user ID)
        date_from  (str YYYY-MM-DD, optional)
        date_to    (str YYYY-MM-DD, optional)
    """
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        limit = min(int(request.query_params.get("limit", "50")), 200)
    except (TypeError, ValueError):
        limit = 50
    try:
        offset = max(int(request.query_params.get("offset", "0")), 0)
    except (TypeError, ValueError):
        offset = 0

    action_type = request.query_params.get("action_type", "").strip()
    date_from = request.query_params.get("date_from", "").strip()
    date_to = request.query_params.get("date_to", "").strip()
    try:
        admin_filter = int(request.query_params.get("admin_id", "0"))
    except (TypeError, ValueError):
        admin_filter = 0

    try:
        rows = db.get_audit_logs(
            limit=limit,
            offset=offset,
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )
        total = db.count_audit_logs(
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )
        entries = [
            {
                "id": r[0],
                "timestamp": r[1],
                "admin_user_id": r[2],
                "admin_name": r[3],
                "action_type": r[4],
                "action_description": r[5],
                "target_entity": r[6],
                "old_value": r[7],
                "new_value": r[8],
                "success": bool(r[9]),
            }
            for r in rows
        ]
        return {
            "entries": entries,
            "total": total,
            "limit": limit,
            "offset": offset,
        }
    except Exception as e:
        logger.exception("api_admin_audit error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_audit_export(request: Request):
    """GET /api/admin/audit/export — експорт журналу дій у Excel."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    if not EXCEL_AVAILABLE:
        return JSONResponse(content={"error": "Модуль openpyxl не встановлено"}, status_code=500)

    action_type = request.query_params.get("action_type", "").strip()
    date_from = request.query_params.get("date_from", "").strip()
    date_to = request.query_params.get("date_to", "").strip()
    try:
        admin_filter = int(request.query_params.get("admin_id", "0"))
    except (TypeError, ValueError):
        admin_filter = 0

    try:
        rows = db.get_audit_logs(
            limit=5000,
            offset=0,
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Журнал дій"

        headers = [
            "#",
            "Час",
            "Адмін ID",
            "Адмін",
            "Тип дії",
            "Опис",
            "Об'єкт",
            "Старе значення",
            "Нове значення",
            "Успішно",
        ]
        header_fill = PatternFill(start_color="2481CC", end_color="2481CC", fill_type="solid")
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

        col_widths = [6, 20, 12, 20, 18, 40, 25, 20, 20, 10]
        from openpyxl.utils import get_column_letter as _gcl

        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[_gcl(ci)].width = w

        for ri, r in enumerate(rows, start=2):
            ws.cell(row=ri, column=1, value=r[0])
            ws.cell(row=ri, column=2, value=r[1])
            ws.cell(row=ri, column=3, value=r[2])
            ws.cell(row=ri, column=4, value=r[3] or "")
            ws.cell(row=ri, column=5, value=r[4] or "")
            ws.cell(row=ri, column=6, value=r[5] or "")
            ws.cell(row=ri, column=7, value=r[6] or "")
            ws.cell(row=ri, column=8, value=r[7] or "")
            ws.cell(row=ri, column=9, value=r[8] or "")
            ws.cell(row=ri, column=10, value="✅" if r[9] else "❌")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now = datetime.now(config.KYIV)
        filename = f"audit_log_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("api_admin_audit_export error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Backup endpoints
# ---------------------------------------------------------------------------


async def api_admin_config_get(request: Request):
    """GET /api/admin/config — поточні налаштування генераторів та глобальні."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        from database.api.config import get_generator_config, get_global_config

        def _build_param_response(cfg: dict, params: tuple) -> dict:
            return {
                p: {
                    "value": cfg[p]["value"] if p in cfg else None,
                    "last_updated": cfg[p]["last_updated"] if p in cfg else "",
                    "updated_by": cfg[p]["updated_by"] if p in cfg else "",
                }
                for p in params
            }

        main_cfg = get_generator_config("main")
        emerg_cfg = get_generator_config("emergency")
        global_cfg = get_global_config()
        gen_params = ("fuel_consumption_rate",)

        return {
            "generators": {
                "main": _build_param_response(main_cfg, gen_params),
                "emergency": _build_param_response(emerg_cfg, gen_params),
            },
            "global": _build_param_response(global_cfg, ("fuel_price",)),
        }
    except Exception as e:
        logger.exception("api_admin_config_get error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_set_generator(request: Request):
    """POST /api/admin/config/generator — змінити параметр генератора."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    generator_id = str(body.get("generator_id", "")).strip()
    param_name = str(body.get("param_name", "")).strip()
    value = body.get("value")

    if not generator_id or not param_name or value is None:
        return JSONResponse(content={"error": "generator_id, param_name та value обов'язкові"}, status_code=400)

    try:
        value = float(value)
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "value має бути числом"}, status_code=400)

    try:
        from database.api.config import (
            set_generator_param,
            get_generator_param,
            VALID_GENERATOR_IDS,
            VALID_GENERATOR_PARAMS,
        )

        if generator_id not in VALID_GENERATOR_IDS:
            return JSONResponse(
                content={"error": f"generator_id має бути одним із: {', '.join(VALID_GENERATOR_IDS)}"}, status_code=400
            )
        if param_name not in VALID_GENERATOR_PARAMS:
            return JSONResponse(
                content={"error": f"param_name має бути одним із: {', '.join(VALID_GENERATOR_PARAMS)}"}, status_code=400
            )

        admin_id, admin_name = _get_admin_info(user)
        old_value = get_generator_param(generator_id, param_name)

        ok = set_generator_param(generator_id, param_name, value, admin_id, admin_name)
        if not ok:
            return JSONResponse(content={"error": "Не вдалося зберегти налаштування"}, status_code=500)

        db.log_admin_action(
            admin_id,
            admin_name,
            "config_generator_set",
            f"Змінено {param_name} для {generator_id}: {old_value} → {value}",
            target_entity=f"generator:{generator_id}",
            old_value=old_value,
            new_value=value,
        )
        return {"ok": True, "message": "Налаштування збережено", "old_value": old_value, "new_value": value}
    except ValueError as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
    except Exception as e:
        logger.exception("api_admin_config_set_generator error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_set_global(request: Request):
    """POST /api/admin/config/global — змінити глобальний параметр."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    param_name = str(body.get("param_name", "")).strip()
    value = body.get("value")

    if not param_name or value is None:
        return JSONResponse(content={"error": "param_name та value обов'язкові"}, status_code=400)

    try:
        value = float(value)
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "value має бути числом"}, status_code=400)

    try:
        from database.api.config import set_global_param, get_global_param, VALID_GLOBAL_PARAMS

        if param_name not in VALID_GLOBAL_PARAMS:
            return JSONResponse(
                content={"error": f"param_name має бути одним із: {', '.join(VALID_GLOBAL_PARAMS)}"}, status_code=400
            )

        admin_id, admin_name = _get_admin_info(user)
        old_value = get_global_param(param_name)

        ok = set_global_param(param_name, value, admin_id, admin_name)
        if not ok:
            return JSONResponse(content={"error": "Не вдалося зберегти налаштування"}, status_code=500)

        db.log_admin_action(
            admin_id,
            admin_name,
            "config_global_set",
            f"Змінено {param_name}: {old_value} → {value}",
            target_entity=f"global:{param_name}",
            old_value=old_value,
            new_value=value,
        )
        return {"ok": True, "message": "Налаштування збережено", "old_value": old_value, "new_value": value}
    except ValueError as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
    except Exception as e:
        logger.exception("api_admin_config_set_global error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_history(request: Request):
    """GET /api/admin/config/history?limit=20 — історія змін налаштувань."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        limit = int(request.query_params.get("limit", "20"))
        limit = max(1, min(limit, 100))
        offset = int(request.query_params.get("offset", "0"))
        offset = max(0, offset)

        from database.api.config import get_config_history

        history = get_config_history(limit=limit, offset=offset)
        return {"history": history}
    except Exception as e:
        logger.exception("api_admin_config_history error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Backup endpoints
# ---------------------------------------------------------------------------


async def api_admin_backups_list(request: Request):
    """GET /api/admin/backups — список резервних копій."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from backup import list_backups, DEFAULT_BACKUP_DIR

        backups = list_backups()
        return {"backups": backups, "count": len(backups)}
    except Exception as e:
        logger.exception("api_admin_backups_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_backup_create(request: Request):
    """POST /api/admin/backup — створити резервну копію вручну."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from backup import create_backup

        backup_path = create_backup()
        size_kb = round(backup_path.stat().st_size / 1024, 1)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "backup_create",
            f"Створено резервну копію вручну: {backup_path.name} ({size_kb} KB)",
            target_entity=backup_path.name,
            new_value={"filename": backup_path.name, "size_kb": size_kb},
        )
        return {
            "ok": True,
            "filename": backup_path.name,
            "size_kb": size_kb,
            "message": f"Резервну копію створено: {backup_path.name}",
        }
    except Exception as e:
        logger.exception("api_admin_backup_create error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_backup_download(request: Request, filename: str):
    """GET /api/admin/backup/download/{filename} — завантажити резервну копію."""
    import re as _re

    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    filename = filename or ""
    # Security: strictly validate the expected filename pattern to prevent path traversal
    # and injection attacks. Pattern: backup_YYYY-MM-DD_HH-MM.sql.gz
    _BACKUP_FILENAME_RE = _re.compile(r'^backup_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}\.sql\.gz$')
    if not filename or not _BACKUP_FILENAME_RE.match(filename):
        return JSONResponse(content={"error": "Невірне ім'я файлу"}, status_code=400)

    try:
        from backup import DEFAULT_BACKUP_DIR

        backup_path = DEFAULT_BACKUP_DIR / filename
        if not backup_path.exists():
            return JSONResponse(content={"error": "Файл не знайдено"}, status_code=404)

        with open(backup_path, "rb") as f:
            data = f.read()

        return Response(
            content=data,
            media_type="application/gzip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("api_admin_backup_download error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Task 5: Notification preferences API
# ---------------------------------------------------------------------------


async def api_notifications_get(request: Request):
    """GET /api/notifications/preferences — get user notification preferences."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from database.api.notifications import get_user_preferences, NOTIFICATION_TYPES

        user_id = int(user.get("id", 0))
        prefs = get_user_preferences(user_id)
        return {
            "preferences": prefs,
            "types": {k: {"label": v["label"], "category": v["category"]} for k, v in NOTIFICATION_TYPES.items()},
        }
    except Exception as e:
        logger.exception("api_notifications_get error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_notifications_set(request: Request):
    """POST /api/notifications/preferences — update user notification preference."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    notification_type = str(body.get("notification_type", "")).strip()
    enabled = body.get("enabled")
    quiet_hours_start = body.get("quiet_hours_start")
    quiet_hours_end = body.get("quiet_hours_end")

    if not notification_type:
        return JSONResponse(content={"error": "notification_type обов'язковий"}, status_code=400)

    try:
        from database.api.notifications import set_user_preference, NOTIFICATION_TYPES

        if notification_type not in NOTIFICATION_TYPES:
            return JSONResponse(content={"error": "Невідомий тип сповіщення"}, status_code=400)
        user_id = int(user.get("id", 0))
        set_user_preference(
            user_id,
            notification_type,
            bool(enabled) if enabled is not None else True,
            quiet_hours_start or None,
            quiet_hours_end or None,
        )
        return {"ok": True, "message": "Налаштування збережено"}
    except Exception as e:
        logger.exception("api_notifications_set error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_notifications_test(request: Request):
    """POST /api/notifications/test — send a test notification to the user."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    # This endpoint is informational — the actual bot send happens via the Telegram bot
    return {
        "ok": True,
        "message": "🔔 Тест сповіщень. Якщо ви бачите це в webapp — система працює.",
    }


# ---------------------------------------------------------------------------
# Task 6: Fuel orders API
# ---------------------------------------------------------------------------


async def api_fuel_orders_list(request: Request):
    """GET /api/fuel/orders — list fuel orders."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from database.api.fuel_orders import get_orders, get_fuel_consumption_stats

        status_filter = request.query_params.get("status") or None
        orders = get_orders(status=status_filter, limit=50)
        stats = get_fuel_consumption_stats(days=30)
        return {"orders": orders, "consumption_stats": stats}
    except Exception as e:
        logger.exception("api_fuel_orders_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_fuel_orders_create(request: Request):
    """POST /api/fuel/orders — create a new fuel order."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    amount = body.get("amount_liters")
    if amount is None:
        return JSONResponse(content={"error": "amount_liters обов'язковий"}, status_code=400)
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірне значення кількості літрів"}, status_code=400)

    if amount <= 0 or amount > 100000:
        return JSONResponse(content={"error": "Кількість літрів поза допустимим діапазоном"}, status_code=400)

    try:
        from database.api.fuel_orders import create_order
        from utils.time import now_kiev

        now = now_kiev()
        user_id = int(user.get("id", 0))
        order_id = create_order(
            created_at=now.strftime("%Y-%m-%d %H:%M:%S"),
            amount_liters=amount,
            requested_by=user_id or None,
            supplier=str(body.get("supplier", "")).strip() or None,
            price=float(body["price"]) if body.get("price") else None,
            delivery_date=str(body.get("delivery_date", "")).strip() or None,
            notes=str(body.get("notes", "")).strip() or None,
        )
        return {"ok": True, "order_id": order_id, "message": "Замовлення створено"}
    except Exception as e:
        logger.exception("api_fuel_orders_create error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_fuel_orders_update(request: Request):
    """POST /api/fuel/orders/update — update a fuel order status."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    order_id = body.get("order_id")
    if not order_id:
        return JSONResponse(content={"error": "order_id обов'язковий"}, status_code=400)

    try:
        from database.api.fuel_orders import update_order, update_order_status, VALID_STATUSES

        new_status = str(body.get("status", "")).strip()
        if new_status and new_status not in VALID_STATUSES:
            return JSONResponse(
                content={"error": f"Статус має бути одним із: {', '.join(VALID_STATUSES)}"}, status_code=400
            )

        updated = update_order(
            int(order_id),
            supplier=str(body.get("supplier", "")).strip() or None,
            price=float(body["price"]) if body.get("price") else None,
            delivery_date=str(body.get("delivery_date", "")).strip() or None,
            status=new_status or None,
            notes=str(body.get("notes", "")).strip() or None,
        )
        if not updated:
            return JSONResponse(content={"error": "Нічого не оновлено"}, status_code=400)

        # If delivered, add fuel to current level
        if new_status == "delivered":
            from database.api.fuel_orders import get_order

            order = get_order(int(order_id))
            if order:
                current_fuel = float(db.get_state_value("current_fuel", "0") or "0")
                new_fuel = current_fuel + order["amount_liters"]
                db.set_state("current_fuel", str(round(new_fuel, 1)))
                user_id = int(user.get("id", 0))
                user_info = db.get_user(user_id)
                actor = user_info[1] if user_info else user.get("first_name", "Адмін")
                db.add_log("refill", actor, str(order["amount_liters"]))

        return {"ok": True, "message": "Замовлення оновлено"}
    except Exception as e:
        logger.exception("api_fuel_orders_update error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Task 8: Shift schedule API
# ---------------------------------------------------------------------------


async def api_shifts_get(request: Request):
    """GET /api/shifts/schedule — get shift schedule for a month or date."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from database.api.shift_schedule import get_month_schedule, get_date_schedule, get_personnel_shift_counts

        date_param = request.query_params.get("date")
        month_param = request.query_params.get("month")  # YYYY-MM

        if date_param:
            entries = get_date_schedule(date_param)
            return {"shifts": entries}
        elif month_param:
            year, month = int(month_param[:4]), int(month_param[5:7])
            entries = get_month_schedule(year, month)
            counts = get_personnel_shift_counts(month_param)
            return {"shifts": entries, "personnel_counts": counts, "month": month_param}
        else:
            from utils.time import now_kiev

            now = now_kiev()
            month_str = now.strftime("%Y-%m")
            year, month = now.year, now.month
            entries = get_month_schedule(year, month)
            counts = get_personnel_shift_counts(month_str)
            return {"shifts": entries, "personnel_counts": counts, "month": month_str}
    except Exception as e:
        logger.exception("api_shifts_get error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_shifts_set(request: Request):
    """POST /api/shifts/schedule — create or update a shift assignment."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    date = str(body.get("date", "")).strip()
    shift_type = str(body.get("shift_type", "")).strip()
    if not date or not shift_type:
        return JSONResponse(content={"error": "date та shift_type обов'язкові"}, status_code=400)

    try:
        from database.api.shift_schedule import upsert_shift, VALID_SHIFT_TYPES

        if shift_type not in VALID_SHIFT_TYPES:
            return JSONResponse(
                content={"error": f"shift_type має бути одним із: {', '.join(VALID_SHIFT_TYPES)}"}, status_code=400
            )

        upsert_shift(
            date=date,
            shift_type=shift_type,
            assigned_personnel_id=str(body.get("assigned_personnel_id", "")).strip() or None,
            status=str(body.get("status", "planned")).strip() or "planned",
            notes=str(body.get("notes", "")).strip() or None,
        )
        return {"ok": True, "message": "Зміну збережено"}
    except Exception as e:
        logger.exception("api_shifts_set error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_shifts_auto(request: Request):
    """POST /api/shifts/auto — generate auto schedule for a month."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    month_param = str(body.get("month", "")).strip()
    save = bool(body.get("save", False))

    if not month_param or len(month_param) < 7:
        return JSONResponse(content={"error": "month (YYYY-MM) обов'язковий"}, status_code=400)

    try:
        from database.api.shift_schedule import auto_schedule_month, upsert_shift

        year, month = int(month_param[:4]), int(month_param[5:7])
        # Get personnel list from DB
        personnel_list = db.get_personnel_names()
        if not personnel_list:
            return JSONResponse(content={"error": "Персонал не знайдено. Додайте персонал спочатку."}, status_code=400)

        assignments = auto_schedule_month(year, month, personnel_list)

        if save:
            for a in assignments:
                upsert_shift(
                    date=a["date"],
                    shift_type=a["shift_type"],
                    assigned_personnel_id=a["assigned_personnel_id"],
                    status=a["status"],
                    notes=a["notes"],
                )

        return {
            "ok": True,
            "assignments": assignments,
            "saved": save,
            "message": f"Згенеровано {len(assignments)} змін" + (" та збережено" if save else " (попередній перегляд)"),
        }
    except Exception as e:
        logger.exception("api_shifts_auto error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_shifts_analytics(request: Request):
    """GET /api/shifts/analytics — shift load analytics."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from database.api.shift_schedule import get_personnel_shift_counts, get_month_schedule
        from utils.time import now_kiev

        now = now_kiev()
        month_str = request.query_params.get("month") or now.strftime("%Y-%m")
        year, month = int(month_str[:4]), int(month_str[5:7])

        counts = get_personnel_shift_counts(month_str)
        entries = get_month_schedule(year, month)

        # Status breakdown
        status_counts: dict = {}
        for e in entries:
            s = e.get("status", "planned")
            status_counts[s] = status_counts.get(s, 0) + 1

        return {
            "month": month_str,
            "total_shifts": len(entries),
            "personnel_counts": counts,
            "status_breakdown": status_counts,
        }
    except Exception as e:
        logger.exception("api_shifts_analytics error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Статичні файли та додаток
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Analytics helpers
# ---------------------------------------------------------------------------


def _safe_round(v: float, ndigits: int = 1) -> float:
    """Round a float, replacing non-finite values with 0.0."""
    try:
        f = float(v)
        return round(f, ndigits) if math.isfinite(f) else 0.0
    except (TypeError, ValueError):
        return 0.0


def _build_daily_stats(start_dt: datetime, end_dt: datetime, generator_id: str | None = None) -> list:
    """Збирає денну статистику з логів за вказаний діапазон дат.

    Повертає список dict:
      {date, work_hours, fuel_consumed, fuel_rate, outage_hours, refill_liters,
       morning_balance, evening_balance}
    """
    from database.api.logs import get_logs_for_period
    from database.api.schedule import get_schedule
    import config as _cfg

    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")

    logs = get_logs_for_period(start_str, end_str, generator_id)

    # Групуємо start/stop пари по датах
    daily: dict = {}
    current_day = start_dt.date()
    while current_day <= end_dt.date():
        daily[current_day.strftime("%Y-%m-%d")] = {
            "date": current_day.strftime("%Y-%m-%d"),
            "work_hours": 0.0,
            "fuel_consumed": 0.0,
            "fuel_rate": 0.0,
            "outage_hours": 0,
            "refill_liters": 0.0,
            "morning_balance": None,
            "evening_balance": None,
        }
        current_day += timedelta(days=1)

    # Відстежуємо старт/стоп по генератору
    pending_start: dict = {}  # gen_id -> datetime

    for row in logs:
        event_type, ts_str, user_name, value, driver_name, receipt_number, gen_id = row
        try:
            ts = datetime.strptime(ts_str[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        day_key = ts.strftime("%Y-%m-%d")
        if day_key not in daily:
            continue

        if event_type in ("m_start", "d_start", "e_start", "x_start"):
            pending_start[gen_id or "main"] = ts

        elif event_type in ("m_end", "d_end", "e_end", "x_end"):
            gen_key = gen_id or "main"
            start_ts = pending_start.pop(gen_key, None)
            if start_ts:
                hours = (ts - start_ts).total_seconds() / 3600.0
                if 0 < hours < 24:
                    fuel_rate = getattr(_cfg, "FUEL_CONSUMPTION", 5.0)
                    daily[day_key]["work_hours"] += hours
                    daily[day_key]["fuel_consumed"] += hours * fuel_rate
                    daily[day_key]["fuel_rate"] = fuel_rate

        elif event_type == "refill":
            try:
                liters = float(value or 0)
                daily[day_key]["refill_liters"] += liters
            except Exception:
                pass

    # Відключення — з таблиці schedule
    for day_key in daily:
        try:
            sched = get_schedule(day_key)
            daily[day_key]["outage_hours"] = sum(1 for v in sched.values() if v == 1)
        except Exception:
            pass

    # Округлення
    for d in daily.values():
        d["fuel_consumed"] = round(d["fuel_consumed"], 2)
        d["work_hours"] = round(d["work_hours"], 2)
        if d["work_hours"] > 0:
            d["fuel_rate"] = round(d["fuel_consumed"] / d["work_hours"], 3)

    # Розрахунок залишків палива (ранок/вечір)
    sorted_days = sorted(daily.values(), key=lambda x: x["date"])
    try:
        current_fuel = float(db.get_state().get("current_fuel", 0) or 0)
    except Exception:
        current_fuel = 0.0
    total_period_refills = sum(d["refill_liters"] for d in sorted_days)
    total_period_consumption = sum(d["fuel_consumed"] for d in sorted_days)
    starting_fuel = current_fuel - total_period_refills + total_period_consumption
    prev_balance: float | None = starting_fuel if starting_fuel > 0 else None
    for d in sorted_days:
        morning_balance = prev_balance
        if morning_balance is not None:
            evening_balance: float | None = round(
                float(morning_balance) + d["refill_liters"] - d["fuel_consumed"], 1
            )
        else:
            evening_balance = None
        d["morning_balance"] = morning_balance
        d["evening_balance"] = evening_balance
        prev_balance = evening_balance if isinstance(evening_balance, float) else None

    return sorted_days


# ---------------------------------------------------------------------------
# Task 9: Analytics API
# ---------------------------------------------------------------------------


async def api_analytics_kpi(request: Request):
    """GET /api/analytics/kpi — KPI картки для дашборду аналітики."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        days = int(request.query_params.get("days", "30"))
        days = max(1, min(days, 365))
        gen_id = request.query_params.get("generator") or None

        from utils.time import now_kiev

        now = now_kiev()
        start_dt = now - timedelta(days=days - 1)

        daily = _build_daily_stats(start_dt, now, gen_id)

        total_hours = sum(d["work_hours"] for d in daily)
        total_fuel = sum(d["fuel_consumed"] for d in daily)
        avg_per_day = total_hours / len(daily) if daily else 0
        avg_rate = total_fuel / total_hours if total_hours > 0 else 0
        fuel_price = db.get_fuel_price_db()
        fuel_cost = total_fuel * fuel_price
        total_outage = sum(d["outage_hours"] for d in daily)
        total_avail = days * 24
        efficiency = round((total_outage / total_avail) * 100, 1) if total_avail > 0 else 0

        # Порівняння з попереднім таким же періодом
        prev_start = start_dt - timedelta(days=days)
        prev_end = start_dt - timedelta(days=1)
        prev_daily = _build_daily_stats(prev_start, prev_end, gen_id)
        prev_hours = sum(d["work_hours"] for d in prev_daily)
        prev_fuel = sum(d["fuel_consumed"] for d in prev_daily)

        def _percent_change(curr, prev):
            if prev == 0:
                return None
            return round((curr - prev) / prev * 100, 1)

        return {
            "period_days": days,
            "total_hours": round(total_hours, 1),
            "avg_hours_per_day": round(avg_per_day, 2),
            "avg_fuel_rate": round(avg_rate, 3),
            "total_fuel": round(total_fuel, 1),
            "fuel_cost": round(fuel_cost, 0),
            "efficiency_pct": efficiency,
            "total_outage_hours": total_outage,
            "prev_total_hours": round(prev_hours, 1),
            "prev_total_fuel": round(prev_fuel, 1),
            "hours_change_pct": _percent_change(total_hours, prev_hours),
            "fuel_change_pct": _percent_change(total_fuel, prev_fuel),
        }
    except Exception as e:
        logger.exception("api_analytics_kpi error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_analytics_fuel_timeline(request: Request):
    """GET /api/analytics/fuel-timeline — дані для графіка витрати палива."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        days = int(request.query_params.get("days", "30"))
        days = max(1, min(days, 365))
        gen_id = request.query_params.get("generator") or None

        from utils.time import now_kiev

        now = now_kiev()
        start_dt = now - timedelta(days=days - 1)

        daily = _build_daily_stats(start_dt, now, gen_id)

        # Прогноз (наступні 7 днів)
        from ml_models import get_fuel_forecast

        forecast_obj = get_fuel_forecast()
        forecast_obj.train(daily)
        avg_outage = sum(d["outage_hours"] for d in daily) / len(daily) if daily else 4.0
        forecast = forecast_obj.predict(7, avg_outage)

        return {
            "actual": [
                {
                    "date": d["date"],
                    "fuel_consumed": d["fuel_consumed"],
                    "work_hours": d["work_hours"],
                    "refill_liters": d["refill_liters"],
                    "outage_hours": d["outage_hours"],
                    "morning_balance": d.get("morning_balance"),
                    "evening_balance": d.get("evening_balance"),
                }
                for d in daily
            ],
            "forecast": forecast,
        }
    except Exception as e:
        logger.exception("api_analytics_fuel_timeline error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_analytics_motor_hours(request: Request):
    """GET /api/analytics/motor-hours — мотогодини генераторів."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        days = int(request.query_params.get("days", "30"))
        days = max(1, min(days, 365))

        from utils.time import now_kiev

        now = now_kiev()
        start_dt = now - timedelta(days=days - 1)

        main_daily = _build_daily_stats(start_dt, now, "main")
        emergency_daily = _build_daily_stats(start_dt, now, "emergency")

        # Об'єднуємо по датах
        date_map: dict = {}
        for d in main_daily:
            date_map[d["date"]] = {"date": d["date"], "main": d["work_hours"], "emergency": 0.0}
        for d in emergency_daily:
            if d["date"] in date_map:
                date_map[d["date"]]["emergency"] = d["work_hours"]
            else:
                date_map[d["date"]] = {"date": d["date"], "main": 0.0, "emergency": d["work_hours"]}

        combined = sorted(date_map.values(), key=lambda x: x["date"])

        from database.api.generator import get_generator_stats

        main_stats = get_generator_stats("main")
        emergency_stats = get_generator_stats("emergency")

        return {
            "daily": combined,
            "totals": {
                "main": {
                    "total_hours": main_stats.get("total_hours", 0),
                    "period_hours": round(sum(d["main"] for d in combined), 1),
                },
                "emergency": {
                    "total_hours": emergency_stats.get("total_hours", 0),
                    "period_hours": round(sum(d["emergency"] for d in combined), 1),
                },
            },
        }
    except Exception as e:
        logger.exception("api_analytics_motor_hours error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_analytics_efficiency(request: Request):
    """GET /api/analytics/efficiency — ефективність роботи."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        days = int(request.query_params.get("days", "30"))
        days = max(1, min(days, 365))

        from utils.time import now_kiev
        from database.api.maintenance import get_maintenance_stats

        now = now_kiev()
        start_dt = now - timedelta(days=days - 1)

        daily = _build_daily_stats(start_dt, now, None)

        total_hours_avail = days * 24
        work_hours = sum(d["work_hours"] for d in daily)
        outage_hours = float(sum(d["outage_hours"] for d in daily))
        maintenance_hours = 0.0
        # Guard against NaN/Inf from summing DB values
        if not math.isfinite(work_hours):
            work_hours = 0.0
        if not math.isfinite(outage_hours):
            outage_hours = 0.0
        idle_hours = max(0.0, total_hours_avail - work_hours - outage_hours - maintenance_hours)

        # Розбивка по змінах
        shift_fuel: dict = {"m": 0.0, "d": 0.0, "e": 0.0, "x": 0.0}
        shift_hours: dict = {"m": 0.0, "d": 0.0, "e": 0.0, "x": 0.0}

        from database.api.logs import get_logs_for_period

        start_str = start_dt.strftime("%Y-%m-%d")
        end_str = now.strftime("%Y-%m-%d")
        logs = get_logs_for_period(start_str, end_str)
        import config as _cfg

        pending: dict = {}
        for row in logs:
            event_type, ts_str, *_ = row
            try:
                ts = datetime.strptime(ts_str[:19], "%Y-%m-%d %H:%M:%S")
            except Exception:
                continue
            if event_type in ("m_start", "d_start", "e_start", "x_start"):
                shift_key = event_type[0]
                pending[shift_key] = ts
            elif event_type in ("m_end", "d_end", "e_end", "x_end"):
                shift_key = event_type[0]
                start_ts = pending.pop(shift_key, None)
                if start_ts:
                    h = (ts - start_ts).total_seconds() / 3600.0
                    if 0 < h < 24:
                        fuel_rate = float(getattr(_cfg, "FUEL_CONSUMPTION", 5.0) or 5.0)
                        shift_hours[shift_key] = shift_hours.get(shift_key, 0.0) + h
                        shift_fuel[shift_key] = shift_fuel.get(shift_key, 0.0) + h * fuel_rate

        return {
            "pie": {
                "work_hours": _safe_round(work_hours),
                "idle_hours": _safe_round(idle_hours),
                "outage_hours": _safe_round(outage_hours),
                "maintenance_hours": _safe_round(maintenance_hours),
            },
            "shifts": {
                shift: {
                    "hours": _safe_round(shift_hours.get(shift, 0.0)),
                    "fuel_consumed": _safe_round(shift_fuel.get(shift, 0.0)),
                }
                for shift in ("m", "d", "e", "x")
            },
        }
    except Exception as e:
        logger.exception("api_analytics_efficiency error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_analytics_calendar(request: Request):
    """GET /api/analytics/calendar — календар відключень (місяць)."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from utils.time import now_kiev
        from database.api.schedule import get_schedule
        import calendar

        now = now_kiev()
        month_str = request.query_params.get("month") or now.strftime("%Y-%m")
        year, month = int(month_str[:4]), int(month_str[5:7])
        _, num_days = calendar.monthrange(year, month)

        result = []
        for day in range(1, num_days + 1):
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            sched = get_schedule(date_str)
            outage_h = sum(1 for v in sched.values() if v == 1)
            result.append({"date": date_str, "outage_hours": outage_h, "schedule": sched})

        return {"month": month_str, "days": result}
    except Exception as e:
        logger.exception("api_analytics_calendar error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Task 10: Trends API
# ---------------------------------------------------------------------------


async def api_analytics_trends(request: Request):
    """GET /api/analytics/trends — тренди та автоматичні інсайти."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        days = int(request.query_params.get("days", "30"))
        days = max(7, min(days, 365))

        from utils.time import now_kiev

        now = now_kiev()
        start_dt = now - timedelta(days=days - 1)

        daily = _build_daily_stats(start_dt, now, None)

        insights = []

        # Тренд витрати палива
        if len(daily) >= 14:
            first_half = daily[: len(daily) // 2]
            second_half = daily[len(daily) // 2 :]
            avg1 = sum(d["fuel_consumed"] for d in first_half) / len(first_half)
            avg2 = sum(d["fuel_consumed"] for d in second_half) / len(second_half)
            if avg1 > 0:
                change = (avg2 - avg1) / avg1 * 100
                if abs(change) >= 5:
                    direction = "зросла" if change > 0 else "знизилась"
                    insights.append(
                        {
                            "type": "fuel_trend",
                            "icon": "📈" if change > 0 else "📉",
                            "text": f"Витрата {direction} на {abs(change):.0f}% за {days} днів",
                            "severity": "warning" if change > 15 else "info",
                        }
                    )

        # День тижня з найбільшою кількістю відключень
        weekday_outage: dict = {}
        weekday_names = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця", "Субота", "Неділя"]
        for d in daily:
            dt = datetime.strptime(d["date"], "%Y-%m-%d")
            wd = dt.weekday()
            weekday_outage[wd] = weekday_outage.get(wd, [])
            weekday_outage[wd].append(d["outage_hours"])
        if weekday_outage:
            avg_by_day = {wd: sum(v) / len(v) for wd, v in weekday_outage.items()}
            worst_day = max(avg_by_day, key=avg_by_day.get)  # type: ignore[arg-type]
            if avg_by_day[worst_day] > 2:
                insights.append(
                    {
                        "type": "outage_weekday",
                        "icon": "📅",
                        "text": f"Найбільше відключень у {weekday_names[worst_day]} (сер. {avg_by_day[worst_day]:.1f} год)",
                        "severity": "info",
                    }
                )

        # Порівняння генераторів
        main_daily = _build_daily_stats(start_dt, now, "main")
        emergency_daily = _build_daily_stats(start_dt, now, "emergency")
        main_hours = sum(d["work_hours"] for d in main_daily)
        emerg_hours = sum(d["work_hours"] for d in emergency_daily)
        total_gen = main_hours + emerg_hours
        if total_gen > 0 and emerg_hours > 0:
            emerg_pct = emerg_hours / total_gen * 100
            if emerg_pct > 10:
                insights.append(
                    {
                        "type": "emergency_usage",
                        "icon": "⚠️",
                        "text": f"Аварійний генератор використовується {emerg_pct:.0f}% часу",
                        "severity": "warning" if emerg_pct > 30 else "info",
                    }
                )

        # Аномалії
        from ml_models import get_anomaly_detector

        anomaly_det = get_anomaly_detector()
        anomaly_det.train(daily)
        anomalies_found = []
        for d in daily[-7:]:  # перевіряємо останній тиждень
            res = anomaly_det.detect(d)
            if res["is_anomaly"]:
                anomalies_found.append(f"{d['date']}: {res['reason']}")
        if anomalies_found:
            insights.append(
                {
                    "type": "anomaly",
                    "icon": "🔴",
                    "text": f"Виявлено аномалії: {', '.join(anomalies_found[:3])}",
                    "severity": "critical",
                }
            )

        return {
            "period_days": days,
            "insights": insights,
        }
    except Exception as e:
        logger.exception("api_analytics_trends error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Task 11: Forecast API
# ---------------------------------------------------------------------------


async def api_analytics_forecast(request: Request):
    """GET /api/analytics/forecast — ML-прогноз витрати палива."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    try:
        from utils.time import now_kiev
        from database.api.generator import get_generator_stats

        now = now_kiev()
        # Тренуємо на останніх 60 днях
        start_dt = now - timedelta(days=60)
        daily = _build_daily_stats(start_dt, now, None)

        from ml_models import get_fuel_forecast

        forecast_obj = get_fuel_forecast()
        ok = forecast_obj.train(daily)

        avg_outage = sum(d["outage_hours"] for d in daily) / len(daily) if daily else 4.0
        forecast = forecast_obj.predict(7, avg_outage)

        total_forecast_fuel = sum(f["predicted_fuel"] for f in forecast)

        # ТО
        main_stats = get_generator_stats("main")
        oil_hours = main_stats.get("last_oil_change", 0)
        oil_interval = getattr(config, "OIL_CHANGE_INTERVAL", 250)
        spark_interval = getattr(config, "SPARK_CHANGE_INTERVAL", 500)
        oil_remaining = max(0, oil_interval - oil_hours)
        spark_remaining = max(0, spark_interval - main_stats.get("last_spark_change", 0))

        avg_daily_hours = sum(d["work_hours"] for d in daily[-7:]) / 7 if len(daily) >= 7 else 2.0
        days_to_oil = round(oil_remaining / avg_daily_hours, 0) if avg_daily_hours > 0 else 0
        days_to_spark = round(spark_remaining / avg_daily_hours, 0) if avg_daily_hours > 0 else 0

        return {
            "model_trained": ok,
            "forecast_days": 7,
            "daily_forecast": forecast,
            "total_forecast_fuel": round(total_forecast_fuel, 1),
            "maintenance": {
                "oil_remaining_hours": round(oil_remaining, 1),
                "spark_remaining_hours": round(spark_remaining, 1),
                "days_to_oil_change": int(days_to_oil),
                "days_to_spark_change": int(days_to_spark),
            },
        }
    except Exception as e:
        logger.exception("api_analytics_forecast error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Enhanced Excel Report API (replaces PDF endpoint)
# ---------------------------------------------------------------------------


async def api_report_excel_v2(request: Request):
    """GET /api/report/excel/v2?type=quick&days=30&generator=main — enhanced Excel report."""
    user = _extract_user(request)
    if not user:
        return JSONResponse(content={"error": "Не авторизовано"}, status_code=401)
    if not _EXCEL_RPT_AVAILABLE:
        return JSONResponse(content={"error": "Модуль openpyxl не встановлено"}, status_code=500)

    try:
        report_type = request.query_params.get("type", "quick")
        valid_types = ("quick", "detailed", "personnel", "technical", "financial")
        if report_type not in valid_types:
            return JSONResponse(content={"error": "Невірний тип звіту"}, status_code=400)

        days = int(request.query_params.get("days", "30"))
        days = max(1, min(days, 365))
        gen_id = (request.query_params.get("generator") or "").strip().lower() or None
        if gen_id not in ("main", "emergency"):
            gen_id = None

        now = datetime.now(config.KYIV)
        excel_bytes = generate_excel_report(report_type, days, gen_id)

        filename = f"generator_report_{report_type}_{now.strftime('%Y%m%d')}.xlsx"
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ValueError as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
    except Exception as e:
        logger.exception("api_report_excel_v2 error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


@asynccontextmanager
async def _webapp_lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    """Lifespan for the webapp FastAPI application.

    Startup: initialise the database so the app works when started
    standalone (``uvicorn webapp_server:create_app()``).
    Shutdown: close the PostgreSQL connection pool gracefully.
    """
    try:
        db_models.init_db()
        logger.info("✅ [lifespan] DB initialised")
    except Exception:
        logger.exception("❌ [lifespan] DB init failed")
        raise
    yield
    try:
        from database.models import close_postgres_pool

        close_postgres_pool()
        logger.info("✅ [lifespan] DB pool closed")
    except Exception:
        logger.warning("⚠️  [lifespan] Error closing DB pool (ignored)")


def create_app() -> FastAPI:
    """Створює FastAPI-додаток з API та статичними файлами."""
    from webapp.middleware.rate_limit import RateLimitMiddleware

    app = FastAPI(
        title="Generator Bot WebApp",
        version=BUILD_VERSION,
        docs_url="/api/docs",
        redoc_url="/api/redoc",
        lifespan=_webapp_lifespan,
    )

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    app.add_middleware(RateLimitMiddleware)

    # API маршрути (читання)
    app.add_api_route("/api/status", api_status, methods=["GET"])
    app.add_api_route("/api/schedule", api_schedule, methods=["GET"])
    app.add_api_route("/api/schedule/week", api_schedule_week, methods=["GET"])
    app.add_api_route("/api/events", api_events, methods=["GET"])
    app.add_api_route("/api/maintenance", api_maintenance, methods=["GET"])

    app.add_api_route("/api/user/role", api_user_role, methods=["GET"])
    app.add_api_route("/api/drivers", api_drivers, methods=["GET"])
    app.add_api_route("/api/generators", api_generators, methods=["GET"])
    app.add_api_route("/api/personnel/me", api_personnel_me, methods=["GET"])
    app.add_api_route("/api/report/excel", api_report_excel, methods=["GET"])

    app.add_api_route("/api/action/start", api_action_start, methods=["POST"])
    app.add_api_route("/api/action/stop", api_action_stop, methods=["POST"])
    app.add_api_route("/api/action/refill", api_action_refill, methods=["POST"])
    app.add_api_route("/api/schedule/toggle", api_schedule_toggle, methods=["POST"])
    app.add_api_route("/api/generator/switch", api_generator_switch, methods=["POST"])
    app.add_api_route("/api/maintenance/perform", api_maintenance_perform, methods=["POST"])
    app.add_api_route("/api/maintenance/set-hours", api_maintenance_set_hours, methods=["POST"])
    app.add_api_route("/api/fuel/set", api_fuel_set, methods=["POST"])

    app.add_api_route("/api/admin/drivers", api_admin_drivers_list, methods=["GET"])
    app.add_api_route("/api/admin/drivers", api_admin_drivers_add, methods=["POST"])
    app.add_api_route("/api/admin/drivers", api_admin_drivers_delete, methods=["DELETE"])
    app.add_api_route("/api/admin/personnel", api_admin_personnel_list, methods=["GET"])
    app.add_api_route("/api/admin/personnel", api_admin_personnel_add, methods=["POST"])
    app.add_api_route("/api/admin/personnel", api_admin_personnel_delete, methods=["DELETE"])
    app.add_api_route("/api/admin/personnel/assign", api_admin_personnel_assign, methods=["POST"])
    app.add_api_route("/api/admin/sync", api_admin_sync, methods=["POST"])

    app.add_api_route("/api/admin/audit", api_admin_audit, methods=["GET"])
    app.add_api_route("/api/admin/audit/export", api_admin_audit_export, methods=["GET"])

    app.add_api_route("/api/admin/config", api_admin_config_get, methods=["GET"])
    app.add_api_route("/api/admin/config/generator", api_admin_config_set_generator, methods=["POST"])
    app.add_api_route("/api/admin/config/global", api_admin_config_set_global, methods=["POST"])
    app.add_api_route("/api/admin/config/history", api_admin_config_history, methods=["GET"])

    app.add_api_route("/api/admin/backups", api_admin_backups_list, methods=["GET"])
    app.add_api_route("/api/admin/backup", api_admin_backup_create, methods=["POST"])
    app.add_api_route("/api/admin/backup/download/{filename}", api_admin_backup_download, methods=["GET"])

    # Task 5: Notification endpoints
    app.add_api_route("/api/notifications/preferences", api_notifications_get, methods=["GET"])
    app.add_api_route("/api/notifications/preferences", api_notifications_set, methods=["POST"])
    app.add_api_route("/api/notifications/test", api_notifications_test, methods=["POST"])

    # Task 6: Fuel orders endpoints
    app.add_api_route("/api/fuel/orders", api_fuel_orders_list, methods=["GET"])
    app.add_api_route("/api/fuel/orders", api_fuel_orders_create, methods=["POST"])
    app.add_api_route("/api/fuel/orders/update", api_fuel_orders_update, methods=["POST"])

    # Task 8: Shift schedule endpoints
    app.add_api_route("/api/shifts/schedule", api_shifts_get, methods=["GET"])
    app.add_api_route("/api/shifts/schedule", api_shifts_set, methods=["POST"])
    app.add_api_route("/api/shifts/auto", api_shifts_auto, methods=["POST"])
    app.add_api_route("/api/shifts/analytics", api_shifts_analytics, methods=["GET"])

    # Tasks 9-12: Analytics, Trends, Forecast, Enhanced Excel report endpoints
    app.add_api_route("/api/analytics/kpi", api_analytics_kpi, methods=["GET"])
    app.add_api_route("/api/analytics/fuel-timeline", api_analytics_fuel_timeline, methods=["GET"])
    app.add_api_route("/api/analytics/motor-hours", api_analytics_motor_hours, methods=["GET"])
    app.add_api_route("/api/analytics/efficiency", api_analytics_efficiency, methods=["GET"])
    app.add_api_route("/api/analytics/calendar", api_analytics_calendar, methods=["GET"])
    app.add_api_route("/api/analytics/trends", api_analytics_trends, methods=["GET"])
    app.add_api_route("/api/analytics/forecast", api_analytics_forecast, methods=["GET"])
    app.add_api_route("/api/report/excel/v2", api_report_excel_v2, methods=["GET"])

    # Static files (CSS, JS) and app routes
    if _webapp_dir.is_dir():
        # Named routes first (before mounts)
        app.add_api_route("/service-worker.js", sw_handler, methods=["GET"])
        app.add_api_route("/", index_handler, methods=["GET"])
        app.add_api_route("/block.html", block_handler, methods=["GET"])

        # Static directories
        css_dir = _webapp_dir / "css"
        js_dir = _webapp_dir / "js"
        if css_dir.is_dir():
            app.mount("/css", StaticFiles(directory=str(css_dir)), name="css")
        if js_dir.is_dir():
            app.mount("/js", StaticFiles(directory=str(js_dir)), name="js")

    return app


def main():
    """Точка входу — запуск веб-сервера."""
    logging.basicConfig(
        level=getattr(logging, config.LOG_LEVEL, logging.INFO),
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )

    # Ініціалізація БД
    logger.info("🔧 Ініціалізація бази даних...")
    logger.info(f"🏷️  BUILD_VERSION: {BUILD_VERSION}")
    db_models.init_db()

    port = int(os.getenv("WEBAPP_PORT", "8080"))
    host = os.getenv("WEBAPP_HOST", "0.0.0.0")

    app = create_app()

    import uvicorn

    logger.info(f"🌐 Mini App сервер запускається на http://{host}:{port}")
    uvicorn.run(app, host=host, port=port, log_level="info")


if __name__ == "__main__":
    main()
