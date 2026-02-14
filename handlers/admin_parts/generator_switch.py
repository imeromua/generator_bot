"""Generator management handler.

Functionality:
- Switching between main and emergency generators (only when status OFF)
- Statistics display for each generator
- Emergency generator Excel report export
- Reports archive
"""

import logging
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional

from aiogram import F, Router, types
from aiogram.fsm.context import FSMContext

import config
import database.db_api as db
from keyboards.builders import InlineKeyboardBuilder

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    MergedCell = None
    get_column_letter = None

router = Router()
logger = logging.getLogger(__name__)

# Maximum reports in archive
MAX_ARCHIVE_SIZE: int = 10

# NOTE: avoid backslash escapes inside f-string expressions (coverage parser / Python syntax)
MARK_MAIN = "🔹"
MARK_EMERGENCY = "🔸"
MARK_INACTIVE = "▫️"


def _generator_keyboard(last_export_time: Optional[str] = None) -> types.InlineKeyboardMarkup:
    """Клавіатура для управління генераторами.

    Args:
        last_export_time: Last export time (e.g. "21:36")

    Returns:
        Inline keyboard markup
    """
    builder = InlineKeyboardBuilder()

    active_gen = db.get_active_generator()

    if active_gen == "main":
        builder.button(text="⚡ Перемкнути на АВАРІЙНИЙ", callback_data="gen_switch_emergency")
    else:
        builder.button(text="🔋 Перемкнути на ОСНОВНИЙ", callback_data="gen_switch_main")

    builder.button(text="📊 Статистика генераторів", callback_data="gen_stats")

    if active_gen == "emergency":
        export_text = "📅 Експорт звіту (Excel)"
        if last_export_time:
            export_text += f" • {last_export_time}"
        builder.button(text=export_text, callback_data="gen_export_excel")
        builder.button(text="📂 Архів звітів", callback_data="gen_archive")

    builder.button(text="🔙 Назад", callback_data="admin_home")
    builder.adjust(1)

    return builder.as_markup()


def _document_keyboard() -> types.InlineKeyboardMarkup:
    """Клавіатура для документа звіту.

    Returns:
        Inline keyboard markup
    """
    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 Повернутися до меню", callback_data="generator_switch")
    builder.adjust(1)
    return builder.as_markup()


@router.callback_query(F.data == "generator_switch")
async def gen_switch_menu(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Головне меню перемикання генераторів.

    Important: supports "single window" concept like admin panel.
    If menu opens from document (Excel report), the document message
    is deleted and new text message is created, which becomes new
    tracked UI (db.set_ui_message).

    Args:
        cb: Callback query
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    await state.clear()

    active_gen = db.get_active_generator()
    gen_name = db.get_generator_name(active_gen)

    st = db.get_state()
    status = st.get("status", "OFF")

    # Information about active generator
    if active_gen == "main":
        stats = db.get_generator_stats("main")
        info_text = (
            f"⚡ <b>Управління генераторами</b>\n"
            f"──────────────────\n"
            f"🔋 Активний: {gen_name}\n"
            f"📊 Статус: {'🟢 ВИМКНЕНО' if status == 'OFF' else '🟩 ПРАЦЮЄ'}\n\n"
            f"📈 Основний генератор:\n"
            f"  ⏱ Мотогодини: {stats['total_hours']:.1f} год\n"
            f"  🛢 Мастило: {stats['last_oil_change']:.1f} год\n"
            f"  🕯 Свічки: {stats['last_spark_change']:.1f} год\n\n"
            f"💡 Для перемикання генератор має бути вимкнений (OFF)"
        )
    else:
        stats = db.get_generator_stats("emergency")
        info_text = (
            f"⚡ <b>Управління генераторами</b>\n"
            f"──────────────────\n"
            f"⚠️ Активний: {gen_name}\n"
            f"📊 Статус: {'🟢 ВИМКНЕНО' if status == 'OFF' else '🟩 ПРАЦЮЄ'}\n\n"
            f"📈 Аварійний генератор:\n"
            f"  ⏱ Мотогодини: {stats['total_hours']:.1f} год\n"
            f"  🛢 Мастило: {stats['last_oil_change']:.1f} год\n"
            f"  🕯 Свічки: {stats['last_spark_change']:.1f} год\n\n"
            f"⚠️ <b>УВАГА!</b> Аварійний генератор НЕ синхронізується з Google Sheets.\n"
            f"Звіт можна експортувати в Excel.\n\n"
            f"💡 Для перемикання генератор має бути вимкнений (OFF)"
        )

    # Check if it's text message
    if cb.message.text:
        # If text - edit within single-window
        await cb.message.edit_text(info_text, reply_markup=_generator_keyboard())
        msg_to_track = cb.message
    else:
        # If document - delete and create new text message
        await cb.message.delete()
        msg_to_track = await cb.message.answer(info_text, reply_markup=_generator_keyboard())

    # Update tracked UI so _is_outdated_ui() in admin panel doesn't
    # consider this old message. This way generator menu fits into
    # "single window" concept together with admin_home.
    try:
        db.set_ui_message(int(cb.from_user.id), int(msg_to_track.chat.id), int(msg_to_track.message_id))
    except Exception:
        pass


@router.callback_query(F.data.startswith("gen_switch_"))
async def gen_switch_action(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Перемикання генератора.

    Args:
        cb: Callback query
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    target = "main" if cb.data == "gen_switch_main" else "emergency"

    # Get admin name
    user_info = db.get_user(int(cb.from_user.id))
    admin_name = user_info[1] if user_info else cb.from_user.full_name

    success, message = db.switch_generator(target, admin_name)

    if success:
        await cb.answer(message, show_alert=True)
        # Update menu
        await gen_switch_menu(cb, state)
    else:
        await cb.answer(message, show_alert=True)


@router.callback_query(F.data == "gen_stats")
async def gen_stats_view(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Показ статистики обох генераторів.

    Args:
        cb: Callback query
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    main_stats = db.get_generator_stats("main")
    emerg_stats = db.get_generator_stats("emergency")

    active_gen = db.get_active_generator()

    # Get fuel balance as float
    try:
        current_fuel = float(db.get_state_value("current_fuel", "0.0"))
    except (ValueError, TypeError):
        current_fuel = 0.0

    text = (
        f"📊 <b>Статистика генераторів</b>\n"
        f"──────────────────\n\n"
        f"{MARK_MAIN if active_gen == 'main' else MARK_INACTIVE} <b>Основний генератор</b>\n"
        f"  ⏱ Мотогодини: {main_stats['total_hours']:.1f} год\n"
        f"  🛢 Від заміни мастила: {main_stats['last_oil_change']:.1f} год\n"
        f"  🕯 Від заміни свічок: {main_stats['last_spark_change']:.1f} год\n\n"
        f"{MARK_EMERGENCY if active_gen == 'emergency' else MARK_INACTIVE} <b>Аварійний генератор</b>\n"
        f"  ⏱ Мотогодини: {emerg_stats['total_hours']:.1f} год\n"
        f"  🛢 Від заміни мастила: {emerg_stats['last_oil_change']:.1f} год\n"
        f"  🕯 Від заміни свічок: {emerg_stats['last_spark_change']:.1f} год\n\n"
        f"{MARK_MAIN if active_gen == 'main' else MARK_EMERGENCY} - Активний генератор\n"
        f"──────────────────\n"
        f"💡 Спільні параметри:\n"
        f"  ⛽ Залишок палива: {current_fuel:.1f} л\n"
        f"  👥 Персонал та водії\n"
    )

    builder = InlineKeyboardBuilder()
    builder.button(text="🔙 Назад", callback_data="generator_switch")

    await cb.message.edit_text(text, reply_markup=builder.as_markup())


@router.callback_query(F.data == "gen_archive")
async def gen_archive_view(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Показ архіву звітів.

    IMPORTANT: callback_data for archive buttons must be short (<64 bytes),
    so we CAN'T embed file_id there (they're long). Instead we use
    index in archive list.

    Args:
        cb: Callback query
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    # Get reports archive from state
    archive_json = db.get_state_value("reports_archive", "[]")
    try:
        import json

        archive = json.loads(archive_json)
    except Exception:
        archive = []

    if not archive:
        await cb.answer("📂 Архів звітів порожній", show_alert=True)
        return

    text = (
        f"📂 <b>Архів звітів</b>\n"
        f"──────────────────\n"
        f"Останні {len(archive)} звітів:\n\n"
    )

    builder = InlineKeyboardBuilder()

    # Go from newest to oldest, but in callback_data put
    # real index in archive list (short int, valid for Telegram).
    indexed_archive = list(enumerate(archive))
    for shown_idx, (real_index, report) in enumerate(reversed(indexed_archive), 1):
        timestamp = report.get("timestamp", "")

        # Format date
        try:
            dt = datetime.fromisoformat(timestamp)
            date_str = dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            date_str = timestamp

        text += f"{shown_idx}. 📊 {date_str}\n"
        builder.button(text=f"📅 Звіт #{shown_idx}", callback_data=f"gen_get_report_{real_index}")

    builder.button(text="🔙 Назад", callback_data="generator_switch")
    builder.adjust(1)

    await cb.message.edit_text(text, reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("gen_get_report_"))
async def gen_get_report(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Get report from archive by index (to keep callback_data short).

    Args:
        cb: Callback query with format "gen_get_report_INDEX"
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    index_str = cb.data.replace("gen_get_report_", "")
    try:
        idx = int(index_str)
    except ValueError:
        return await cb.answer("❌ Невірний індекс звіту", show_alert=True)

    # Read actual archive from DB
    archive_json = db.get_state_value("reports_archive", "[]")
    try:
        import json

        archive = json.loads(archive_json)
    except Exception:
        archive = []

    if not archive or idx < 0 or idx >= len(archive):
        return await cb.answer("❌ Звіт не знайдено в архіві", show_alert=True)

    report = archive[idx]
    file_id = report.get("file_id")

    if not file_id:
        return await cb.answer("❌ У записі відсутній file_id", show_alert=True)

    await cb.answer("📤 Відправляю звіт...")

    try:
        # Delete old menu (single window)
        await cb.message.delete()

        # Send document with buttons
        await cb.message.answer_document(
            document=file_id,
            caption="📊 Звіт аварійного генератора з архіву",
            reply_markup=_document_keyboard(),
        )
    except Exception as e:
        logger.error(f"Помилка отримання звіту з архіву: {e}")
        await cb.answer(f"❌ Помилка: {e}", show_alert=True)


@router.callback_query(F.data == "gen_export_excel")
async def gen_export_excel(cb: types.CallbackQuery, state: FSMContext) -> None:
    """Експорт звіту аварійного генератора в Excel.

    Args:
        cb: Callback query
        state: FSM context
    """
    if cb.from_user.id not in config.ADMIN_IDS:
        return await cb.answer("⛔ Тільки для адмінів", show_alert=True)

    if not EXCEL_AVAILABLE:
        return await cb.answer(
            "❌ Модуль openpyxl не встановлено.\nВиконайте: pip install openpyxl",
            show_alert=True,
        )

    await cb.answer("📤 Генерую звіт...")

    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Аварійний генератор"

        # Styles
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Title
        ws["A1"] = "Звіт: Аварійний генератор"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # General information
        stats = db.get_generator_stats("emergency")
        st = db.get_state()

        ws["A3"] = "Мотогодини:"
        ws["B3"] = f"{stats['total_hours']:.2f} год"

        ws["A4"] = "Від заміни мастила:"
        ws["B4"] = f"{stats['last_oil_change']:.2f} год"

        ws["A5"] = "Від заміни свічок:"
        ws["B5"] = f"{stats['last_spark_change']:.2f} год"

        ws["A6"] = "Поточний залишок палива:"
        ws["B6"] = f"{float(st.get('current_fuel', 0.0)):.2f} л"

        # Events table
        ws["A8"] = "Журнал подій (аварійний генератор)"
        ws["A8"].font = Font(bold=True, size=12)

        # Table headers
        headers = ["Дата/Час", "Подія", "Користувач", "Значення", "Водій"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get all emergency generator logs for last 30 days
        end_date = datetime.now(config.KYIV).strftime("%Y-%m-%d")
        start_date = (datetime.now(config.KYIV) - timedelta(days=30)).strftime("%Y-%m-%d")

        logs = db.get_logs_for_period(start_date, end_date, generator_id="emergency")

        # Fill table
        row = 10
        event_names = {
            "m_start": "🌅 Зміна 1 (початок)",
            "m_end": "🌅 Зміна 1 (кінець)",
            "d_start": "☀️ Зміна 2 (початок)",
            "d_end": "☀️ Зміна 2 (кінець)",
            "e_start": "🌙 Зміна 3 (початок)",
            "e_end": "🌙 Зміна 3 (кінець)",
            "x_start": "⚡ Екстра (початок)",
            "x_end": "⚡ Екстра (кінець)",
            "refill": "⛽ Заправка",
            "correction": "🔧 Корекція палива",
        }

        for log in logs:
            event_type, timestamp, user_name, value, driver_name, receipt_number, _ = log

            ws.cell(row=row, column=1).value = timestamp
            ws.cell(row=row, column=2).value = event_names.get(event_type, event_type)
            ws.cell(row=row, column=3).value = user_name or "-"
            ws.cell(row=row, column=4).value = value or "-"
            ws.cell(row=row, column=5).value = driver_name or "-"

            row += 1

        # Auto-width columns (safely handle MergedCell)
        for col_idx in range(1, 6):  # Columns A-E
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if MergedCell and isinstance(cell, MergedCell):
                    continue
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Send file
        filename = f"emergency_generator_{datetime.now(config.KYIV).strftime('%Y%m%d_%H%M')}.xlsx"
        file = types.BufferedInputFile(buffer.read(), filename=filename)

        # Delete old menu (single window concept)
        await cb.message.delete()

        # Send document with buttons
        sent_msg = await cb.message.answer_document(
            document=file,
            caption=f"📊 Звіт аварійного генератора\n🗓 Період: {start_date} — {end_date}\n📁 {len(logs)} подій",
            reply_markup=_document_keyboard(),
        )

        # Save to archive
        import json

        archive_json = db.get_state_value("reports_archive", "[]")
        try:
            archive = json.loads(archive_json)
        except Exception:
            archive = []

        archive.append(
            {
                "file_id": sent_msg.document.file_id,
                "timestamp": datetime.now(config.KYIV).isoformat(),
                "filename": filename,
            }
        )

        if len(archive) > MAX_ARCHIVE_SIZE:
            archive = archive[-MAX_ARCHIVE_SIZE:]

        db.set_state_value("reports_archive", json.dumps(archive))

    except Exception as e:
        logger.error(f"Помилка експорту Excel: {e}", exc_info=True)
        await cb.answer(f"❌ Помилка експорту: {e}", show_alert=True)
