"""Admin management API endpoints."""
import io
import re
import logging
from datetime import datetime
from fastapi import Request
from fastapi.responses import JSONResponse, Response
import config
import database.db_api as db
from webapp.utils.validation import extract_user as _extract_user
from webapp.utils.permissions import is_admin as _is_admin
from webapp.utils.db_helpers import get_admin_info as _get_admin_info
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)
MAX_NAME_LENGTH = 100


async def api_admin_drivers_list(request: Request):
    """GET /api/admin/drivers — список водіїв (лише для адмінів)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        drivers = db.get_drivers()
        return {"drivers": list(drivers) if drivers else []}
    except Exception as e:
        logger.exception("api_admin_drivers_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_drivers_add(request: Request):
    """POST /api/admin/drivers — додати водія."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name or len(name) > MAX_NAME_LENGTH:
        return JSONResponse(content={"error": "Невірне ім'я водія (1–100 символів)"}, status_code=400)

    try:
        ok = db.add_driver(name)
        if not ok:
            return JSONResponse(content={"error": f"Водій «{name}» вже існує"}, status_code=409)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "driver_add",
            f"Додано водія «{name}»",
            target_entity=f"driver:{name}",
            new_value=name,
        )
        return {"ok": True, "message": f"Водія «{name}» додано"}
    except Exception as e:
        logger.exception("api_admin_drivers_add error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_drivers_delete(request: Request):
    """DELETE /api/admin/drivers — видалити водія."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name:
        return JSONResponse(content={"error": "Ім'я водія обов'язкове"}, status_code=400)

    try:
        ok = db.delete_driver(name)
        if not ok:
            return JSONResponse(content={"error": f"Водія «{name}» не знайдено"}, status_code=404)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "driver_delete",
            f"Видалено водія «{name}»",
            target_entity=f"driver:{name}",
            old_value=name,
        )
        return {"ok": True, "message": f"Водія «{name}» видалено"}
    except Exception as e:
        logger.exception("api_admin_drivers_delete error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_list(request: Request):
    """GET /api/admin/personnel — список персоналу (лише для адмінів)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        names = db.get_personnel_names()
        users_with_p = db.get_all_users_with_personnel()
        users_list = [{"user_id": row[0], "full_name": row[1] or "", "personnel": row[2] or ""} for row in users_with_p]
        return {"personnel": names, "users": users_list}
    except Exception as e:
        logger.exception("api_admin_personnel_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_add(request: Request):
    """POST /api/admin/personnel — додати ПІБ персоналу."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name or len(name) > MAX_NAME_LENGTH:
        return JSONResponse(content={"error": "Невірне ім'я (1–100 символів)"}, status_code=400)

    try:
        ok = db.add_personnel_name(name)
        if not ok:
            return JSONResponse(content={"error": f"Персонал «{name}» вже існує"}, status_code=409)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_add",
            f"Додано персонал «{name}»",
            target_entity=f"personnel:{name}",
            new_value=name,
        )
        return {"ok": True, "message": f"Персонал «{name}» додано"}
    except Exception as e:
        logger.exception("api_admin_personnel_add error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_delete(request: Request):
    """DELETE /api/admin/personnel — видалити ПІБ персоналу."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    name = (body.get("name") or "").strip()
    if not name:
        return JSONResponse(content={"error": "Ім'я обов'язкове"}, status_code=400)

    try:
        ok = db.delete_personnel_name(name)
        if not ok:
            return JSONResponse(content={"error": f"Персонал «{name}» не знайдено"}, status_code=404)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_delete",
            f"Видалено персонал «{name}»",
            target_entity=f"personnel:{name}",
            old_value=name,
        )
        return {"ok": True, "message": f"Персонал «{name}» видалено"}
    except Exception as e:
        logger.exception("api_admin_personnel_delete error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_personnel_assign(request: Request):
    """POST /api/admin/personnel/assign — прив'язати персонал до Telegram-користувача."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    try:
        target_user_id = int(body.get("user_id", 0))
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "Невірний user_id"}, status_code=400)

    personnel_name = (body.get("personnel") or "").strip() or None

    if not target_user_id:
        return JSONResponse(content={"error": "user_id обов'язковий"}, status_code=400)

    try:
        old_personnel = db.get_personnel_for_user(target_user_id)
        db.set_personnel_for_user(target_user_id, personnel_name)
        admin_id, admin_name = _get_admin_info(user)
        if personnel_name:
            msg = f"Прив'язано: user {target_user_id} → «{personnel_name}»"
        else:
            msg = f"Прив'язку для user {target_user_id} знято"
        db.log_admin_action(
            admin_id,
            admin_name,
            "personnel_assign",
            msg,
            target_entity=f"user:{target_user_id}",
            old_value=old_personnel,
            new_value=personnel_name,
        )
        return {"ok": True, "message": msg}
    except Exception as e:
        logger.exception("api_admin_personnel_assign error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_sync(request: Request):
    """POST /api/admin/sync — запуск синхронізації з Google Sheets (експорт)."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from services.sheets_export import full_export

        result = full_export()
        updated = result.get("updated", [])
        skipped = result.get("skipped", [])
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "export_sheets",
            f"Синхронізація з Google Sheets: {len(updated)} дн. оновлено",
            new_value={"updated": len(updated), "skipped": len(skipped)},
        )
        return {
            "ok": True,
            "message": f"Синхронізовано: {len(updated)} дн., пропущено: {len(skipped)} дн.",
            "updated": updated,
            "skipped": skipped,
        }
    except Exception as e:
        logger.exception("api_admin_sync error")
        return JSONResponse(content={"error": f"Помилка синхронізації: {e}"}, status_code=500)


async def api_admin_audit(request: Request):
    """GET /api/admin/audit — журнал дій адміністраторів.

    Query params:
        limit      (int, default 50, max 200)
        offset     (int, default 0)
        action_type (str, optional filter)
        admin_id   (int, optional filter by admin user ID)
        date_from  (str YYYY-MM-DD, optional)
        date_to    (str YYYY-MM-DD, optional)
    """
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        limit = min(int(request.query_params.get("limit", "50")), 200)
    except (TypeError, ValueError):
        limit = 50
    try:
        offset = max(int(request.query_params.get("offset", "0")), 0)
    except (TypeError, ValueError):
        offset = 0

    action_type = request.query_params.get("action_type", "").strip()
    date_from = request.query_params.get("date_from", "").strip()
    date_to = request.query_params.get("date_to", "").strip()
    try:
        admin_filter = int(request.query_params.get("admin_id", "0"))
    except (TypeError, ValueError):
        admin_filter = 0

    try:
        rows = db.get_audit_logs(
            limit=limit,
            offset=offset,
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )
        total = db.count_audit_logs(
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )
        entries = [
            {
                "id": r[0],
                "timestamp": r[1],
                "admin_user_id": r[2],
                "admin_name": r[3],
                "action_type": r[4],
                "action_description": r[5],
                "target_entity": r[6],
                "old_value": r[7],
                "new_value": r[8],
                "success": bool(r[9]),
            }
            for r in rows
        ]
        return {
            "entries": entries,
            "total": total,
            "limit": limit,
            "offset": offset,
        }
    except Exception as e:
        logger.exception("api_admin_audit error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_audit_export(request: Request):
    """GET /api/admin/audit/export — експорт журналу дій у Excel."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    if not EXCEL_AVAILABLE:
        return JSONResponse(content={"error": "Модуль openpyxl не встановлено"}, status_code=500)

    action_type = request.query_params.get("action_type", "").strip()
    date_from = request.query_params.get("date_from", "").strip()
    date_to = request.query_params.get("date_to", "").strip()
    try:
        admin_filter = int(request.query_params.get("admin_id", "0"))
    except (TypeError, ValueError):
        admin_filter = 0

    try:
        rows = db.get_audit_logs(
            limit=5000,
            offset=0,
            action_type=action_type,
            admin_user_id=admin_filter,
            date_from=date_from,
            date_to=date_to,
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Журнал дій"

        headers = [
            "#",
            "Час",
            "Адмін ID",
            "Адмін",
            "Тип дії",
            "Опис",
            "Об'єкт",
            "Старе значення",
            "Нове значення",
            "Успішно",
        ]
        header_fill = PatternFill(start_color="2481CC", end_color="2481CC", fill_type="solid")
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

        col_widths = [6, 20, 12, 20, 18, 40, 25, 20, 20, 10]
        from openpyxl.utils import get_column_letter as _gcl

        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[_gcl(ci)].width = w

        for ri, r in enumerate(rows, start=2):
            ws.cell(row=ri, column=1, value=r[0])
            ws.cell(row=ri, column=2, value=r[1])
            ws.cell(row=ri, column=3, value=r[2])
            ws.cell(row=ri, column=4, value=r[3] or "")
            ws.cell(row=ri, column=5, value=r[4] or "")
            ws.cell(row=ri, column=6, value=r[5] or "")
            ws.cell(row=ri, column=7, value=r[6] or "")
            ws.cell(row=ri, column=8, value=str(r[7]) if isinstance(r[7], (dict, list)) else (r[7] or ""))
            ws.cell(row=ri, column=9, value=str(r[8]) if isinstance(r[8], (dict, list)) else (r[8] or ""))
            ws.cell(row=ri, column=10, value="✅" if r[9] else "❌")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now = datetime.now(config.KYIV)
        filename = f"audit_log_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("api_admin_audit_export error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_get(request: Request):
    """GET /api/admin/config — поточні налаштування генераторів та глобальні."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        from database.api.config import get_generator_config, get_global_config

        def _build_param_response(cfg: dict, params: tuple) -> dict:
            return {
                p: {
                    "value": cfg[p]["value"] if p in cfg else None,
                    "last_updated": cfg[p]["last_updated"] if p in cfg else "",
                    "updated_by": cfg[p]["updated_by"] if p in cfg else "",
                }
                for p in params
            }

        main_cfg = get_generator_config("main")
        emerg_cfg = get_generator_config("emergency")
        global_cfg = get_global_config()
        gen_params = ("fuel_consumption_rate",)

        return {
            "generators": {
                "main": _build_param_response(main_cfg, gen_params),
                "emergency": _build_param_response(emerg_cfg, gen_params),
            },
            "global": _build_param_response(global_cfg, ("fuel_price",)),
        }
    except Exception as e:
        logger.exception("api_admin_config_get error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_set_generator(request: Request):
    """POST /api/admin/config/generator — змінити параметр генератора."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    generator_id = str(body.get("generator_id", "")).strip()
    param_name = str(body.get("param_name", "")).strip()
    value = body.get("value")

    if not generator_id or not param_name or value is None:
        return JSONResponse(content={"error": "generator_id, param_name та value обов'язкові"}, status_code=400)

    try:
        value = float(value)
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "value має бути числом"}, status_code=400)

    try:
        from database.api.config import (
            set_generator_param,
            get_generator_param,
            VALID_GENERATOR_IDS,
            VALID_GENERATOR_PARAMS,
        )

        if generator_id not in VALID_GENERATOR_IDS:
            return JSONResponse(
                content={"error": f"generator_id має бути одним із: {', '.join(VALID_GENERATOR_IDS)}"}, status_code=400
            )
        if param_name not in VALID_GENERATOR_PARAMS:
            return JSONResponse(
                content={"error": f"param_name має бути одним із: {', '.join(VALID_GENERATOR_PARAMS)}"}, status_code=400
            )

        admin_id, admin_name = _get_admin_info(user)
        old_value = get_generator_param(generator_id, param_name)

        ok = set_generator_param(generator_id, param_name, value, admin_id, admin_name)
        if not ok:
            return JSONResponse(content={"error": "Не вдалося зберегти налаштування"}, status_code=500)

        db.log_admin_action(
            admin_id,
            admin_name,
            "config_generator_set",
            f"Змінено {param_name} для {generator_id}: {old_value} → {value}",
            target_entity=f"generator:{generator_id}",
            old_value=old_value,
            new_value=value,
        )
        return {"ok": True, "message": "Налаштування збережено", "old_value": old_value, "new_value": value}
    except ValueError as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
    except Exception as e:
        logger.exception("api_admin_config_set_generator error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_set_global(request: Request):
    """POST /api/admin/config/global — змінити глобальний параметр."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(content={"error": "Невірний JSON"}, status_code=400)

    param_name = str(body.get("param_name", "")).strip()
    value = body.get("value")

    if not param_name or value is None:
        return JSONResponse(content={"error": "param_name та value обов'язкові"}, status_code=400)

    try:
        value = float(value)
    except (TypeError, ValueError):
        return JSONResponse(content={"error": "value має бути числом"}, status_code=400)

    try:
        from database.api.config import set_global_param, get_global_param, VALID_GLOBAL_PARAMS

        if param_name not in VALID_GLOBAL_PARAMS:
            return JSONResponse(
                content={"error": f"param_name має бути одним із: {', '.join(VALID_GLOBAL_PARAMS)}"}, status_code=400
            )

        admin_id, admin_name = _get_admin_info(user)
        old_value = get_global_param(param_name)

        ok = set_global_param(param_name, value, admin_id, admin_name)
        if not ok:
            return JSONResponse(content={"error": "Не вдалося зберегти налаштування"}, status_code=500)

        db.log_admin_action(
            admin_id,
            admin_name,
            "config_global_set",
            f"Змінено {param_name}: {old_value} → {value}",
            target_entity=f"global:{param_name}",
            old_value=old_value,
            new_value=value,
        )
        return {"ok": True, "message": "Налаштування збережено", "old_value": old_value, "new_value": value}
    except ValueError as e:
        return JSONResponse(content={"error": str(e)}, status_code=400)
    except Exception as e:
        logger.exception("api_admin_config_set_global error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_config_history(request: Request):
    """GET /api/admin/config/history?limit=20 — історія змін налаштувань."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)
    try:
        limit = int(request.query_params.get("limit", "20"))
        limit = max(1, min(limit, 100))
        offset = int(request.query_params.get("offset", "0"))
        offset = max(0, offset)

        from database.api.config import get_config_history

        history = get_config_history(limit=limit, offset=offset)
        return {"history": history}
    except Exception as e:
        logger.exception("api_admin_config_history error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_backups_list(request: Request):
    """GET /api/admin/backups — список резервних копій."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from backup import list_backups, DEFAULT_BACKUP_DIR

        backups = list_backups()
        return {"backups": backups, "count": len(backups)}
    except Exception as e:
        logger.exception("api_admin_backups_list error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_backup_create(request: Request):
    """POST /api/admin/backup — створити резервну копію вручну."""
    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    try:
        from backup import create_backup

        backup_path = create_backup()
        size_kb = round(backup_path.stat().st_size / 1024, 1)
        admin_id, admin_name = _get_admin_info(user)
        db.log_admin_action(
            admin_id,
            admin_name,
            "backup_create",
            f"Створено резервну копію вручну: {backup_path.name} ({size_kb} KB)",
            target_entity=backup_path.name,
            new_value={"filename": backup_path.name, "size_kb": size_kb},
        )
        return {
            "ok": True,
            "filename": backup_path.name,
            "size_kb": size_kb,
            "message": f"Резервну копію створено: {backup_path.name}",
        }
    except Exception as e:
        logger.exception("api_admin_backup_create error")
        return JSONResponse(content={"error": str(e)}, status_code=500)


async def api_admin_backup_download(request: Request, filename: str):
    """GET /api/admin/backup/download/{filename} — завантажити резервну копію."""
    import re as _re

    user = _extract_user(request)
    if not _is_admin(user):
        return JSONResponse(content={"error": "Тільки для адміністраторів"}, status_code=403)

    filename = filename or ""
    # Security: strictly validate the expected filename pattern to prevent path traversal
    # and injection attacks. Pattern: backup_YYYY-MM-DD_HH-MM.sql.gz
    _BACKUP_FILENAME_RE = _re.compile(r'^backup_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}\.sql\.gz$')
    if not filename or not _BACKUP_FILENAME_RE.match(filename):
        return JSONResponse(content={"error": "Невірне ім'я файлу"}, status_code=400)

    try:
        from backup import DEFAULT_BACKUP_DIR

        backup_path = DEFAULT_BACKUP_DIR / filename
        if not backup_path.exists():
            return JSONResponse(content={"error": "Файл не знайдено"}, status_code=404)

        with open(backup_path, "rb") as f:
            data = f.read()

        return Response(
            content=data,
            media_type="application/gzip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("api_admin_backup_download error")
        return JSONResponse(content={"error": str(e)}, status_code=500)
