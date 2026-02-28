"""Report generation business logic services."""
import io
import logging
from collections import defaultdict
from datetime import datetime, timedelta
import database.db_api as db
import config
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    MergedCell = None
    get_column_letter = None

logger = logging.getLogger(__name__)


def _build_daily_report_wb(generator_id: str, period_days: int, now: datetime) -> "Workbook":
    """Будує Excel-книгу з детальним щоденним звітом для одного генератора.

    Стовпці: Дата | Зміна 1 (поч/кін) | Зміна 2 | Зміна 3 | Екстра |
             Залишок ранок | Витрата | Залишок вечір | Мотогодини |
             Заправка (л) | Хто привіз | № чека
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError("openpyxl не встановлено")

    wb = Workbook()

    gen_name = db.get_generator_name(generator_id)
    end_date = now.strftime("%Y-%m-%d")
    start_date = (now - timedelta(days=period_days)).strftime("%Y-%m-%d")

    # --- Кольори ---
    BLUE_FILL = PatternFill(start_color="2481CC", end_color="2481CC", fill_type="solid")
    LBLUE_FILL = PatternFill(start_color="D6E8FA", end_color="D6E8FA", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True, size=11)
    BORDER_SIDE = None
    try:
        from openpyxl.styles import Border, Side

        thin = Side(style="thin", color="AAAAAA")
        BORDER_SIDE = Border(left=thin, right=thin, top=thin, bottom=thin)
    except Exception:
        pass

    def _style_header(cell, fill=BLUE_FILL):
        cell.font = WHITE_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if BORDER_SIDE:
            cell.border = BORDER_SIDE

    def _style_data(cell, bold=False, align="center"):
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if bold:
            cell.font = Font(bold=True)
        if BORDER_SIDE:
            cell.border = BORDER_SIDE

    # ---- Аркуш «Щоденний звіт» ----
    ws = wb.active
    ws.title = "Щоденний звіт"

    # Шапка
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    header_text = f"Звіт генератора «{gen_name}» за {period_days} днів | Сформовано: {now.strftime('%d.%m.%Y %H:%M')}"
    ws["A1"] = header_text
    ws["A1"].font = Font(bold=True, size=13, color="1A1A2E")
    ws.merge_cells("A1:M1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")

    # Рядки заголовків стовпців
    col_headers_r2 = [
        "Дата",
        "Зміна 1\nпочаток",
        "Зміна 1\nкінець",
        "Зміна 2\nпочаток",
        "Зміна 2\nкінець",
        "Зміна 3\nпочаток",
        "Зміна 3\nкінець",
        "Залишок\nранок, л",
        "Витрата\nза день, л",
        "Залишок\nвечір, л",
        "Мотогодини\n(накопичено)",
        "Заправка\n(прихід), л",
        "Хто привіз / № чека",
    ]
    for ci, h in enumerate(col_headers_r2, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        if ci == 1:
            _style_header(c, BLUE_FILL)
        elif ci in (2, 3):
            _style_header(c, PatternFill(start_color="1A7A44", end_color="1A7A44", fill_type="solid"))
        elif ci in (4, 5):
            _style_header(c, PatternFill(start_color="D4AC0D", end_color="D4AC0D", fill_type="solid"))
        elif ci in (6, 7):
            _style_header(c, PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"))
        elif ci in (8, 9, 10):
            _style_header(c, PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid"))
        elif ci == 11:
            _style_header(c, PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid"))
        elif ci in (12, 13):
            _style_header(c, PatternFill(start_color="117A65", end_color="117A65", fill_type="solid"))
        ws.row_dimensions[2].height = 40

    # Ширини стовпців
    col_widths = [12, 11, 11, 11, 11, 11, 11, 13, 13, 13, 15, 13, 30]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Отримуємо всі логи за період для цього генератора
    logs = db.get_logs_for_period(start_date, end_date, generator_id)

    # Агрегуємо по датах
    days_data = defaultdict(
        lambda: {
            "shifts": {"m": {}, "d": {}, "e": {}, "x": {}},
            "refills": [],
            "morning_fuel": None,
            "evening_fuel": None,
            "hours_start": None,
            "hours_end": None,
        }
    )

    for row_data in logs:
        event_type, ts_str, user_name, value, driver_name, receipt_number, *_ = row_data
        if not ts_str:
            continue
        try:
            ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        date_str = ts.strftime("%Y-%m-%d")
        day = days_data[date_str]

        if event_type.endswith("_start"):
            shift = event_type.split("_")[0]
            if shift in day["shifts"]:
                day["shifts"][shift]["start"] = ts.strftime("%H:%M")
        elif event_type.endswith("_end"):
            shift = event_type.split("_")[0]
            if shift in day["shifts"]:
                day["shifts"][shift]["end"] = ts.strftime("%H:%M")
        elif event_type == "refill":
            try:
                liters = float(value or 0)
            except Exception:
                liters = 0.0
            day["refills"].append((liters, (driver_name or "").strip(), (receipt_number or "").strip()))
        elif event_type == "corr_fuel_set":
            # Використовуємо останню корекцію дня як залишок
            try:
                day["evening_fuel"] = float(value or 0)
            except Exception:
                pass

    # Рядок початку — отримуємо поточний стан
    state = db.get_state()
    current_fuel = float(state.get("current_fuel", 0))

    # Генеруємо рядки за відсортованими датами
    data_row = 3
    prev_fuel = None
    fuel_rate = db.get_fuel_consumption_rate()

    for date_str in sorted(days_data.keys()):
        day = days_data[date_str]
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            date_fmt = dt.strftime("%d.%m.%Y")
        except Exception:
            date_fmt = date_str

        # Розраховуємо витрату
        total_shift_mins = 0
        for shift_data in day["shifts"].values():
            s_str = shift_data.get("start")
            e_str = shift_data.get("end")
            if s_str and e_str:
                try:
                    s_t = datetime.strptime(s_str, "%H:%M")
                    e_t = datetime.strptime(e_str, "%H:%M")
                    diff = (e_t - s_t).total_seconds() / 60
                    if diff < 0:
                        diff += 24 * 60
                    total_shift_mins += diff
                except Exception:
                    pass

        total_hours = round(total_shift_mins / 60, 2)
        consumption = round(total_hours * fuel_rate, 1) if total_hours > 0 else 0.0
        refill_total = round(sum(r[0] for r in day["refills"]), 1) if day["refills"] else 0.0

        morning_fuel = day.get("morning_fuel") or prev_fuel
        if morning_fuel is not None:
            evening_fuel = round(float(morning_fuel) + refill_total - consumption, 1)
        else:
            morning_fuel = ""
            evening_fuel = ""

        prev_fuel = evening_fuel if isinstance(evening_fuel, float) else None

        drivers_str = ", ".join(f"{drv} (чек {rec})" if rec else drv for _, drv, rec in day["refills"] if drv) or "—"

        row_vals = [
            date_fmt,
            day["shifts"]["m"].get("start", ""),
            day["shifts"]["m"].get("end", ""),
            day["shifts"]["d"].get("start", ""),
            day["shifts"]["d"].get("end", ""),
            day["shifts"]["e"].get("start", ""),
            day["shifts"]["e"].get("end", ""),
            morning_fuel if morning_fuel != "" else "—",
            consumption if consumption > 0 else "—",
            evening_fuel if evening_fuel != "" else "—",
            total_hours if total_hours > 0 else "—",
            refill_total if refill_total > 0 else "—",
            drivers_str,
        ]

        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=data_row, column=ci, value=val)
            bold = ci == 1
            align = "left" if ci == 13 else "center"
            _style_data(c, bold=bold, align=align)
            # Підсвітлення критичних залишків
            if ci == 10 and isinstance(val, float):
                if val < 15:
                    c.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                elif val < 40:
                    c.fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # ---- Аркуш ТО ----
    ws_mnt = wb.create_sheet("Технічне обслуговування")
    stats = db.get_maintenance_stats(generator_id)
    mnt_history = db.get_maintenance_history(generator_id, 100)

    ws_mnt["A1"] = f"Технічне обслуговування — {gen_name}"
    ws_mnt["A1"].font = Font(bold=True, size=13)
    ws_mnt.merge_cells("A1:E1")
    ws_mnt["A1"].alignment = Alignment(horizontal="center")
    ws_mnt["A1"].fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")

    ws_mnt["A3"] = "Мотогодини (загалом):"
    ws_mnt["B3"] = f"{float(stats.get('total_hours', 0)):.1f} год"
    ws_mnt["A3"].font = BOLD_FONT
    ws_mnt["B3"].font = Font(size=11)

    mnt_col_hdrs = ["Дата", "Тип ТО", "Мотогодини на момент ТО", "Виконав", "Примітки"]
    for ci, h in enumerate(mnt_col_hdrs, start=1):
        c = ws_mnt.cell(row=5, column=ci, value=h)
        _style_header(c)

    ws_mnt.column_dimensions["A"].width = 14
    ws_mnt.column_dimensions["B"].width = 22
    ws_mnt.column_dimensions["C"].width = 26
    ws_mnt.column_dimensions["D"].width = 20
    ws_mnt.column_dimensions["E"].width = 20

    mnt_map = {"oil": "Заміна мастила", "spark": "Заміна свічок", "maintenance": "Планове ТО"}
    for ri, rec in enumerate(mnt_history, start=6):
        rec_id, date_s, action, hours, admin_name, *_ = rec
        ws_mnt.cell(row=ri, column=1, value=date_s)
        ws_mnt.cell(row=ri, column=2, value=mnt_map.get(action, action))
        ws_mnt.cell(row=ri, column=3, value=f"{float(hours):.1f} год")
        ws_mnt.cell(row=ri, column=4, value=admin_name or "—")
        for ci in range(1, 5):
            _style_data(ws_mnt.cell(row=ri, column=ci))

    return wb
