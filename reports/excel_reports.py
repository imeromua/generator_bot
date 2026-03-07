"""Enhanced Excel Report Generator with Charts and Advanced Formatting.

Features:
- Multiple report types (quick, detailed, technical, financial)
- Conditional formatting (low fuel = red, high efficiency = green)
- Multi-sheet workbooks
- Auto-sizing columns
- Professional styling
"""

import io
import logging
from calendar import monthrange
from collections import defaultdict
from datetime import datetime, timedelta

import config
import database.db_api as db

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:  # pragma: no cover
    EXCEL_AVAILABLE = False


MONTH_NAMES = {
    1: 'Січень', 2: 'Лютий', 3: 'Березень', 4: 'Квітень',
    5: 'Травень', 6: 'Червень', 7: 'Липень', 8: 'Серпень',
    9: 'Вересень', 10: 'Жовтень', 11: 'Листопад', 12: 'Грудень',
}

MONTH_NAMES_GENITIVE = {
    1: 'Січня', 2: 'Лютого', 3: 'Березня', 4: 'Квітня',
    5: 'Травня', 6: 'Червня', 7: 'Липня', 8: 'Серпня',
    9: 'Вересня', 10: 'Жовтня', 11: 'Листопада', 12: 'Грудня',
}

WEEKDAY_NAMES = {
    0: 'Понеділок', 1: 'Вівторок', 2: 'Середа', 3: 'Четвер',
    4: "П'ятниця", 5: 'Субота', 6: 'Неділя',
}

WEEKDAY_ABBREV = {
    0: 'Пн', 1: 'Вт', 2: 'Ср', 3: 'Чт', 4: 'Пт', 5: 'Сб', 6: 'Нд',
}


class ExcelReportGenerator:
    """Generate professional Excel reports with charts."""

    # Color scheme
    COLORS = {
        'header_blue': '2481CC',
        'light_blue': 'D6E8FA',
        'header_bg': 'EAF2FB',
        'green': '27AE60',
        'dark_green': '1A7A44',
        'orange': 'F39C12',
        'dark_yellow': 'D4AC0D',
        'red': 'E74C3C',
        'yellow': 'F4D03F',
        'gray': 'BDC3C7',
        'purple': '6C3483',
        'dark_blue': '1565C0',
        'teal': '117A65',
        'dark_text': '1A1A2E',
        'light_green': 'D5F5E3',
        # Detailed report column-group section fills (light tints for data cells)
        'sect_date': 'EBF5FB',
        'sect_morning': 'EAFAF1',
        'sect_day': 'E8F8F5',
        'sect_evening': 'F5EEF8',
        'sect_runtime': 'EBF5FB',
        'sect_fuel': 'FEF9E7',
        'sect_docs': 'F2F3F4',
        # Row-level fills
        'alt_row': 'F4F6F7',
        'weekend_row': 'FFFDE7',
        'total_row': 'D5D8DC',
        # Group header backgrounds (darker, white text)
        'grp_date': '1A5276',
        'grp_morning': '1E8449',
        'grp_day': '148F77',
        'grp_evening': '6C3483',
        'grp_runtime': '1A5276',
        'grp_fuel': 'B7770D',
        'grp_docs': '5D6D7E',
        # Fuel alert fills
        'fuel_critical': 'FADBD8',
        'fuel_low': 'FDEBD0',
        'fuel_ok': 'D5F5E3',
        'mid_gray': '7F8C8D',
    }

    def __init__(self):
        self.wb = None
        self._border = None
        self._init_border()

    def _init_border(self):
        if not EXCEL_AVAILABLE:
            return
        try:
            thin = Side(style='thin', color='AAAAAA')
            self._border = Border(left=thin, right=thin, top=thin, bottom=thin)
        except Exception:
            pass

    def _month_range(self, year: int, month: int) -> tuple:
        """Return (start_dt, end_dt) for a calendar month — 00:00:00 to 23:59:59."""
        last_day = monthrange(year, month)[1]
        tz = config.KYIV
        start_dt = datetime(year, month, 1, 0, 0, 0, tzinfo=tz)
        end_dt = datetime(year, month, last_day, 23, 59, 59, tzinfo=tz)
        return start_dt, end_dt

    def _fill(self, color_key: str) -> "PatternFill":
        color = self.COLORS.get(color_key, color_key)
        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _style_header(self, cell, color_key: str = 'header_blue'):
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = self._fill(color_key)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if self._border:
            cell.border = self._border

    def _style_data(self, cell, bold: bool = False, align: str = 'center'):
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if bold:
            cell.font = Font(bold=True)
        if self._border:
            cell.border = self._border

    def generate_report(
        self,
        report_type: str,
        days: int,
        generator_id: str = None,
        year: int = None,
        month: int = None,
    ) -> bytes:
        """Generate Excel report based on type."""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl не встановлено")

        self.wb = Workbook()

        if report_type == 'quick':
            self._generate_quick_report(days, generator_id, year=year, month=month)
        elif report_type == 'detailed':
            self._generate_detailed_report(days, generator_id, year=year, month=month)
        elif report_type == 'technical':
            self._generate_technical_report(days, generator_id, year=year, month=month)
        elif report_type == 'financial':
            self._generate_financial_report(days, generator_id, year=year, month=month)
        elif report_type == 'personnel':
            self._generate_personnel_report(days, generator_id, year=year, month=month)
        else:
            raise ValueError(f"Невідомий тип звіту: {report_type}")

        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ------------------------------------------------------------------
    # Quick summary report
    # ------------------------------------------------------------------

    def _generate_quick_report(self, days: int, gen_id: str, year: int = None, month: int = None):
        """Quick summary report with KPI cards and daily table."""
        ws = self.wb.active
        ws.title = "Швидкий звіт"

        gen_id = gen_id or db.get_active_generator()
        gen_name = db.get_generator_name(gen_id)
        if year and month:
            start_dt, end_dt = self._month_range(year, month)
            period_label = f"{MONTH_NAMES[month]} {year}"
        else:
            end_dt = datetime.now(config.KYIV)
            start_dt = end_dt - timedelta(days=days)
            period_label = f"за {days} днів"

        # Title
        ws['A1'] = f"Генератор «{gen_name}» — Швидкий звіт: {period_label}"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['dark_text'])
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 28

        # KPI Section header
        ws['A3'] = 'Ключові показники'
        ws['A3'].font = Font(bold=True, size=12)

        daily = self._build_daily_stats(start_dt, end_dt, gen_id)
        total_hours = round(sum(d['work_hours'] for d in daily), 1)
        total_fuel = round(sum(d['fuel_consumed'] for d in daily), 1)
        avg_rate = round(total_fuel / total_hours, 2) if total_hours > 0 else 0.0
        fuel_price = getattr(config, 'FUEL_PRICE', 50.0)
        fuel_cost = round(total_fuel * fuel_price, 0)

        kpi_labels = ['Мотогодини', 'Витрата палива, л', 'Сер. витрата, л/год', 'Вартість палива, грн']
        kpi_values = [total_hours, total_fuel, avg_rate, fuel_cost]

        for ci, (lbl, val) in enumerate(zip(kpi_labels, kpi_values), start=1):
            lc = ws.cell(row=4, column=ci, value=lbl)
            lc.font = Font(bold=True, size=10)
            lc.fill = self._fill('light_blue')
            lc.alignment = Alignment(horizontal='center')
            vc = ws.cell(row=5, column=ci, value=val)
            vc.font = Font(bold=True, size=13)
            vc.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = 22

        # Daily summary table
        ws['A7'] = 'Щоденна статистика'
        ws['A7'].font = Font(bold=True, size=12)

        headers = ['Дата', 'Мотогодини', 'Витрата, л', 'Залишок ранок, л', 'Залишок вечір, л', 'Заправка, л']
        for ci, h in enumerate(headers, start=1):
            self._style_header(ws.cell(row=8, column=ci, value=h))
        ws.row_dimensions[8].height = 30

        for ri, d in enumerate(daily, start=9):
            ws.cell(row=ri, column=1, value=d['date'])
            ws.cell(row=ri, column=2, value=d['work_hours'])
            ws.cell(row=ri, column=3, value=d['fuel_consumed'])
            ws.cell(row=ri, column=4, value=d.get('morning_fuel', ''))
            eve_cell = ws.cell(row=ri, column=5, value=d.get('evening_fuel', ''))
            ws.cell(row=ri, column=6, value=d.get('refill_total', 0) or '')
            for ci in range(1, 7):
                self._style_data(ws.cell(row=ri, column=ci), bold=(ci == 1))
            # Conditional formatting on evening fuel
            eve_val = d.get('evening_fuel')
            if isinstance(eve_val, (int, float)):
                if eve_val < 15:
                    eve_cell.fill = self._fill('red')
                elif eve_val < 40:
                    eve_cell.fill = self._fill('orange')

        # Fuel timeline chart
        if len(daily) > 1:
            chart = LineChart()
            chart.title = f"Витрата палива — {gen_name}"
            chart.x_axis.title = "Дата"
            chart.y_axis.title = "Літри"
            chart.width = 20
            chart.height = 10

            data_start_row = 9
            data_end_row = 8 + len(daily)
            fuel_ref = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)
            date_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            chart.add_data(fuel_ref, titles_from_data=False)
            chart.set_categories(date_ref)
            if chart.series and chart.series[0] is not None and chart.series[0].title is not None:
                chart.series[0].title.value = "Витрата, л"
            ws.add_chart(chart, f"A{data_end_row + 3}")

    # ------------------------------------------------------------------
    # Detailed multi-sheet operational monthly report
    # ------------------------------------------------------------------

    def _generate_detailed_report(self, days: int, gen_id: str, year: int = None, month: int = None):
        """Detailed monthly operational workbook — always uses the current calendar month.

        The ``days`` parameter is intentionally ignored for this report type so
        that the output is always a clean, complete monthly table regardless of
        whatever period-selector value the caller passes.
        """
        now = datetime.now(config.KYIV)
        report_year = year or now.year
        report_month = month or now.month
        start_dt, end_dt = self._month_range(report_year, report_month)

        gen_id = gen_id or db.get_active_generator()
        gen_name = db.get_generator_name(gen_id)

        # Build enriched per-day data
        daily_data = self._build_detailed_daily_data(start_dt, end_dt, gen_id)

        # Sheet 1 — Main operational journal (primary fallback for Google Sheets)
        ws_ops = self.wb.active
        ws_ops.title = "Операційний журнал"
        self._fill_operational_sheet(ws_ops, daily_data, gen_name, gen_id, report_year, report_month)

        # Sheet 2 — Monthly KPI summary
        ws_sum = self.wb.create_sheet("Зведення місяця")
        self._fill_monthly_summary_sheet(ws_sum, daily_data, gen_name, gen_id, report_year, report_month)

        # Sheet 3 — Maintenance history
        ws_maint = self.wb.create_sheet("Технічне обслуговування")
        self._fill_maintenance_sheet(ws_maint, gen_id, gen_name)

    # ------------------------------------------------------------------
    # Detailed report: rich per-day data builder
    # ------------------------------------------------------------------

    def _build_detailed_daily_data(self, start_dt: datetime, end_dt: datetime, gen_id: str) -> list:
        """Return one dict per calendar day in [start_dt, end_dt] with full shift, fuel and
        personnel details — all days are present even if no activity occurred.
        """
        start_date = start_dt.strftime('%Y-%m-%d')
        end_date = end_dt.strftime('%Y-%m-%d')
        logs = db.get_logs_for_period(start_date, end_date, gen_id)
        fuel_rate = db.get_fuel_consumption_rate()

        # Parse logs into per-day buckets
        days_raw: dict = defaultdict(lambda: {
            'shifts': {'m': {}, 'd': {}, 'e': {}, 'x': {}},
            'refills': [],
            'corr_fuel': None,
        })

        for row_data in logs:
            event_type, ts_str, user_name, value, driver_name, receipt_number, *_ = row_data
            if not ts_str:
                continue
            try:
                ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                continue
            date_str = ts.strftime('%Y-%m-%d')
            raw = days_raw[date_str]

            if event_type.endswith('_start'):
                shift = event_type.split('_')[0]
                if shift in raw['shifts']:
                    raw['shifts'][shift]['start'] = ts.strftime('%H:%M')
                    raw['shifts'][shift]['operator'] = user_name or ''
            elif event_type.endswith('_end'):
                shift = event_type.split('_')[0]
                if shift in raw['shifts']:
                    raw['shifts'][shift]['end'] = ts.strftime('%H:%M')
            elif event_type == 'refill':
                try:
                    liters = float(value or 0)
                except Exception:
                    liters = 0.0
                raw['refills'].append({
                    'liters': liters,
                    'driver': driver_name or '',
                    'receipt': receipt_number or '',
                })
            elif event_type == 'corr_fuel_set':
                try:
                    raw['corr_fuel'] = float(value or 0)
                except Exception:
                    pass

        # Generate one record for every calendar day in the month
        prev_evening_fuel = None
        result = []
        current = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)

        while current.date() <= end_dt.date():
            date_str = current.strftime('%Y-%m-%d')
            weekday = current.weekday()
            raw = days_raw.get(date_str) or {
                'shifts': {'m': {}, 'd': {}, 'e': {}, 'x': {}},
                'refills': [],
                'corr_fuel': None,
            }

            # Compute shift durations and collect personnel
            total_shift_mins = 0.0
            shifts_active = []
            shift_details: dict = {}
            operators: list = []

            for sname in ('m', 'd', 'e', 'x'):
                sdata = raw['shifts'].get(sname, {})
                s_str = sdata.get('start')
                e_str = sdata.get('end')
                operator = sdata.get('operator', '')
                duration = 0.0
                if s_str and e_str:
                    try:
                        s_t = datetime.strptime(s_str, '%H:%M')
                        e_t = datetime.strptime(e_str, '%H:%M')
                        diff = (e_t - s_t).total_seconds() / 60
                        if diff < 0:
                            diff += 24 * 60
                        total_shift_mins += diff
                        shifts_active.append(sname)
                        duration = round(diff / 60, 2)
                    except Exception:
                        pass
                shift_details[sname] = {'start': s_str, 'end': e_str, 'duration': duration, 'operator': operator}
                if operator and operator not in operators:
                    operators.append(operator)

            work_hours = round(total_shift_mins / 60, 2)
            fuel_consumed = round(work_hours * fuel_rate, 1) if work_hours > 0 else 0.0
            refill_total = round(sum(r['liters'] for r in raw['refills']), 1)
            receipt_numbers = [r['receipt'] for r in raw['refills'] if r.get('receipt')]

            # Fuel balance propagation — same logic as _build_daily_stats
            corr_fuel = raw.get('corr_fuel')
            if corr_fuel is not None:
                evening_fuel: float | None = corr_fuel
                morning_fuel: float | None = round(corr_fuel - refill_total + fuel_consumed, 1)
                if morning_fuel < 0:
                    morning_fuel = 0.0
            elif prev_evening_fuel is not None:
                morning_fuel = prev_evening_fuel
                evening_fuel = round(morning_fuel + refill_total - fuel_consumed, 1)
                if evening_fuel < 0:
                    evening_fuel = 0.0
            elif refill_total > 0:
                morning_fuel = 0.0
                evening_fuel = round(refill_total - fuel_consumed, 1)
                if evening_fuel < 0:
                    evening_fuel = 0.0
            else:
                morning_fuel = None
                evening_fuel = None

            if evening_fuel is not None:
                prev_evening_fuel = evening_fuel

            fuel_rate_actual = round(fuel_consumed / work_hours, 2) if work_hours > 0 else None

            # Auto-generate warning notes
            notes = ''
            if isinstance(evening_fuel, (int, float)):
                if evening_fuel < 15:
                    notes = '⚠ КРИТИЧНО: низький рівень палива'
                elif evening_fuel < 40:
                    notes = '⚠ Низький рівень палива'
            if work_hours == 0 and not notes:
                notes = '—'

            result.append({
                'day_num': current.day,
                'date_str': date_str,
                'date': current.strftime('%d.%m.%Y'),
                'weekday': WEEKDAY_NAMES.get(weekday, ''),
                'weekday_abbrev': WEEKDAY_ABBREV.get(weekday, ''),
                'is_weekend': weekday >= 5,
                'shift_details': shift_details,
                'shifts_active': shifts_active,
                'work_hours': work_hours,
                'operators': operators,
                'morning_fuel': morning_fuel,
                'evening_fuel': evening_fuel,
                'fuel_consumed': fuel_consumed,
                'refill_total': refill_total if refill_total > 0 else None,
                'refills': raw['refills'],
                'fuel_rate': fuel_rate_actual,
                'receipt_numbers': receipt_numbers,
                'notes': notes,
            })
            current += timedelta(days=1)

        return result

    # ------------------------------------------------------------------
    # Detailed report: main operational sheet
    # ------------------------------------------------------------------

    def _fill_operational_sheet(
        self, ws, daily_data: list, gen_name: str, gen_id: str, year: int, month: int
    ):
        """Fill the primary operational journal sheet with full month data."""

        # ------------------------------------------------------------------ column spec
        # Each entry: (letter, group_key, group_label, col_header, width)
        # Groups:  date | morning | day | evening | runtime | fuel | docs
        COL_DEFS = [
            ('A', 'date',    '',         '#',           4),
            ('B', 'date',    'НАВІГАЦІЯ','Дата',        11),
            ('C', 'date',    '',         'День',        5),
            ('D', 'morning', 'РАНОК',    'Р↑ Поч.',     8),
            ('E', 'morning', '',         'Р↓ Кін.',     8),
            ('F', 'day',     'ДЕНЬ',     'Д↑ Поч.',     8),
            ('G', 'day',     '',         'Д↓ Кін.',     8),
            ('H', 'evening', 'ВЕЧІР',    'В↑ Поч.',     8),
            ('I', 'evening', '',         'В↓ Кін.',     8),
            ('J', 'runtime', 'ВИРОБІТОК','Год.',         7),
            ('K', 'runtime', '',         'Оператор',   16),
            ('L', 'fuel',    'ПАЛИВО',   'Пал.⬆, л',   11),
            ('M', 'fuel',    '',         'Пал.⬇, л',   11),
            ('N', 'fuel',    '',         'Витрата, л',  11),
            ('O', 'fuel',    '',         'Заправка, л', 11),
            ('P', 'fuel',    '',         'л/год',        7),
            ('Q', 'docs',    'ДОКУМЕНТИ','Чек #',       11),
            ('R', 'docs',    '',         'Примітки',    24),
        ]
        N_COLS = len(COL_DEFS)
        LAST_COL = COL_DEFS[-1][0]

        GROUP_HDR_COLORS = {
            'date': 'grp_date',
            'morning': 'grp_morning',
            'day': 'grp_day',
            'evening': 'grp_evening',
            'runtime': 'grp_runtime',
            'fuel': 'grp_fuel',
            'docs': 'grp_docs',
        }
        GROUP_CELL_COLORS = {
            'date': 'sect_date',
            'morning': 'sect_morning',
            'day': 'sect_day',
            'evening': 'sect_evening',
            'runtime': 'sect_runtime',
            'fuel': 'sect_fuel',
            'docs': 'sect_docs',
        }

        TITLE_ROW = 1
        META_ROW = 2
        GRP_HDR_ROW = 3
        COL_HDR_ROW = 4
        DATA_START = 5

        now_str = datetime.now(config.KYIV).strftime('%d.%m.%Y %H:%M')
        month_title = f"{MONTH_NAMES[month]} {year}"

        # ------------------------------------------------------------------ row 1: title
        ws.merge_cells(f'A{TITLE_ROW}:{LAST_COL}{TITLE_ROW}')
        tc = ws[f'A{TITLE_ROW}']
        tc.value = f"Операційний журнал роботи генератора «{gen_name}»  —  {month_title}"
        tc.font = Font(bold=True, size=14, color='FFFFFF')
        tc.fill = self._fill('grp_date')
        tc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[TITLE_ROW].height = 30

        # ------------------------------------------------------------------ row 2: metadata
        ws.merge_cells(f'A{META_ROW}:{LAST_COL}{META_ROW}')
        mc = ws[f'A{META_ROW}']
        mc.value = (
            f"Генератор: «{gen_name}»  |  Місяць: {month_title}"
            f"  |  Сформовано: {now_str}  |  Рядків: {len(daily_data)}"
        )
        mc.font = Font(italic=True, size=10, color='FFFFFF')
        mc.fill = self._fill('grp_docs')
        mc.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[META_ROW].height = 18

        # ------------------------------------------------------------------ row 3: group headers
        # Build merged ranges per group
        group_ranges: dict = {}
        for col_letter, group_key, group_label, col_header, width in COL_DEFS:
            col_idx = ord(col_letter) - ord('A') + 1
            if group_key not in group_ranges:
                group_ranges[group_key] = [col_idx, col_idx, group_label or col_header]
            else:
                group_ranges[group_key][1] = col_idx

        for group_key, (start_ci, end_ci, label) in group_ranges.items():
            start_l = get_column_letter(start_ci)
            end_l = get_column_letter(end_ci)
            if start_ci != end_ci:
                ws.merge_cells(f'{start_l}{GRP_HDR_ROW}:{end_l}{GRP_HDR_ROW}')
            c = ws.cell(row=GRP_HDR_ROW, column=start_ci, value=label)
            c.font = Font(bold=True, size=9, color='FFFFFF')
            c.fill = self._fill(GROUP_HDR_COLORS.get(group_key, 'grp_docs'))
            c.alignment = Alignment(horizontal='center', vertical='center')
            if self._border:
                c.border = self._border
        ws.row_dimensions[GRP_HDR_ROW].height = 18

        # ------------------------------------------------------------------ row 4: column headers
        col_group_map = {col_letter: group_key for col_letter, group_key, *_ in COL_DEFS}
        for col_letter, group_key, _, col_header, width in COL_DEFS:
            col_idx = ord(col_letter) - ord('A') + 1
            c = ws.cell(row=COL_HDR_ROW, column=col_idx, value=col_header)
            c.font = Font(bold=True, size=9, color='FFFFFF')
            c.fill = self._fill(GROUP_HDR_COLORS.get(group_key, 'grp_docs'))
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if self._border:
                c.border = self._border
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[COL_HDR_ROW].height = 28

        # Auto-filter on column header row
        ws.auto_filter.ref = f'A{COL_HDR_ROW}:{LAST_COL}{COL_HDR_ROW}'

        # Freeze panes: freeze columns A-B and rows 1-4
        ws.freeze_panes = f'C{DATA_START}'

        # ------------------------------------------------------------------ data rows
        total_hours_sum = 0.0
        total_consumed_sum = 0.0
        total_refill_sum = 0.0

        for i, day in enumerate(daily_data):
            ri = DATA_START + i

            # Row background
            if day['is_weekend']:
                row_fill = self._fill('weekend_row')
            elif i % 2 == 0:
                row_fill = self._fill('alt_row')
            else:
                row_fill = PatternFill(fill_type=None)  # white

            def _write(col_letter, value, bold=False, align='center', fill=None, num_fmt=None):
                ci = ord(col_letter) - ord('A') + 1
                c = ws.cell(row=ri, column=ci, value=value)
                c.font = Font(bold=bold, size=9)
                c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=(col_letter == 'R'))
                if fill:
                    c.fill = fill
                elif row_fill.fill_type:
                    c.fill = row_fill
                if num_fmt:
                    c.number_format = num_fmt
                if self._border:
                    c.border = self._border
                return c

            sd = day['shift_details']

            _write('A', day['day_num'], bold=True, align='center')
            _write('B', day['date'], bold=day['is_weekend'], align='center')
            _write('C', day['weekday_abbrev'], bold=day['is_weekend'])
            _write('D', sd['m'].get('start') or '—')
            _write('E', sd['m'].get('end') or '—')
            _write('F', sd['d'].get('start') or '—')
            _write('G', sd['d'].get('end') or '—')
            _write('H', sd['e'].get('start') or '—')
            _write('I', sd['e'].get('end') or '—')

            hrs_val = day['work_hours'] if day['work_hours'] > 0 else None
            _write('J', hrs_val, bold=(hrs_val is not None), num_fmt='0.00')
            _write('K', ', '.join(day['operators']) if day['operators'] else '—', align='left')

            # Fuel columns — with conditional color coding
            morning_f = day['morning_fuel']
            evening_f = day['evening_fuel']
            consumed = day['fuel_consumed'] if day['fuel_consumed'] > 0 else None
            refill = day['refill_total']
            rate = day['fuel_rate']

            _write('L', morning_f if morning_f is not None else '—', num_fmt='0.0')
            eve_c = _write('M', evening_f if evening_f is not None else '—', num_fmt='0.0')
            _write('N', consumed, num_fmt='0.0')
            _write('O', refill, num_fmt='0.0')
            _write('P', rate, num_fmt='0.00')

            # Fuel alert coloring on evening fuel cell
            if isinstance(evening_f, (int, float)):
                if evening_f < 15:
                    eve_c.fill = self._fill('fuel_critical')
                    eve_c.font = Font(bold=True, size=9, color=self.COLORS['red'])
                elif evening_f < 40:
                    eve_c.fill = self._fill('fuel_low')
                elif evening_f > 60:
                    eve_c.fill = self._fill('fuel_ok')

            _write('Q', ', '.join(day['receipt_numbers']) if day['receipt_numbers'] else '—')
            notes_c = _write('R', day['notes'], align='left')
            if '⚠ КРИТИЧНО' in (day['notes'] or ''):
                notes_c.font = Font(bold=True, size=9, color=self.COLORS['red'])
            elif '⚠' in (day['notes'] or ''):
                notes_c.font = Font(bold=False, size=9, color=self.COLORS['orange'])

            # Weekend indicator: bold date+day
            if day['is_weekend']:
                for col_l in ('B', 'C'):
                    ci = ord(col_l) - ord('A') + 1
                    ws.cell(row=ri, column=ci).font = Font(bold=True, size=9, color=self.COLORS['dark_blue'])

            total_hours_sum += day['work_hours']
            total_consumed_sum += day['fuel_consumed']
            total_refill_sum += day['refill_total'] or 0.0

            ws.row_dimensions[ri].height = 16

        # ------------------------------------------------------------------ totals row
        totals_row = DATA_START + len(daily_data)
        ws.merge_cells(f'A{totals_row}:C{totals_row}')
        tc = ws.cell(row=totals_row, column=1, value='ПІДСУМОК МІСЯЦЯ')
        tc.font = Font(bold=True, size=10, color='FFFFFF')
        tc.fill = self._fill('grp_date')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        if self._border:
            tc.border = self._border

        def _tot(col_letter, value, num_fmt=None):
            ci = ord(col_letter) - ord('A') + 1
            c = ws.cell(row=totals_row, column=ci, value=value)
            c.font = Font(bold=True, size=10)
            c.fill = self._fill('total_row')
            c.alignment = Alignment(horizontal='center', vertical='center')
            if num_fmt:
                c.number_format = num_fmt
            if self._border:
                c.border = self._border
            return c

        for col_l in ('D', 'E', 'F', 'G', 'H', 'I', 'K', 'P', 'Q', 'R'):
            _tot(col_l, '')

        _tot('J', round(total_hours_sum, 2), '0.00')
        _tot('L', '')
        _tot('M', '')
        _tot('N', round(total_consumed_sum, 1), '0.0')
        _tot('O', round(total_refill_sum, 1), '0.0')
        ws.row_dimensions[totals_row].height = 22

    # ------------------------------------------------------------------
    # Detailed report: monthly summary sheet
    # ------------------------------------------------------------------

    def _fill_monthly_summary_sheet(
        self, ws, daily_data: list, gen_name: str, gen_id: str, year: int, month: int
    ):
        """Fill the monthly KPI summary sheet."""
        month_title = f"{MONTH_NAMES[month]} {year}"
        now_str = datetime.now(config.KYIV).strftime('%d.%m.%Y %H:%M')

        # Title
        ws.merge_cells('A1:F1')
        tc = ws['A1']
        tc.value = f"Зведення місяця — «{gen_name}»  —  {month_title}"
        tc.font = Font(bold=True, size=14, color='FFFFFF')
        tc.fill = self._fill('grp_date')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:F2')
        mc = ws['A2']
        mc.value = f"Сформовано: {now_str}"
        mc.font = Font(italic=True, size=10)
        mc.alignment = Alignment(horizontal='left')
        ws.row_dimensions[2].height = 16

        # Compute KPIs
        active_days = [d for d in daily_data if d['work_hours'] > 0]
        total_hours = round(sum(d['work_hours'] for d in daily_data), 2)
        total_fuel = round(sum(d['fuel_consumed'] for d in daily_data), 1)
        total_refill = round(sum(d['refill_total'] or 0 for d in daily_data), 1)
        avg_rate = round(total_fuel / total_hours, 2) if total_hours > 0 else 0.0
        fuel_price = getattr(config, 'FUEL_PRICE', 50.0)
        fuel_cost = round(total_fuel * fuel_price, 0)
        working_days = len(active_days)
        total_days = len(daily_data)
        weekend_days = sum(1 for d in daily_data if d['is_weekend'])

        # Section header
        ws['A4'] = 'Ключові показники'
        ws['A4'].font = Font(bold=True, size=12)
        ws.row_dimensions[4].height = 22

        kpi_data = [
            ('Мотогодини за місяць', f"{total_hours:.2f} год"),
            ('Робочих днів', f"{working_days} / {total_days}"),
            ('Вихідних у місяці', str(weekend_days)),
            ('Загальна витрата палива', f"{total_fuel:.1f} л"),
            ('Загальне поповнення', f"{total_refill:.1f} л"),
            ('Середня витрата', f"{avg_rate:.2f} л/год"),
            ('Ціна палива', f"{fuel_price:.2f} грн/л"),
            ('Вартість палива', f"{fuel_cost:.0f} грн"),
        ]

        for ri, (label, value) in enumerate(kpi_data, start=5):
            lc = ws.cell(row=ri, column=1, value=label)
            lc.font = Font(bold=True, size=10)
            lc.fill = self._fill('light_blue')
            lc.alignment = Alignment(horizontal='left', vertical='center')
            if self._border:
                lc.border = self._border

            vc = ws.cell(row=ri, column=2, value=value)
            vc.font = Font(size=10)
            vc.alignment = Alignment(horizontal='center', vertical='center')
            if self._border:
                vc.border = self._border
            ws.row_dimensions[ri].height = 18

        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 20

        # Fuel warning summary
        critical_days = [d for d in daily_data if isinstance(d['evening_fuel'], (int, float)) and d['evening_fuel'] < 15]
        low_days = [d for d in daily_data if isinstance(d['evening_fuel'], (int, float)) and 15 <= d['evening_fuel'] < 40]

        warn_row = 5 + len(kpi_data) + 1
        if critical_days or low_days:
            ws.cell(row=warn_row, column=1, value='Попередження').font = Font(bold=True, size=11)
            warn_row += 1
            if critical_days:
                c = ws.cell(row=warn_row, column=1,
                            value=f"⚠ КРИТИЧНО низький рівень палива: {len(critical_days)} дн.")
                c.font = Font(bold=True, size=10, color=self.COLORS['red'])
                warn_row += 1
            if low_days:
                c = ws.cell(row=warn_row, column=1,
                            value=f"⚠ Низький рівень палива: {len(low_days)} дн.")
                c.font = Font(bold=False, size=10, color=self.COLORS['orange'])
                warn_row += 1

        # Per-day summary table
        tbl_start = warn_row + 1
        ws.cell(row=tbl_start, column=1, value='Щоденна зведена таблиця').font = Font(bold=True, size=11)
        tbl_start += 1

        sum_headers = ['Дата', 'День', 'Год.', 'Пал.⬆, л', 'Пал.⬇, л', 'Витрата, л', 'Заправка, л', 'Примітки']
        for ci, h in enumerate(sum_headers, start=1):
            c = ws.cell(row=tbl_start, column=ci, value=h)
            c.font = Font(bold=True, size=9, color='FFFFFF')
            c.fill = self._fill('grp_date')
            c.alignment = Alignment(horizontal='center', vertical='center')
            if self._border:
                c.border = self._border
        ws.row_dimensions[tbl_start].height = 20

        for i, day in enumerate(daily_data):
            ri = tbl_start + 1 + i
            row_fill = self._fill('weekend_row') if day['is_weekend'] else (
                self._fill('alt_row') if i % 2 == 0 else PatternFill(fill_type=None)
            )

            def _sw(col, value, num_fmt=None):
                c = ws.cell(row=ri, column=col, value=value)
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                if row_fill.fill_type:
                    c.fill = row_fill
                if num_fmt:
                    c.number_format = num_fmt
                if self._border:
                    c.border = self._border
                return c

            _sw(1, day['date'])
            _sw(2, day['weekday_abbrev'])
            _sw(3, day['work_hours'] if day['work_hours'] > 0 else None, '0.00')
            _sw(4, day['morning_fuel'], '0.0')
            eve_c = _sw(5, day['evening_fuel'], '0.0')
            _sw(6, day['fuel_consumed'] if day['fuel_consumed'] > 0 else None, '0.0')
            _sw(7, day['refill_total'], '0.0')
            notes_c = _sw(8, day['notes'] or '')

            if isinstance(day['evening_fuel'], (int, float)):
                if day['evening_fuel'] < 15:
                    eve_c.fill = self._fill('fuel_critical')
                elif day['evening_fuel'] < 40:
                    eve_c.fill = self._fill('fuel_low')

            if '⚠ КРИТИЧНО' in (day['notes'] or ''):
                notes_c.font = Font(bold=True, size=9, color=self.COLORS['red'])
            elif '⚠' in (day['notes'] or ''):
                notes_c.font = Font(size=9, color=self.COLORS['orange'])

            ws.row_dimensions[ri].height = 15

        for ci, w in enumerate([11, 5, 7, 10, 10, 11, 11, 26], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = f'A{tbl_start + 1}'

    def _fill_summary_sheet(self, ws, days, gen_id, gen_name, end_dt, start_dt, period_label=None):
        if period_label is None:
            period_label = f"за {days} днів"
        ws['A1'] = f"Детальний звіт: «{gen_name}» {period_label}"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['dark_text'])
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 28

        ws['A2'] = f"Сформовано: {end_dt.strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)

        daily = self._build_daily_stats(start_dt, end_dt, gen_id)
        total_hours = round(sum(d['work_hours'] for d in daily), 1)
        total_fuel = round(sum(d['fuel_consumed'] for d in daily), 1)
        avg_rate = round(total_fuel / total_hours, 2) if total_hours > 0 else 0.0
        fuel_price = getattr(config, 'FUEL_PRICE', 50.0)
        fuel_cost = round(total_fuel * fuel_price, 0)
        total_refills = round(sum(d.get('refill_total', 0) or 0 for d in daily), 1)

        summary_data = [
            ('Генератор', gen_name),
            ('Період', f"{start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')}"),
            ('Кількість днів', days),
            ('Загальні мотогодини', total_hours),
            ('Загальна витрата палива, л', total_fuel),
            ('Середня витрата, л/год', avg_rate),
            ('Загальна вартість палива, грн', fuel_cost),
            ('Загальне поповнення, л', total_refills),
        ]

        for ri, (label, value) in enumerate(summary_data, start=4):
            lc = ws.cell(row=ri, column=1, value=label)
            lc.font = Font(bold=True)
            lc.fill = self._fill('light_blue')
            vc = ws.cell(row=ri, column=2, value=value)
            self._style_data(lc, bold=True, align='left')
            self._style_data(vc)
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 28

    def _fill_daily_sheet(self, ws, gen_id, gen_name, days, end_dt, start_dt):
        ws['A1'] = f"Щоденна статистика — «{gen_name}»"
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 24

        headers = [
            'Дата',
            'Мотогодини',
            'Витрата, л',
            'Сер. витрата, л/год',
            'Залишок ранок, л',
            'Залишок вечір, л',
            'Заправка, л',
            'Зміни',
        ]
        for ci, h in enumerate(headers, start=1):
            self._style_header(ws.cell(row=2, column=ci, value=h))
        ws.row_dimensions[2].height = 36

        col_widths = [14, 14, 13, 18, 18, 18, 13, 20]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        daily = self._build_daily_stats(start_dt, end_dt, gen_id)
        for ri, d in enumerate(daily, start=3):
            rate = round(d['fuel_consumed'] / d['work_hours'], 2) if d['work_hours'] > 0 else ''
            shifts_str = ', '.join(d.get('shifts_active', []))
            row_vals = [
                d['date'],
                d['work_hours'],
                d['fuel_consumed'],
                rate,
                d.get('morning_fuel', ''),
                d.get('evening_fuel', ''),
                d.get('refill_total', 0) or '',
                shifts_str,
            ]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                self._style_data(c, bold=(ci == 1), align='left' if ci == 8 else 'center')
            # Highlight low evening fuel
            eve_val = d.get('evening_fuel')
            if isinstance(eve_val, (int, float)):
                eve_cell = ws.cell(row=ri, column=6)
                if eve_val < 15:
                    eve_cell.fill = self._fill('red')
                elif eve_val < 40:
                    eve_cell.fill = self._fill('orange')
                elif eve_val > 60:
                    eve_cell.fill = self._fill('light_green')

    def _fill_maintenance_sheet(self, ws, gen_id, gen_name):
        ws['A1'] = f"Технічне обслуговування — «{gen_name}»"
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')

        stats = db.get_maintenance_stats(gen_id)
        ws['A3'] = 'Мотогодини (загалом):'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f"{float(stats.get('total_hours', 0)):.1f} год"

        oil_interval = getattr(config, 'OIL_CHANGE_INTERVAL', 250)
        spark_interval = getattr(config, 'SPARK_CHANGE_INTERVAL', 500)
        oil_h = float(stats.get('last_oil_change', 0) or 0)
        spark_h = float(stats.get('last_spark_change', 0) or 0)
        total_h = float(stats.get('total_hours', 0) or 0)

        ws['A4'] = 'До заміни мастила:'
        ws['A4'].font = Font(bold=True)
        ws['B4'] = f"{max(0, oil_interval - (total_h - oil_h)):.0f} год"
        ws['A5'] = 'До заміни свічок:'
        ws['A5'].font = Font(bold=True)
        ws['B5'] = f"{max(0, spark_interval - (total_h - spark_h)):.0f} год"

        mnt_col_hdrs = ['Дата', 'Тип ТО', 'Мотогодини', 'Виконав', 'Примітки']
        for ci, h in enumerate(mnt_col_hdrs, start=1):
            self._style_header(ws.cell(row=7, column=ci, value=h))

        col_widths_mnt = [14, 22, 18, 20, 20]
        for ci, w in enumerate(col_widths_mnt, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        mnt_map = {'oil': 'Заміна мастила', 'spark': 'Заміна свічок', 'maintenance': 'Планове ТО'}
        mnt_history = db.get_maintenance_history(gen_id, 100)
        for ri, rec in enumerate(mnt_history, start=8):
            rec_id, date_s, action, hours, admin_name, *_ = rec
            ws.cell(row=ri, column=1, value=date_s)
            ws.cell(row=ri, column=2, value=mnt_map.get(action, action))
            ws.cell(row=ri, column=3, value=f"{float(hours):.1f} год")
            ws.cell(row=ri, column=4, value=admin_name or '—')
            for ci in range(1, 5):
                self._style_data(ws.cell(row=ri, column=ci))

    # ------------------------------------------------------------------
    # Technical report
    # ------------------------------------------------------------------

    def _generate_technical_report(self, days: int, gen_id: str, year: int = None, month: int = None):
        """Technical report focused on generator performance."""
        gen_id = gen_id or db.get_active_generator()
        gen_name = db.get_generator_name(gen_id)
        if year and month:
            start_dt, end_dt = self._month_range(year, month)
            period_label = f"{MONTH_NAMES[month]} {year}"
        else:
            end_dt = datetime.now(config.KYIV)
            start_dt = end_dt - timedelta(days=days)
            period_label = f"за {days} днів"

        ws = self.wb.active
        ws.title = "Технічний звіт"

        ws['A1'] = f"Технічний звіт: «{gen_name}» {period_label}"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['dark_text'])
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 28

        stats = db.get_maintenance_stats(gen_id)
        main_stats = db.get_generator_stats(gen_id)
        daily = self._build_daily_stats(start_dt, end_dt, gen_id)
        total_hours = round(sum(d['work_hours'] for d in daily), 1)
        total_fuel = round(sum(d['fuel_consumed'] for d in daily), 1)
        avg_rate = round(total_fuel / total_hours, 2) if total_hours > 0 else 0.0

        oil_interval = getattr(config, 'OIL_CHANGE_INTERVAL', 250)
        spark_interval = getattr(config, 'SPARK_CHANGE_INTERVAL', 500)
        total_h = float(stats.get('total_hours', 0) or 0)
        oil_h = float(stats.get('last_oil_change', 0) or 0)
        spark_h = float(stats.get('last_spark_change', 0) or 0)
        oil_remaining = max(0, oil_interval - (total_h - oil_h))
        spark_remaining = max(0, spark_interval - (total_h - spark_h))

        tech_data = [
            ('Загальні мотогодини (всього)', f"{float(main_stats.get('total_hours', 0)):.1f} год"),
            ('Мотогодини за звітний період', f"{total_hours:.1f} год"),
            ('Загальна витрата палива', f"{total_fuel:.1f} л"),
            ('Середня витрата, л/год', f"{avg_rate:.2f}"),
            ('До заміни мастила', f"{oil_remaining:.0f} год"),
            ('До заміни свічок', f"{spark_remaining:.0f} год"),
        ]

        ws['A3'] = 'Технічні показники'
        ws['A3'].font = Font(bold=True, size=12)
        for ri, (label, value) in enumerate(tech_data, start=4):
            lc = ws.cell(row=ri, column=1, value=label)
            vc = ws.cell(row=ri, column=2, value=value)
            lc.font = Font(bold=True)
            lc.fill = self._fill('light_blue')
            self._style_data(lc, bold=True, align='left')
            self._style_data(vc)
            # Highlight if maintenance is due
            if 'мастила' in label and oil_remaining < oil_interval * 0.1:
                vc.fill = self._fill('red')
            elif 'свічок' in label and spark_remaining < spark_interval * 0.1:
                vc.fill = self._fill('red')

        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 20

        # Maintenance history sheet
        ws_maint = self.wb.create_sheet("ТО Історія")
        self._fill_maintenance_sheet(ws_maint, gen_id, gen_name)

    # ------------------------------------------------------------------
    # Financial report
    # ------------------------------------------------------------------

    def _generate_financial_report(self, days: int, gen_id: str, year: int = None, month: int = None):
        """Financial report with cost analysis."""
        gen_id = gen_id or db.get_active_generator()
        gen_name = db.get_generator_name(gen_id)
        if year and month:
            start_dt, end_dt = self._month_range(year, month)
            period_label = f"{MONTH_NAMES[month]} {year}"
        else:
            end_dt = datetime.now(config.KYIV)
            start_dt = end_dt - timedelta(days=days)
            period_label = f"за {days} днів"

        ws = self.wb.active
        ws.title = "Фінансовий звіт"

        ws['A1'] = f"Фінансовий звіт: «{gen_name}» {period_label}"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['dark_text'])
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 28

        daily = self._build_daily_stats(start_dt, end_dt, gen_id)
        total_hours = round(sum(d['work_hours'] for d in daily), 1)
        total_fuel = round(sum(d['fuel_consumed'] for d in daily), 1)
        avg_rate = round(total_fuel / total_hours, 2) if total_hours > 0 else 0.0
        fuel_price = getattr(config, 'FUEL_PRICE', 50.0)
        fuel_cost = round(total_fuel * fuel_price, 0)
        cost_per_hour = round(fuel_cost / total_hours, 2) if total_hours > 0 else 0.0

        fin_data = [
            ('Ціна палива, грн/л', fuel_price),
            ('Загальна витрата палива, л', total_fuel),
            ('Загальна вартість палива, грн', fuel_cost),
            ('Мотогодин', total_hours),
            ('Вартість 1 мотогодини, грн', cost_per_hour),
            ('Середня витрата, л/год', avg_rate),
        ]

        ws['A3'] = 'Фінансові показники'
        ws['A3'].font = Font(bold=True, size=12)
        for ri, (label, value) in enumerate(fin_data, start=4):
            lc = ws.cell(row=ri, column=1, value=label)
            vc = ws.cell(row=ri, column=2, value=value)
            lc.font = Font(bold=True)
            lc.fill = self._fill('light_blue')
            self._style_data(lc, bold=True, align='left')
            self._style_data(vc)
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 20

        # Daily cost breakdown
        ws['A11'] = 'Щоденні витрати'
        ws['A11'].font = Font(bold=True, size=12)
        fin_headers = ['Дата', 'Мотогодини', 'Витрата, л', 'Вартість, грн']
        for ci, h in enumerate(fin_headers, start=1):
            self._style_header(ws.cell(row=12, column=ci, value=h))
        for ri, d in enumerate(daily, start=13):
            day_cost = round(d['fuel_consumed'] * fuel_price, 0)
            row_vals = [d['date'], d['work_hours'], d['fuel_consumed'], day_cost]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                self._style_data(c, bold=(ci == 1))
        for ci, w in enumerate([14, 14, 13, 16], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ------------------------------------------------------------------
    # Personnel report
    # ------------------------------------------------------------------

    def _generate_personnel_report(self, days: int, gen_id: str, year: int = None, month: int = None):
        """Personnel report with shift statistics."""
        gen_id = gen_id or db.get_active_generator()
        gen_name = db.get_generator_name(gen_id)
        if year and month:
            start_dt, end_dt = self._month_range(year, month)
            period_label = f"{MONTH_NAMES[month]} {year}"
        else:
            end_dt = datetime.now(config.KYIV)
            start_dt = end_dt - timedelta(days=days)
            period_label = f"за {days} днів"

        ws = self.wb.active
        ws.title = "По персоналу"

        ws['A1'] = f"Звіт по персоналу: «{gen_name}» {period_label}"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['dark_text'])
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = self._fill('header_bg')
        ws.row_dimensions[1].height = 28

        start_date = start_dt.strftime('%Y-%m-%d')
        end_date = end_dt.strftime('%Y-%m-%d')
        logs = db.get_logs_for_period(start_date, end_date, gen_id)

        pname_map: dict = {}
        for row in logs:
            event_type, ts_str, user_name = row[0], row[1], row[2]
            if event_type in ('m_start', 'd_start', 'e_start', 'x_start'):
                if user_name not in pname_map:
                    pname_map[user_name] = {'name': user_name, 'shifts': 0}
                pname_map[user_name]['shifts'] += 1

        personnel_headers = ['Ім\'я', 'Кількість змін']
        ws['A3'] = 'Статистика по персоналу'
        ws['A3'].font = Font(bold=True, size=12)
        for ci, h in enumerate(personnel_headers, start=1):
            self._style_header(ws.cell(row=4, column=ci, value=h))
        for ri, p in enumerate(pname_map.values(), start=5):
            ws.cell(row=ri, column=1, value=p['name'])
            ws.cell(row=ri, column=2, value=p['shifts'])
            for ci in range(1, 3):
                self._style_data(ws.cell(row=ri, column=ci))
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    # ------------------------------------------------------------------
    # Helper: build daily statistics
    # ------------------------------------------------------------------

    def _build_daily_stats(self, start_dt: datetime, end_dt: datetime, gen_id: str) -> list:
        """Build daily statistics list from database logs."""
        start_date = start_dt.strftime('%Y-%m-%d')
        end_date = end_dt.strftime('%Y-%m-%d')
        logs = db.get_logs_for_period(start_date, end_date, gen_id)
        fuel_rate = db.get_fuel_consumption_rate()

        days_data = defaultdict(
            lambda: {
                'shifts': {'m': {}, 'd': {}, 'e': {}, 'x': {}},
                'refills': [],
                'morning_fuel': None,
                'evening_fuel': None,
            }
        )

        for row_data in logs:
            event_type, ts_str, user_name, value, driver_name, receipt_number, *_ = row_data
            if not ts_str:
                continue
            try:
                ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                continue
            date_str = ts.strftime('%Y-%m-%d')
            day = days_data[date_str]

            if event_type.endswith('_start'):
                shift = event_type.split('_')[0]
                if shift in day['shifts']:
                    day['shifts'][shift]['start'] = ts.strftime('%H:%M')
            elif event_type.endswith('_end'):
                shift = event_type.split('_')[0]
                if shift in day['shifts']:
                    day['shifts'][shift]['end'] = ts.strftime('%H:%M')
            elif event_type == 'refill':
                try:
                    liters = float(value or 0)
                except Exception:
                    liters = 0.0
                day['refills'].append(liters)
            elif event_type == 'corr_fuel_set':
                try:
                    day['evening_fuel'] = float(value or 0)
                except Exception:
                    pass

        # Forward calculation: propagate fuel balances day by day,
        # using corr_fuel_set events as anchor points when available.
        prev_evening_fuel = None

        result = []
        for date_str in sorted(days_data.keys()):
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                date_fmt = dt.strftime('%d.%m.%Y')
            except Exception:
                date_fmt = date_str

            day = days_data[date_str]
            total_shift_mins = 0
            shifts_active = []
            for sname, shift_data in day['shifts'].items():
                s_str = shift_data.get('start')
                e_str = shift_data.get('end')
                if s_str and e_str:
                    try:
                        s_t = datetime.strptime(s_str, '%H:%M')
                        e_t = datetime.strptime(e_str, '%H:%M')
                        diff = (e_t - s_t).total_seconds() / 60
                        if diff < 0:
                            diff += 24 * 60
                        total_shift_mins += diff
                        shifts_active.append(sname)
                    except Exception:
                        pass

            work_hours = round(total_shift_mins / 60, 2)
            fuel_consumed = round(work_hours * fuel_rate, 1) if work_hours > 0 else 0.0
            refill_total = round(sum(day['refills']), 1) if day['refills'] else 0.0

            if day['evening_fuel'] is not None:
                # Day has a manual fuel correction — use it as evening fuel and
                # calculate morning fuel backwards from it.
                evening_fuel = day['evening_fuel']
                morning_fuel = round(evening_fuel - refill_total + fuel_consumed, 1)
                if morning_fuel < 0:
                    morning_fuel = 0.0
            elif prev_evening_fuel is not None:
                # Propagate forward: previous day's evening fuel is today's morning fuel.
                morning_fuel = prev_evening_fuel
                evening_fuel = round(morning_fuel + refill_total - fuel_consumed, 1)
                if evening_fuel < 0:
                    evening_fuel = 0.0
            elif refill_total > 0:
                # Refill logged but no prior state known — assume tank was empty before refill.
                morning_fuel = 0.0
                evening_fuel = round(refill_total - fuel_consumed, 1)
                if evening_fuel < 0:
                    evening_fuel = 0.0
            else:
                morning_fuel = None
                evening_fuel = None

            if evening_fuel is not None:
                prev_evening_fuel = evening_fuel

            result.append(
                {
                    'date': date_fmt,
                    'work_hours': work_hours,
                    'fuel_consumed': fuel_consumed,
                    'morning_fuel': morning_fuel,
                    'evening_fuel': evening_fuel,
                    'morning_balance': morning_fuel,
                    'evening_balance': evening_fuel,
                    'refill_total': refill_total if refill_total > 0 else None,
                    # outage_hours = work_hours because the generator runs during power outages
                    'outage_hours': work_hours,
                    'shifts_active': shifts_active,
                }
            )

        return result


def generate_excel_report(
    report_type: str,
    days: int,
    generator_id: str = None,
    year: int = None,
    month: int = None,
) -> bytes:
    """Main entry point for Excel report generation."""
    generator = ExcelReportGenerator()
    return generator.generate_report(report_type, days, generator_id, year=year, month=month)
